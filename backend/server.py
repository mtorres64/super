from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Query, Request, Body
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import re
import json
import logging
from pathlib import Path
from pydantic import BaseModel, Field, root_validator
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import calendar
import jwt
from passlib.hash import bcrypt
from enum import Enum
import base64
import pandas as pd
import mercadopago
import httpx
import asyncio
import random
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from afip import AfipService, encrypt_private_key, decrypt_private_key, extract_p12

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url, tz_aware=True)
db = client[os.environ['DB_NAME']]

# JWT settings
SECRET_KEY = os.environ.get('JWT_SECRET', 'your-secret-key-here')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480  # 8 horas — duración de un turno de trabajo

# MercadoPago settings
MP_ACCESS_TOKEN = os.environ.get('MP_ACCESS_TOKEN', '')
MP_TEST_PAYER_EMAIL = os.environ.get('MP_TEST_PAYER_EMAIL', '')
APP_URL = os.environ.get('APP_URL', 'http://localhost:8000')
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
SUSCRIPCION_PRECIO = float(os.environ.get('SUSCRIPCION_PRECIO', '50000'))
SUSCRIPCION_PLAN_NOMBRE = os.environ.get('SUSCRIPCION_PLAN_NOMBRE', 'Plan Mensual')

# SMTP settings
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
EMAIL_FROM = os.environ.get('EMAIL_FROM', 'PULS <noreply@example.com>')

async def get_precio_suscripcion() -> float:
    doc = await db.system_config.find_one({"key": "suscripcion_precio"})
    return float(doc["value"]) if doc else SUSCRIPCION_PRECIO

async def get_plan_nombre_suscripcion() -> str:
    doc = await db.system_config.find_one({"key": "suscripcion_plan_nombre"})
    return doc["value"] if doc else SUSCRIPCION_PLAN_NOMBRE

async def get_whatsapp_numero() -> str:
    doc = await db.system_config.find_one({"key": "whatsapp_numero"})
    return doc["value"] if doc else "+5493815156095"

async def get_trial_dias() -> int:
    doc = await db.system_config.find_one({"key": "trial_dias"})
    return int(doc["value"]) if doc else 15

async def get_grace_days() -> int:
    doc = await db.system_config.find_one({"key": "grace_days"})
    return int(doc["value"]) if doc else 15

async def get_dias_alerta() -> list:
    doc = await db.system_config.find_one({"key": "dias_alerta"})
    return doc["value"] if doc else [10, 5]

async def get_descuento_anual_pct() -> int:
    doc = await db.system_config.find_one({"key": "descuento_anual_pct"})
    return int(doc["value"]) if doc else 20

async def get_precio_emprendedor() -> float:
    doc = await db.system_config.find_one({"key": "precio_emprendedor"})
    return float(doc["value"]) if doc else 30000.0

async def get_precio_profesional() -> float:
    doc = await db.system_config.find_one({"key": "precio_profesional"})
    return float(doc["value"]) if doc else 45000.0

async def get_precio_empresarial() -> float:
    doc = await db.system_config.find_one({"key": "precio_empresarial"})
    return float(doc["value"]) if doc else 60000.0

async def get_precio_por_tier(tier: str) -> float:
    if tier == "emprendedor":
        return await get_precio_emprendedor()
    elif tier == "empresarial":
        return await get_precio_empresarial()
    return await get_precio_profesional()

async def get_precio_tienda() -> float:
    doc = await db.system_config.find_one({"key": "precio_tienda"})
    return float(doc["value"]) if doc else 0.0

async def get_precio_sucursal_extra() -> float:
    doc = await db.system_config.find_one({"key": "precio_sucursal_extra"})
    return float(doc["value"]) if doc else 0.0

async def get_precio_pack_usuarios() -> float:
    doc = await db.system_config.find_one({"key": "precio_pack_usuarios"})
    return float(doc["value"]) if doc else 0.0

async def get_saas_nombre() -> str:
    doc = await db.system_config.find_one({"key": "saas_nombre"})
    return doc["value"] if doc else "PULS"

async def get_statement_descriptor() -> str:
    doc = await db.system_config.find_one({"key": "statement_descriptor"})
    return doc["value"] if doc else "SuperMarket POS"

async def get_smtp_config() -> dict:
    doc = await db.system_config.find_one({"key": "smtp_config"})
    if doc and isinstance(doc.get("value"), dict):
        return doc["value"]
    return {
        "host": SMTP_HOST,
        "port": SMTP_PORT,
        "user": SMTP_USER,
        "password": SMTP_PASSWORD,
        "from_address": EMAIL_FROM,
    }

def calcular_siguiente_vencimiento(dia_facturacion: int, meses: int, desde: datetime) -> datetime:
    """Mismo día del mes, n meses hacia adelante. Clamped al último día del mes si es necesario."""
    mes = desde.month + meses
    anio = desde.year + (mes - 1) // 12
    mes = ((mes - 1) % 12) + 1
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    dia = min(dia_facturacion, ultimo_dia)
    tz = desde.tzinfo or timezone.utc
    return datetime(anio, mes, dia, 0, 0, 0, tzinfo=tz)

def _aplicar_renovacion(suscripcion: dict, meses: int, now: datetime):
    """Calcula (base, nueva_fecha) respetando día de facturación fijo. No acumulativo.
    Durante el periodo de gracia (suscripción paga vencida hace <= GRACE_DAYS días),
    usa la fecha de vencimiento como base para respetar el ciclo de facturación.
    Ej: venció el 1, paga el día 10 (gracia) → renueva hasta el 1 del mes siguiente."""
    vencimiento = suscripcion["fecha_vencimiento"]
    if isinstance(vencimiento, str):
        vencimiento = datetime.fromisoformat(vencimiento)
    if not vencimiento.tzinfo:
        vencimiento = vencimiento.replace(tzinfo=timezone.utc)
    fue_pagada = suscripcion.get("fue_pagada", False)
    # Si fue paga y venció dentro del periodo de gracia, respetar el ciclo de facturación
    if fue_pagada and vencimiento < now and (now - vencimiento).days <= GRACE_DAYS:
        base = vencimiento
    else:
        base = vencimiento if vencimiento > now else now
    dia = suscripcion.get("dia_facturacion") or min(base.day, 28)
    nueva_fecha = calcular_siguiente_vencimiento(dia, meses, base)
    return base, nueva_fecha

DIAS_ALERTA = [10, 5]
GRACE_DAYS = 15  # Días de gracia para suscripciones pagas vencidas

async def generar_alertas_vencimiento() -> int:
    """Genera notificaciones para suscripciones próximas a vencer. Retorna cantidad generadas."""
    now = datetime.now(timezone.utc)
    dias_alerta = await get_dias_alerta()
    suscripciones = await db.suscripciones.find(
        {"status": {"$in": [SuscripcionStatus.TRIAL, SuscripcionStatus.ACTIVA]}}
    ).to_list(None)
    generadas = 0
    for sus in suscripciones:
        vencimiento = sus.get("fecha_vencimiento")
        if isinstance(vencimiento, str):
            vencimiento = datetime.fromisoformat(vencimiento)
        if vencimiento and not vencimiento.tzinfo:
            vencimiento = vencimiento.replace(tzinfo=timezone.utc)
        if not vencimiento:
            continue
        dias_restantes = (vencimiento - now).days
        empresa_id = sus["empresa_id"]
        periodo_ref = vencimiento.date().isoformat()
        plan_nombre = sus.get("plan_nombre", "Plan")
        status_label = "trial" if sus.get("status") == SuscripcionStatus.TRIAL else "suscripción"
        # Usar solo el threshold más urgente que aplique
        applicable = sorted([t for t in dias_alerta if dias_restantes <= t])
        if not applicable:
            continue
        threshold = applicable[0]  # el más chico = más urgente
        tipo = f"plan_por_vencer_{threshold}"
        existing = await db.notificaciones.find_one({
            "empresa_id": empresa_id,
            "tipo": tipo,
            "periodo_ref": periodo_ref,
        })
        if not existing:
            dias_str = f"Quedan {max(0, dias_restantes)} días" if dias_restantes >= 0 else "Ya venció"
            notif = Notificacion(
                empresa_id=empresa_id,
                tipo=tipo,
                titulo=f"Tu {status_label} vence pronto",
                mensaje=(
                    f"Tu {plan_nombre} vence el {vencimiento.strftime('%d/%m/%Y')}. "
                    f"{dias_str}. Renovalo para continuar usando el sistema."
                ),
                dias_restantes=max(0, dias_restantes),
                periodo_ref=periodo_ref,
            )
            await db.notificaciones.insert_one(notif.dict())
            generadas += 1
    return generadas

# Owner (system admin) credentials
OWNER_USERNAME = os.environ.get('OWNER_USERNAME', 'owner')
OWNER_PASSWORD = os.environ.get('OWNER_PASSWORD', 'owner1234')

# Security
security = HTTPBearer()

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Enums
class UserRole(str, Enum):
    ADMIN = "admin"
    CAJERO = "cajero"
    SUPERVISOR = "supervisor"

class PaymentMethod(str, Enum):
    EFECTIVO = "efectivo"
    TARJETA = "tarjeta"
    TRANSFERENCIA = "transferencia"

class ProductType(str, Enum):
    CODIGO_BARRAS = "codigo_barras"
    POR_PESO = "por_peso"

class ProductKind(str, Enum):
    NORMAL = "normal"
    COMBO = "combo"

class CashSessionStatus(str, Enum):
    ABIERTA = "abierta"
    CERRADA = "cerrada"

class MovementType(str, Enum):
    APERTURA = "apertura"
    VENTA = "venta"
    RETIRO = "retiro"
    CIERRE = "cierre"
    DEVOLUCION = "devolucion"

class SuscripcionStatus(str, Enum):
    TRIAL = "trial"
    ACTIVA = "activa"
    VENCIDA = "vencida"
    SUSPENDIDA = "suspendida"

# --- Email helpers ---
TD_DIGITO = (
    '<td width="10"></td>'
    '<td style="width:56px;height:64px;background:#ecfdf5;border:2px solid #10b981;'
    'border-radius:12px;text-align:center;vertical-align:middle;'
    'font-size:32px;font-weight:800;color:#065f46;">{d}</td>'
)

def _digitos_html(codigo: str) -> str:
    partes = []
    for d in codigo:
        partes.append(TD_DIGITO.format(d=d))
    return "".join(partes)

def _anio() -> str:
    return str(datetime.now().year)

def _build_email_html(titulo: str, subtitulo: str, codigo: str, nota_pie: str, saas_nombre: str = "PULS") -> str:
    return (
        '<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1"></head>'
        '<body style="margin:0;padding:0;background:#f3f4f6;font-family:\'Segoe UI\',Arial,sans-serif;">'
        '<table width="100%" cellpadding="0" cellspacing="0" style="background:#f3f4f6;padding:40px 0;">'
        '<tr><td align="center">'
        '<table width="520" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:16px;'
        'overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">'
        # Header
        '<tr><td style="background:linear-gradient(135deg,#059669,#10b981);padding:36px 40px;text-align:center;">'
        '<div style="font-size:28px;font-weight:800;color:#ffffff;letter-spacing:-0.5px;">' + saas_nombre + '</div>'
        '<div style="font-size:13px;color:rgba(255,255,255,0.8);margin-top:4px;">Sistema de gesti\u00f3n para comercios</div>'
        '</td></tr>'
        # Cuerpo
        '<tr><td style="padding:40px 40px 32px;">'
        '<p style="margin:0 0 8px;font-size:22px;font-weight:700;color:#111827;">' + titulo + '</p>'
        '<p style="margin:0 0 28px;font-size:15px;color:#6b7280;line-height:1.6;">' + subtitulo + '</p>'
        '<table cellpadding="0" cellspacing="0" style="margin:0 auto 28px;"><tr>'
        + _digitos_html(codigo) +
        '</tr></table>'
        '<p style="margin:0;font-size:13px;color:#9ca3af;text-align:center;">' + nota_pie + '</p>'
        '</td></tr>'
        # Footer
        '<tr><td style="background:#f9fafb;padding:20px 40px;border-top:1px solid #e5e7eb;text-align:center;">'
        '<p style="margin:0;font-size:12px;color:#9ca3af;">'
        '\u00a9 ' + _anio() + ' ' + saas_nombre + ' \u00b7 Sistema de gesti\u00f3n para comercios'
        '</p></td></tr>'
        '</table></td></tr></table></body></html>'
    )

def _smtp_send(destinatario: str, subject: str, html: str, smtp_cfg: dict = None):
    cfg = smtp_cfg or {"host": SMTP_HOST, "port": SMTP_PORT, "user": SMTP_USER, "password": SMTP_PASSWORD, "from_address": EMAIL_FROM}
    host = cfg.get("host", SMTP_HOST)
    port = int(cfg.get("port", SMTP_PORT))
    user = cfg.get("user", SMTP_USER)
    password = cfg.get("password", SMTP_PASSWORD)
    from_address = cfg.get("from_address", EMAIL_FROM)
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = from_address
    msg['To'] = destinatario
    msg.attach(MIMEText(html, 'html', 'utf-8'))
    if port == 465:
        with smtplib.SMTP_SSL(host, port) as server:
            server.login(user, password)
            server.sendmail(user, [destinatario], msg.as_string())
    else:
        with smtplib.SMTP(host, port) as server:
            server.ehlo()
            server.starttls()
            server.login(user, password)
            server.sendmail(user, [destinatario], msg.as_string())

def _send_email_otp_sync(destinatario: str, codigo: str, saas_nombre: str = "PULS", smtp_cfg: dict = None):
    html = _build_email_html(
        titulo="Verific\u00e1 tu correo",
        subtitulo=f"Us\u00e1 el siguiente c\u00f3digo para completar el registro de tu empresa en {saas_nombre}.<br>"
                  "V\u00e1lido por <strong style=\"color:#111827;\">10 minutos</strong>.",
        codigo=codigo,
        nota_pie="Si no solicitaste este c\u00f3digo, pod\u00e9s ignorar este mensaje.",
        saas_nombre=saas_nombre,
    )
    _smtp_send(destinatario, f"Tu código de verificación {saas_nombre}: {codigo}", html, smtp_cfg=smtp_cfg)

def _send_password_reset_sync(destinatario: str, codigo: str, saas_nombre: str = "PULS", smtp_cfg: dict = None):
    html = _build_email_html(
        titulo="Recuperar contrase\u00f1a",
        subtitulo="Recibimos una solicitud para restablecer la contrase\u00f1a de tu cuenta.<br>"
                  "Us\u00e1 este c\u00f3digo. V\u00e1lido por <strong style=\"color:#111827;\">15 minutos</strong>.",
        codigo=codigo,
        nota_pie="Si no solicitaste este cambio, pod\u00e9s ignorar este mensaje. Tu contrase\u00f1a no se modificar\u00e1.",
        saas_nombre=saas_nombre,
    )
    _smtp_send(destinatario, f"Recuperar contrase\u00f1a {saas_nombre}: {codigo}", html, smtp_cfg=smtp_cfg)

async def send_email_otp(destinatario: str, codigo: str):
    saas_nombre = await get_saas_nombre()
    smtp_cfg = await get_smtp_config()
    await asyncio.to_thread(_send_email_otp_sync, destinatario, codigo, saas_nombre, smtp_cfg)

async def send_password_reset_email(destinatario: str, codigo: str):
    saas_nombre = await get_saas_nombre()
    smtp_cfg = await get_smtp_config()
    await asyncio.to_thread(_send_password_reset_sync, destinatario, codigo, saas_nombre, smtp_cfg)

def _send_solicitud_notify_sync(empresa_nombre: str, admin_nombre: str, admin_email: str, telefono: str, smtp_cfg: dict = None):
    html = (
        '<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;background:#f9fafb;padding:32px;">'
        '<div style="max-width:480px;margin:0 auto;background:white;border-radius:12px;padding:32px;border:1px solid #e5e7eb;">'
        '<h2 style="color:#059669;margin-top:0;">Nueva solicitud de cuenta - PULS</h2>'
        f'<p style="margin:8px 0;"><strong>Empresa:</strong> {empresa_nombre}</p>'
        f'<p style="margin:8px 0;"><strong>Nombre:</strong> {admin_nombre}</p>'
        f'<p style="margin:8px 0;"><strong>Email:</strong> {admin_email}</p>'
        f'<p style="margin:8px 0;"><strong>Teléfono:</strong> {telefono}</p>'
        '<hr style="border:none;border-top:1px solid #e5e7eb;margin:24px 0;">'
        '<p style="color:#6b7280;font-size:12px;margin:0;">Contactar al cliente y crear su cuenta manualmente.</p>'
        '</div></body></html>'
    )
    _smtp_send("ing.torresma@gmail.com", f"Nueva solicitud PULS: {empresa_nombre}", html, smtp_cfg=smtp_cfg)

async def send_solicitud_notify(empresa_nombre: str, admin_nombre: str, admin_email: str, telefono: str):
    smtp_cfg = await get_smtp_config()
    await asyncio.to_thread(_send_solicitud_notify_sync, empresa_nombre, admin_nombre, admin_email, telefono, smtp_cfg)


class OTPEnviar(BaseModel):
    email: str

class OTPVerificar(BaseModel):
    email: str
    codigo: str

# ─── Módulos del sistema ──────────────────────────────────────────────────────

MODULES: dict[str, str] = {
    "pos":           "POS / Ventas",
    "caja":          "Caja",
    "inventario":    "Inventario",
    "clientes":      "Clientes",
    "facturacion":   "Facturación Electrónica",
    "reportes":      "Reportes de Ventas",
    "compras":       "Compras y Proveedores",
    "alertas_stock": "Alertas de Stock",
    "usuarios":      "Usuarios y Roles",
    "multi_sucursal":"Multi-sucursal",
    "configuracion": "Configuración",
    "notificaciones":"Notificaciones",
}

PLAN_MODULES: dict[str, list[str]] = {
    "emprendedor": [
        "pos", "caja", "inventario", "notificaciones",
    ],
    "profesional": [
        "pos", "caja", "inventario", "clientes", "facturacion", "reportes", "compras",
        "alertas_stock", "usuarios", "configuracion", "notificaciones",
    ],
    "empresarial": [
        "pos", "caja", "inventario", "clientes", "facturacion", "reportes", "compras",
        "alertas_stock", "usuarios", "multi_sucursal", "configuracion", "notificaciones",
    ],
}

def calcular_modules_activos(plan_tipo: str, modules_extra: list, modules_removidos: list) -> list[str]:
    base = list(PLAN_MODULES.get(plan_tipo, PLAN_MODULES["profesional"]))
    activos = [m for m in base if m not in modules_removidos]
    for m in modules_extra:
        if m not in activos and m in MODULES:
            activos.append(m)
    return activos


# --- Empresa models ---
class Empresa(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nombre: str
    email_verificado: bool = False
    activo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmpresaRegister(BaseModel):
    empresa_nombre: str
    admin_nombre: str
    admin_email: str
    admin_password: str
    otp_token: str

class EmpresaSolicitud(BaseModel):
    empresa_nombre: str
    admin_nombre: str
    admin_email: str
    telefono: str

class OwnerCreateCliente(BaseModel):
    empresa_nombre: str
    admin_nombre: str
    admin_email: str
    admin_password: str

# --- Subscription models ---
class Suscripcion(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    plan_nombre: str = "Plan Mensual"
    precio: float = 0.0
    moneda: str = "ARS"
    status: SuscripcionStatus = SuscripcionStatus.TRIAL
    fecha_inicio: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    fecha_vencimiento: datetime
    dia_facturacion: Optional[int] = None
    plan_tipo: str = "mensual"               # periodo de facturación: "mensual" | "anual"
    plan_tier: str = "empresarial"           # nivel de plan: "emprendedor" | "profesional" | "empresarial"
    fue_pagada: bool = False
    tipo_cobro: str = "manual"               # "manual" | "automatico"
    mp_preapproval_id: Optional[str] = None  # ID del preapproval en MP (débito automático)
    modules_extra: List[str] = Field(default_factory=list)     # módulos añadidos por owner
    modules_removidos: List[str] = Field(default_factory=list) # módulos quitados por owner
    addon_tienda: bool = False
    sucursales_extra: int = 0       # sucursales adicionales sobre las 3 base del plan empresarial
    usuarios_extra_packs: int = 0   # packs de 5 usuarios adicionales sobre los 15 base
    descuento_pct: int = 0          # porcentaje de descuento personalizado asignado por owner (0-100)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PagoSuscripcion(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    monto: float
    moneda: str = "ARS"
    estado: str  # pending, approved, rejected, cancelled
    concepto: str
    mp_payment_id: Optional[str] = None
    mp_preference_id: Optional[str] = None
    mp_preapproval_id: Optional[str] = None  # Vinculado si el pago es de débito automático
    origen: str = "manual"                   # "manual" | "preapproval"
    plan_tipo: str = "mensual"
    plan_tier: Optional[str] = None          # "emprendedor" | "profesional" | "empresarial"
    addon_tienda: bool = False
    sucursales_extra: int = 0
    usuarios_extra_packs: int = 0
    fecha: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    periodo_inicio: Optional[datetime] = None
    periodo_fin: Optional[datetime] = None

class Notificacion(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    tipo: str  # "plan_por_vencer_10", "plan_por_vencer_5"
    titulo: str
    mensaje: str
    leida: bool = False
    dias_restantes: Optional[int] = None
    periodo_ref: str = ""  # fecha_vencimiento ISO date para deduplicación
    fecha: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Models
class Branch(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nombre: str
    direccion: str
    telefono: Optional[str] = None
    activo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BranchCreate(BaseModel):
    nombre: str
    direccion: str
    telefono: Optional[str] = None
    margen_ajuste: Optional[float] = None  # % de ajuste sobre precios de referencia al crear

class BranchUpdate(BaseModel):
    nombre: Optional[str] = None
    direccion: Optional[str] = None
    telefono: Optional[str] = None
    activo: Optional[bool] = None

class CashSession(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    branch_id: str
    user_id: str
    monto_inicial: float
    monto_ventas: float = 0.0
    monto_retiros: float = 0.0
    monto_final: Optional[float] = None
    monto_esperado: Optional[float] = None
    diferencia: Optional[float] = None
    status: CashSessionStatus = CashSessionStatus.ABIERTA
    fecha_apertura: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    fecha_cierre: Optional[datetime] = None
    observaciones: Optional[str] = None

class CashSessionCreate(BaseModel):
    monto_inicial: float
    observaciones: Optional[str] = None
    branch_id: Optional[str] = None

class CashSessionClose(BaseModel):
    monto_final: float
    observaciones: Optional[str] = None

class CashSessionHistory(BaseModel):
    id: str
    empresa_id: str
    branch_id: str
    branch_nombre: str
    user_id: str
    user_nombre: str
    monto_inicial: float
    monto_ventas: float = 0
    monto_retiros: float = 0
    monto_final: Optional[float] = None
    diferencia: Optional[float] = None
    status: CashSessionStatus
    fecha_apertura: datetime
    fecha_cierre: Optional[datetime] = None
    observaciones: Optional[str] = None

class CashMovement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    session_id: str
    tipo: MovementType
    monto: float
    descripcion: str
    venta_id: Optional[str] = None
    fecha: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BranchProduct(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    product_id: str
    branch_id: str
    precio: float
    precio_por_peso: Optional[float] = None
    stock: int = 0
    stock_minimo: int = 10
    margen: Optional[float] = None
    costo: Optional[float] = None
    activo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BranchProductCreate(BaseModel):
    product_id: str
    branch_id: Optional[str] = None
    precio: float
    precio_por_peso: Optional[float] = None
    stock: int = 0
    stock_minimo: int = 10
    margen: Optional[float] = None
    costo: Optional[float] = None

class BranchProductUpdate(BaseModel):
    precio: Optional[float] = None
    precio_por_peso: Optional[float] = None
    stock: Optional[int] = None
    stock_minimo: Optional[int] = None
    margen: Optional[float] = None
    costo: Optional[float] = None
    activo: Optional[bool] = None

class BranchProductBulkDeleteRequest(BaseModel):
    ids: Optional[List[str]] = None
    branch_id: Optional[str] = None
    delete_all: bool = False
    search: Optional[str] = None

class Configuration(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    # Company Information
    company_name: str = "SuperMarket POS"
    company_address: str = ""
    company_phone: str = ""
    company_email: str = ""
    company_tax_id: str = ""

    # Tax & Financial Settings
    tax_rate: float = 0.12  # 12% default
    currency_symbol: str = "$"
    currency_code: str = "USD"
    redondeo_precio: int = 100  # 0 = sin redondeo

    # POS Settings
    sounds_enabled: bool = True
    auto_focus_barcode: bool = True
    barcode_scan_timeout: int = 100  # milliseconds
    unified_barcode_search: bool = False
    receipt_footer_text: str = "¡Gracias por su compra!"

    # Inventory Settings
    default_minimum_stock: int = 10
    low_stock_alert_enabled: bool = True
    auto_update_inventory: bool = True
    auto_update_prices: bool = True

    # System Settings
    date_format: str = "DD/MM/YYYY"
    time_format: str = "24h"
    language: str = "es"

    # Receipt Settings
    print_receipt_auto: bool = False
    show_receipt_after_sale: bool = True
    receipt_width: int = 80  # characters
    receipt_format: str = "ticket"  # "ticket" | "a4"

    # Pagination Settings
    items_per_page: int = 10

    # POS Display Settings
    product_view_mode: str = "cards"  # "cards" | "rows"

    # Company Branding
    company_logo: Optional[str] = None  # URL or base64 of logo

    # Color Theme
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    tertiary_color: Optional[str] = None

    # Payment Method Adjustments (percentage: -5 = 5% discount, +3 = 3% surcharge)
    payment_method_adjustments: dict = Field(default_factory=lambda: {
        "efectivo": 0.0,
        "tarjeta": 0.0,
        "transferencia": 0.0
    })

    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ConfigurationUpdate(BaseModel):
    company_name: Optional[str] = None
    company_address: Optional[str] = None
    company_phone: Optional[str] = None
    company_email: Optional[str] = None
    company_tax_id: Optional[str] = None
    tax_rate: Optional[float] = None
    currency_symbol: Optional[str] = None
    currency_code: Optional[str] = None
    redondeo_precio: Optional[int] = None
    sounds_enabled: Optional[bool] = None
    auto_focus_barcode: Optional[bool] = None
    barcode_scan_timeout: Optional[int] = None
    unified_barcode_search: Optional[bool] = None
    receipt_footer_text: Optional[str] = None
    default_minimum_stock: Optional[int] = None
    low_stock_alert_enabled: Optional[bool] = None
    auto_update_inventory: Optional[bool] = None
    auto_update_prices: Optional[bool] = None
    date_format: Optional[str] = None
    time_format: Optional[str] = None
    language: Optional[str] = None
    print_receipt_auto: Optional[bool] = None
    show_receipt_after_sale: Optional[bool] = None
    receipt_width: Optional[int] = None
    receipt_format: Optional[str] = None
    items_per_page: Optional[int] = None
    product_view_mode: Optional[str] = None
    company_logo: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    tertiary_color: Optional[str] = None
    payment_method_adjustments: Optional[dict] = None

class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nombre: str
    email: str
    rol: UserRole
    branch_ids: List[str] = []
    active_branch_id: Optional[str] = None
    activo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

    @root_validator(pre=True)
    def migrate_branch_id(cls, values):
        if not values.get('branch_ids') and values.get('branch_id'):
            values['branch_ids'] = [values['branch_id']]
        return values

    @property
    def branch_id(self) -> Optional[str]:
        active = self.active_branch_id
        if active:
            # Admin puede operar en cualquier sucursal seleccionada, otros solo en las asignadas
            if self.rol == UserRole.ADMIN or active in self.branch_ids:
                return active
        if len(self.branch_ids) == 1:
            return self.branch_ids[0]
        return None

    def dict(self, **kwargs):
        d = super().dict(**kwargs)
        d['branch_id'] = self.branch_id
        return d

class UserCreate(BaseModel):
    nombre: str
    email: str
    password: str
    rol: UserRole
    branch_ids: List[str] = []

class UserUpdate(BaseModel):
    nombre: Optional[str] = None
    email: Optional[str] = None
    rol: Optional[UserRole] = None
    branch_ids: Optional[List[str]] = None
    activo: Optional[bool] = None

class UserLogin(BaseModel):
    email: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nombre: str
    descripcion: Optional[str] = None
    icono: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoryCreate(BaseModel):
    nombre: str
    descripcion: Optional[str] = None
    icono: Optional[str] = None

class ComboItem(BaseModel):
    product_id: str
    cantidad: float = 1.0

class Product(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nombre: str
    codigo_barras: Optional[str] = None
    tipo: ProductType
    kind: ProductKind = ProductKind.NORMAL
    precio: float
    precio_por_peso: Optional[float] = None
    categoria_id: str
    stock: int = 0
    stock_minimo: int = 10
    control_stock: bool = True
    combo_items: List[ComboItem] = []
    activo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductCreate(BaseModel):
    nombre: str
    codigo_barras: Optional[str] = None
    tipo: ProductType
    kind: ProductKind = ProductKind.NORMAL
    precio: float
    precio_por_peso: Optional[float] = None
    categoria_id: str
    stock: int = 0
    stock_minimo: int = 10
    control_stock: bool = True
    combo_items: List[ComboItem] = []
    precio_costo: Optional[float] = None

class ProductUpdate(BaseModel):
    nombre: Optional[str] = None
    codigo_barras: Optional[str] = None
    tipo: Optional[ProductType] = None
    kind: Optional[ProductKind] = None
    precio: Optional[float] = None
    precio_por_peso: Optional[float] = None
    categoria_id: Optional[str] = None
    stock: Optional[int] = None
    stock_minimo: Optional[int] = None
    control_stock: Optional[bool] = None
    combo_items: Optional[List[ComboItem]] = None
    activo: Optional[bool] = None
    precio_costo: Optional[float] = None

class BulkDeleteRequest(BaseModel):
    ids: Optional[List[str]] = None
    delete_all: bool = False
    search: Optional[str] = None

class BulkUpdateControlStockRequest(BaseModel):
    ids: Optional[List[str]] = None
    delete_all: bool = False
    search: Optional[str] = None
    control_stock: bool

class SaleItem(BaseModel):
    producto_id: str
    nombre: Optional[str] = None
    cantidad: float
    precio_unitario: float
    subtotal: Optional[float] = None
    total: Optional[float] = None

    @property
    def total_item(self):
        return self.subtotal or self.total or 0

class Sale(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    cajero_id: str
    branch_id: str
    session_id: Optional[str] = None
    items: List[SaleItem]
    subtotal: float
    impuestos: float = 0.0
    total: float
    metodo_pago: PaymentMethod
    fecha: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    numero_factura: str
    estado: str = "activo"
    # Campos AFIP/ARCA
    cae: Optional[str] = None
    cae_vencimiento: Optional[str] = None       # formato "AAAAMMDD"
    afip_estado: str = "no_configurado"          # no_configurado | autorizado | contingencia | error
    afip_error: Optional[str] = None
    tipo_comprobante: Optional[int] = None       # None=Ticket, 1=Factura A, 6=Factura B, 11=Factura C
    nro_comprobante_afip: Optional[int] = None
    cuit_receptor: Optional[str] = None
    cliente_id: Optional[str] = None
    descuento: float = 0.0
    impuestos_extra_total: float = 0.0
    condicion_iva_receptor: Optional[str] = None
    observaciones_comprobante: Optional[str] = None
    factura_payload: Optional[dict] = None      # Datos mínimos AFIP/ARCA para generación de comprobante

class SaleCreate(BaseModel):
    items: List[SaleItem]
    metodo_pago: PaymentMethod
    tipo_comprobante: Optional[int] = None      # si None → usa default de AfipConfig
    cuit_receptor: Optional[str] = None         # requerido si tipo_comprobante = 1
    cliente_id: Optional[str] = None
    descuento: Optional[float] = 0.0
    impuestos_extra_total: Optional[float] = 0.0
    condicion_iva_receptor: Optional[str] = None
    observaciones_comprobante: Optional[str] = None

# ─── Modelos AFIP/ARCA ────────────────────────────────────────────────────────

class AfipConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    cuit: str
    punto_venta: int = 1
    ambiente: str = "homologacion"              # "homologacion" | "produccion"
    tipo_comprobante_default: int = 6           # 1=A, 6=B, 11=C
    razon_social: str = ""
    # Datos fiscales del emisor (requeridos para generar comprobante)
    domicilio_fiscal: Optional[str] = None      # Dirección fiscal del comercio
    condicion_iva_emisor: str = "RI"            # RI=Resp.Inscripto, MO=Monotributista, EX=Exento
    inicio_actividades: Optional[str] = None    # Formato dd/mm/yyyy
    iibb: Optional[str] = None                  # Nro. Ingresos Brutos o "CM" (Conv. Multilateral)
    concepto_default: int = 1                   # 1=Productos, 2=Servicios, 3=Productos y Servicios
    cert_pem: Optional[str] = None              # Certificado X.509 en PEM codificado en base64
    key_pem_encrypted: Optional[str] = None     # Private key cifrada con Fernet
    activo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AfipConfigCreate(BaseModel):
    cuit: str
    punto_venta: int = 1
    ambiente: str = "homologacion"
    tipo_comprobante_default: int = 6
    razon_social: Optional[str] = ""
    domicilio_fiscal: Optional[str] = None
    condicion_iva_emisor: Optional[str] = "RI"
    inicio_actividades: Optional[str] = None
    iibb: Optional[str] = None
    concepto_default: Optional[int] = 1

class AfipConfigUpdate(BaseModel):
    cuit: Optional[str] = None
    punto_venta: Optional[int] = None
    ambiente: Optional[str] = None
    tipo_comprobante_default: Optional[int] = None
    razon_social: Optional[str] = None
    domicilio_fiscal: Optional[str] = None
    condicion_iva_emisor: Optional[str] = None
    inicio_actividades: Optional[str] = None
    iibb: Optional[str] = None
    concepto_default: Optional[int] = None

class ReturnItemCreate(BaseModel):
    producto_id: str
    cantidad: float

class ReturnCreate(BaseModel):
    items: List[ReturnItemCreate]
    motivo: Optional[str] = None

class RetryNcRequest(BaseModel):
    cuit_receptor: Optional[str] = None

class SaleReturn(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    sale_id: str
    cajero_id: str
    items: List[SaleItem]
    total: float
    motivo: Optional[str] = None
    fecha: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    numero_devolucion: str

class CreditNote(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    sale_id: str
    return_id: str
    cajero_id: str
    numero_nota_credito: str
    numero_factura_original: str
    items: List[SaleItem]
    total: float
    motivo: Optional[str] = None
    fecha: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    tipo: str  # "parcial" | "total"
    # Campos AFIP (3=NC-A, 8=NC-B, 13=NC-C)
    cae: Optional[str] = None
    cae_vencimiento: Optional[str] = None
    afip_estado: str = "no_aplica"
    tipo_comprobante_nc: Optional[int] = None
    nro_comprobante_afip: Optional[int] = None
    afip_error: Optional[str] = None

# --- Proveedor models ---
class Proveedor(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nombre: str
    ruc_cuit: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None
    activo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProveedorCreate(BaseModel):
    nombre: str
    ruc_cuit: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None

class ProveedorUpdate(BaseModel):
    nombre: Optional[str] = None
    ruc_cuit: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None
    activo: Optional[bool] = None

# --- Compra models ---
class CompraItem(BaseModel):
    descripcion: str
    cantidad: float
    precio_unitario: float
    subtotal: float
    product_id: Optional[str] = None
    actualizar_precio: bool = False
    nuevo_precio: Optional[float] = None
    nuevo_margen: Optional[float] = None
    costo_anterior: Optional[float] = None

class Compra(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    sucursal_id: Optional[str] = None
    proveedor_id: Optional[str] = None
    proveedor_nombre: Optional[str] = None
    numero_factura: Optional[str] = None
    fecha: datetime
    items: List[CompraItem] = []
    subtotal: float
    impuestos: float = 0.0
    total: float
    notas: Optional[str] = None
    distribuciones: List[dict] = []
    registrado_por: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CompraCreate(BaseModel):
    sucursal_id: Optional[str] = None
    proveedor_id: Optional[str] = None
    numero_factura: Optional[str] = None
    fecha: datetime
    items: List[CompraItem] = []
    subtotal: float
    impuestos: float = 0.0
    total: float
    notas: Optional[str] = None

class CompraUpdate(BaseModel):
    sucursal_id: Optional[str] = None
    proveedor_id: Optional[str] = None
    numero_factura: Optional[str] = None
    fecha: Optional[datetime] = None
    items: Optional[List[CompraItem]] = None
    subtotal: Optional[float] = None
    impuestos: Optional[float] = None
    total: Optional[float] = None
    notas: Optional[str] = None

class DistribucionItem(BaseModel):
    product_id: str
    cantidad: float = 0
    nuevo_precio: Optional[float] = None
    nuevo_margen: Optional[float] = None
    actualizar_stock: bool = True
    actualizar_precio: bool = False

class DistribucionCreate(BaseModel):
    sucursal_id: str
    items: List[DistribucionItem]

# Utility functions
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        empresa_id: str = payload.get("empresa_id")
        active_branch_id: Optional[str] = payload.get("active_branch_id")
        if user_id is None or empresa_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")

        user_doc = await db.users.find_one({"id": user_id, "empresa_id": empresa_id})
        if user_doc is None:
            raise HTTPException(status_code=401, detail="User not found")

        user = User(**user_doc)
        if active_branch_id:
            # Admin puede usar cualquier sucursal de su empresa; otros usuarios solo las asignadas
            if user.rol == UserRole.ADMIN or active_branch_id in user.branch_ids:
                user = user.copy(update={"active_branch_id": active_branch_id})
        return user
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

def require_role(required_roles: List[UserRole]):
    def role_checker(user: User = Depends(get_current_user)):
        if user.rol not in required_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

# Branch routes
@api_router.post("/branches", response_model=Branch)
async def create_branch(branch_data: BranchCreate, user: User = Depends(require_role([UserRole.ADMIN]))):
    empresa = await db.empresas.find_one({"id": user.empresa_id})
    plan = (empresa or {}).get("plan", "emprendedor")
    suscripcion_doc = await db.suscripciones.find_one({"empresa_id": user.empresa_id})
    suc_extra = int((suscripcion_doc or {}).get("sucursales_extra", 0))
    limite_base = {"empresarial": 3, "profesional": 1, "emprendedor": 1}.get(plan, 1)
    limite_sucursales = limite_base + (suc_extra if plan == "empresarial" else 0)
    count = await db.branches.count_documents({"empresa_id": user.empresa_id, "activo": True})
    if count >= limite_sucursales:
        raise HTTPException(status_code=400, detail=f"Tu plan permite hasta {limite_sucursales} sucursal(es). Actualizá tu plan para agregar más.")
    branch = Branch(**branch_data.dict(), empresa_id=user.empresa_id)
    await db.branches.insert_one(branch.dict())

    # Build a price/margin reference map from the first existing branch (if any)
    first_branch = await db.branches.find_one(
        {"empresa_id": user.empresa_id, "id": {"$ne": branch.id}},
        sort=[("created_at", 1)]
    )
    ref_map: dict = {}
    if first_branch:
        ref_bps = await db.branch_products.find(
            {"branch_id": first_branch["id"], "empresa_id": user.empresa_id}
        ).to_list(10000)
        ref_map = {bp["product_id"]: bp for bp in ref_bps}

    # Auto-sync: create branch_products for all existing active products of this empresa
    products = await db.products.find({"empresa_id": user.empresa_id, "activo": True}).to_list(10000)
    for product in products:
        existing = await db.branch_products.find_one({
            "product_id": product["id"],
            "branch_id": branch.id,
            "empresa_id": user.empresa_id
        })
        if not existing:
            ref = ref_map.get(product["id"])
            base_precio = ref["precio"] if ref else product["precio"]
            base_precio_por_peso = ref.get("precio_por_peso") if ref else product.get("precio_por_peso")
            base_costo = ref.get("costo") if ref else None
            ajuste = branch_data.margen_ajuste
            if ajuste is not None and ajuste != 0:
                factor = 1 + ajuste / 100
                base_precio = round(base_precio * factor, 2)
                if base_precio_por_peso:
                    base_precio_por_peso = round(base_precio_por_peso * factor, 2)
            nuevo_margen = (
                round((base_precio - base_costo) / base_costo * 100, 2)
                if base_costo and base_costo > 0 else (ref.get("margen") if ref else None)
            )
            bp = BranchProduct(
                empresa_id=user.empresa_id,
                product_id=product["id"],
                branch_id=branch.id,
                precio=base_precio,
                precio_por_peso=base_precio_por_peso,
                stock=product.get("stock", 0),
                stock_minimo=product.get("stock_minimo", 10),
                margen=nuevo_margen,
                costo=base_costo,
            )
            await db.branch_products.insert_one(bp.dict())
    return branch

@api_router.get("/branches", response_model=List[Branch])
async def get_branches(user: User = Depends(get_current_user)):
    branches = await db.branches.find({"empresa_id": user.empresa_id, "activo": True}).to_list(1000)
    return [Branch(**branch) for branch in branches]

@api_router.get("/branches/{branch_id}/products/export")
async def export_branch_products(
    branch_id: str,
    format: str = Query("csv"),
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    if format not in ("csv", "xlsx"):
        format = "csv"
    branch = await db.branches.find_one({"id": branch_id, "empresa_id": user.empresa_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")

    products = await db.products.find({"empresa_id": user.empresa_id, "activo": True}).to_list(10000)
    categories = await db.categories.find({"empresa_id": user.empresa_id}).to_list(1000)
    cat_map = {c["id"]: c["nombre"] for c in categories}

    rows = []
    for product in products:
        bp = await db.branch_products.find_one({
            "product_id": product.get("id"),
            "branch_id": branch_id,
            "empresa_id": user.empresa_id
        })
        rows.append({
            "nombre": product.get("nombre"),
            "codigo_barras": product.get("codigo_barras", ""),
            "tipo": product.get("tipo"),
            "categoria": cat_map.get(product.get("categoria_id"), ""),
            "precio_global": product.get("precio"),
            "stock_global": product.get("stock", 0),
            "precio_sucursal": bp.get("precio") if bp else "",
            "precio_por_peso_sucursal": bp.get("precio_por_peso") if bp else "",
            "stock_sucursal": bp.get("stock") if bp else "",
            "stock_minimo_sucursal": bp.get("stock_minimo") if bp else "",
        })

    df = pd.DataFrame(rows)
    branch_nombre = branch.get("nombre", branch_id).replace(" ", "_")

    if format == "xlsx":
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Productos")
        output.seek(0)
        filename = f"productos_{branch_nombre}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
        filename = f"productos_{branch_nombre}.csv"
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

@api_router.get("/branches/{branch_id}/products")
async def get_branch_products_admin(
    branch_id: str,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=10000),
    search: Optional[str] = Query(None),
    category_id: Optional[str] = Query(None),
    kind: Optional[str] = Query(None),
    activo_sucursal: Optional[bool] = Query(None),
    all: bool = Query(False),
    user: User = Depends(get_current_user)
):
    if user.rol not in [UserRole.ADMIN, UserRole.SUPERVISOR]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    branch = await db.branches.find_one({"id": branch_id, "empresa_id": user.empresa_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    query = {"empresa_id": user.empresa_id, "activo": True}
    if search:
        regex = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [{"nombre": regex}, {"codigo_barras": regex}]
    if category_id:
        query["categoria_id"] = category_id
    if kind:
        query["kind"] = kind
    if activo_sucursal is not None:
        bp_query = {"branch_id": branch_id, "empresa_id": user.empresa_id, "activo": activo_sucursal}
        filtered_bps = await db.branch_products.find(bp_query, {"product_id": 1}).to_list(None)
        filtered_ids = [bp["product_id"] for bp in filtered_bps]
        if activo_sucursal:
            # activo=True: products with explicit activo=True entry (exclude those with activo=False)
            inactive_bps = await db.branch_products.find(
                {"branch_id": branch_id, "empresa_id": user.empresa_id, "activo": False}, {"product_id": 1}
            ).to_list(None)
            inactive_ids = {bp["product_id"] for bp in inactive_bps}
            query["id"] = {"$nin": list(inactive_ids)}
        else:
            query["id"] = {"$in": filtered_ids}
    total = await db.products.count_documents(query)
    if all:
        products = await db.products.find(query).to_list(None)
    else:
        skip = (page - 1) * per_page
        products = await db.products.find(query).skip(skip).limit(per_page).to_list(per_page)
    product_ids = [p.get("id") for p in products]
    bps = await db.branch_products.find({
        "product_id": {"$in": product_ids},
        "branch_id": branch_id,
        "empresa_id": user.empresa_id
    }).to_list(None)
    bp_map = {bp["product_id"]: bp for bp in bps}
    result = []
    for product in products:
        bp = bp_map.get(product.get("id"))
        result.append({
            "product_id": product.get("id"),
            "nombre": product.get("nombre"),
            "codigo_barras": product.get("codigo_barras"),
            "tipo": product.get("tipo"),
            "categoria_id": product.get("categoria_id"),
            "precio_global": product.get("precio"),
            "stock_global": product.get("stock", 0),
            "branch_product_id": bp.get("id") if bp else None,
            "precio_sucursal": bp.get("precio") if bp else None,
            "precio_por_peso_sucursal": bp.get("precio_por_peso") if bp else None,
            "stock_sucursal": bp.get("stock") if bp else None,
            "stock_minimo_sucursal": bp.get("stock_minimo") if bp else None,
            "margen_sucursal": bp.get("margen") if bp else None,
            "costo_sucursal": bp.get("costo") if bp else None,
            "activo_sucursal": bp.get("activo", True) if bp else True,
        })
    return {
        "items": result,
        "total": total,
        "page": page,
        "per_page": len(result) if all else per_page,
        "total_pages": 1 if all else max(1, -(-total // per_page))
    }

@api_router.get("/branches/{branch_id}", response_model=Branch)
async def get_branch(branch_id: str, user: User = Depends(get_current_user)):
    branch = await db.branches.find_one({"id": branch_id, "empresa_id": user.empresa_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    return Branch(**branch)

@api_router.put("/branches/{branch_id}", response_model=Branch)
async def update_branch(branch_id: str, branch_data: BranchUpdate, user: User = Depends(require_role([UserRole.ADMIN]))):
    branch = await db.branches.find_one({"id": branch_id, "empresa_id": user.empresa_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    update_data = {k: v for k, v in branch_data.dict().items() if v is not None}
    if update_data:
        await db.branches.update_one({"id": branch_id, "empresa_id": user.empresa_id}, {"$set": update_data})
    updated = await db.branches.find_one({"id": branch_id, "empresa_id": user.empresa_id})
    return Branch(**updated)

@api_router.delete("/branches/{branch_id}", status_code=204)
async def delete_branch(branch_id: str, user: User = Depends(require_role([UserRole.ADMIN]))):
    branch = await db.branches.find_one({"id": branch_id, "empresa_id": user.empresa_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")

    # Collect session ids to cascade-delete movements
    session_ids = [s["id"] async for s in db.cash_sessions.find(
        {"branch_id": branch_id, "empresa_id": user.empresa_id}, {"id": 1}
    )]

    if session_ids:
        await db.cash_movements.delete_many({"session_id": {"$in": session_ids}})
    await db.cash_sessions.delete_many({"branch_id": branch_id, "empresa_id": user.empresa_id})
    await db.branch_products.delete_many({"branch_id": branch_id, "empresa_id": user.empresa_id})
    await db.users.update_many(
        {"empresa_id": user.empresa_id},
        {"$pull": {"branch_ids": branch_id}}
    )
    await db.users.update_many(
        {"branch_id": branch_id, "empresa_id": user.empresa_id},
        {"$unset": {"branch_id": ""}}
    )
    await db.branches.delete_one({"id": branch_id, "empresa_id": user.empresa_id})

# Cash Session routes
@api_router.post("/cash-sessions", response_model=CashSession)
async def open_cash_session(session_data: CashSessionCreate, user: User = Depends(get_current_user)):
    if not user.branch_ids:
        raise HTTPException(status_code=400, detail="El usuario debe estar asignado a una sucursal")

    session_branch_id = session_data.branch_id
    if not session_branch_id:
        if len(user.branch_ids) == 1:
            session_branch_id = user.branch_ids[0]
        else:
            raise HTTPException(status_code=400, detail="Debe seleccionar una sucursal para abrir la caja")
    elif session_branch_id not in user.branch_ids:
        raise HTTPException(status_code=403, detail="No tiene acceso a esa sucursal")

    # Check if there's already an open session for this user in this branch
    existing_session = await db.cash_sessions.find_one({
        "user_id": user.id,
        "empresa_id": user.empresa_id,
        "branch_id": session_branch_id,
        "status": CashSessionStatus.ABIERTA
    })
    if existing_session:
        raise HTTPException(status_code=400, detail="Ya hay una caja abierta en esta sucursal")

    # Create new session
    session = CashSession(
        empresa_id=user.empresa_id,
        branch_id=session_branch_id,
        user_id=user.id,
        monto_inicial=session_data.monto_inicial,
        observaciones=session_data.observaciones
    )

    await db.cash_sessions.insert_one(session.dict())

    # Create opening movement
    movement = CashMovement(
        empresa_id=user.empresa_id,
        session_id=session.id,
        tipo=MovementType.APERTURA,
        monto=session_data.monto_inicial,
        descripcion=f"Apertura de caja - {session_data.observaciones or ''}"
    )
    await db.cash_movements.insert_one(movement.dict())

    return session

@api_router.put("/cash-sessions/{session_id}/close", response_model=CashSession)
async def close_cash_session(session_id: str, close_data: CashSessionClose, user: User = Depends(get_current_user)):
    query = {"id": session_id, "empresa_id": user.empresa_id}
    if user.rol != UserRole.ADMIN:
        query["user_id"] = user.id
    session = await db.cash_sessions.find_one(query)
    if not session:
        raise HTTPException(status_code=404, detail="Sesión de caja no encontrada")

    if session["status"] == CashSessionStatus.CERRADA:
        raise HTTPException(status_code=400, detail="La sesión ya está cerrada")

    # Calculate expected amount
    monto_esperado = session["monto_inicial"] + session["monto_ventas"] - session["monto_retiros"]
    diferencia = close_data.monto_final - monto_esperado

    # Update session
    update_data = {
        "monto_final": close_data.monto_final,
        "monto_esperado": monto_esperado,
        "diferencia": diferencia,
        "status": CashSessionStatus.CERRADA,
        "fecha_cierre": datetime.now(timezone.utc),
        "observaciones": close_data.observaciones
    }

    await db.cash_sessions.update_one({"id": session_id, "empresa_id": user.empresa_id}, {"$set": update_data})

    # Create closing movement
    movement = CashMovement(
        empresa_id=user.empresa_id,
        session_id=session_id,
        tipo=MovementType.CIERRE,
        monto=close_data.monto_final,
        descripcion=f"Cierre de caja - {close_data.observaciones or ''}"
    )
    await db.cash_movements.insert_one(movement.dict())

    updated_session = await db.cash_sessions.find_one({"id": session_id, "empresa_id": user.empresa_id})
    return CashSession(**updated_session)

@api_router.get("/cash-sessions/current", response_model=Optional[CashSession])
async def get_current_cash_session(user: User = Depends(get_current_user)):
    if not user.branch_id:
        return None
    session = await db.cash_sessions.find_one({
        "user_id": user.id,
        "empresa_id": user.empresa_id,
        "branch_id": user.branch_id,
        "status": CashSessionStatus.ABIERTA
    })
    return CashSession(**session) if session else None

@api_router.get("/cash-sessions", response_model=List[CashSession])
async def get_cash_sessions(user: User = Depends(get_current_user)):
    if user.rol == UserRole.CAJERO:
        sessions = await db.cash_sessions.find({"user_id": user.id, "empresa_id": user.empresa_id}).sort("fecha_apertura", -1).to_list(100)
    else:
        query = {"empresa_id": user.empresa_id}
        if user.branch_ids and user.rol != UserRole.ADMIN:
            query["branch_id"] = {"$in": user.branch_ids}
        sessions = await db.cash_sessions.find(query).sort("fecha_apertura", -1).to_list(1000)

    return [CashSession(**session) for session in sessions]

@api_router.get("/cash-sessions/history")
async def get_cash_sessions_history(
    user: User = Depends(get_current_user),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    branch_id: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
):
    if user.rol == UserRole.CAJERO:
        raise HTTPException(status_code=403, detail="No tiene permisos para ver el historial")

    query = {"empresa_id": user.empresa_id}
    if user.rol != UserRole.ADMIN and user.branch_ids:
        query["branch_id"] = {"$in": user.branch_ids}
    if branch_id:
        query["branch_id"] = branch_id
    if user_id:
        query["user_id"] = user_id

    total = await db.cash_sessions.count_documents(query)
    skip = (page - 1) * per_page
    sessions = await db.cash_sessions.find(query).sort("fecha_apertura", -1).skip(skip).limit(per_page).to_list(per_page)

    user_cache: dict = {}
    branch_cache: dict = {}
    items = []

    for s in sessions:
        uid = s.get("user_id", "")
        bid = s.get("branch_id", "")

        if uid not in user_cache:
            u = await db.users.find_one({"id": uid})
            user_cache[uid] = u["nombre"] if u else "Usuario desconocido"

        if bid not in branch_cache:
            b = await db.branches.find_one({"id": bid})
            branch_cache[bid] = b["nombre"] if b else "Sucursal desconocida"

        items.append(CashSessionHistory(
            id=s["id"],
            empresa_id=s["empresa_id"],
            branch_id=bid,
            branch_nombre=branch_cache[bid],
            user_id=uid,
            user_nombre=user_cache[uid],
            monto_inicial=s.get("monto_inicial", 0),
            monto_ventas=s.get("monto_ventas", 0),
            monto_retiros=s.get("monto_retiros", 0),
            monto_final=s.get("monto_final"),
            diferencia=s.get("diferencia"),
            status=s["status"],
            fecha_apertura=s["fecha_apertura"],
            fecha_cierre=s.get("fecha_cierre"),
            observaciones=s.get("observaciones"),
        ))

    return {
        "items": [i.dict() for i in items],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, -(-total // per_page)),
    }

# Auth routes
@api_router.post("/auth/otp/enviar")
async def otp_enviar(data: OTPEnviar):
    email = data.email.strip().lower()

    # Rate limiting: máx 3 códigos activos por email
    recientes = await db.email_otps.count_documents({
        "email": email,
        "expires_at": {"$gt": datetime.now(timezone.utc)},
        "usado": False,
    })
    if recientes >= 3:
        raise HTTPException(status_code=429, detail="Demasiados intentos. Esperá unos minutos.")

    codigo = str(random.randint(1000, 9999))
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)
    await db.email_otps.insert_one({
        "email": email,
        "codigo": codigo,
        "expires_at": expires_at,
        "usado": False,
    })

    try:
        await send_email_otp(email, codigo)
    except Exception as e:
        logging.exception("Error enviando email OTP")
        raise HTTPException(status_code=500, detail="No se pudo enviar el código. Verificá el correo e intentá de nuevo.")

    return {"ok": True, "mensaje": "Código enviado al correo"}


@api_router.post("/auth/otp/verificar")
async def otp_verificar(data: OTPVerificar):
    email = data.email.strip().lower()
    otp_doc = await db.email_otps.find_one({
        "email": email,
        "codigo": data.codigo,
        "usado": False,
        "expires_at": {"$gt": datetime.now(timezone.utc)},
    })
    if not otp_doc:
        raise HTTPException(status_code=400, detail="Código inválido o expirado")

    await db.email_otps.update_one({"_id": otp_doc["_id"]}, {"$set": {"usado": True}})

    token = jwt.encode(
        {"type": "email_otp", "email": email, "exp": datetime.now(timezone.utc) + timedelta(minutes=15)},
        SECRET_KEY,
        algorithm=ALGORITHM,
    )
    return {"ok": True, "verificacion_token": token}


@api_router.post("/auth/empresa/solicitar")
async def solicitar_cuenta(data: EmpresaSolicitud):
    try:
        await send_solicitud_notify(data.empresa_nombre, data.admin_nombre, data.admin_email, data.telefono)
    except Exception:
        logging.exception("Error enviando notificación de solicitud")
        raise HTTPException(status_code=500, detail="No se pudo enviar la solicitud. Intentá de nuevo.")
    return {"ok": True, "mensaje": "Solicitud enviada. Te contactaremos pronto."}


@api_router.post("/auth/empresa/register", response_model=Token)
async def register_empresa(data: EmpresaRegister):
    # Validar token OTP
    try:
        payload = jwt.decode(data.otp_token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("type") != "email_otp" or payload.get("email") != data.admin_email.strip().lower():
            raise ValueError()
    except Exception:
        raise HTTPException(status_code=400, detail="Token de verificación de correo inválido o expirado")

    # Check if admin email already exists
    existing_user = await db.users.find_one({"email": data.admin_email})
    if existing_user:
        raise HTTPException(status_code=400, detail="El email ya está registrado")

    # Create empresa
    empresa = Empresa(nombre=data.empresa_nombre, email_verificado=True)
    await db.empresas.insert_one(empresa.dict())

    # Create admin user
    hashed_password = bcrypt.hash(data.admin_password)
    user = User(
        empresa_id=empresa.id,
        nombre=data.admin_nombre,
        email=data.admin_email,
        rol=UserRole.ADMIN
    )
    user_doc = user.dict()
    user_doc['password'] = hashed_password
    await db.users.insert_one(user_doc)

    # Create default configuration for this empresa
    config = Configuration(empresa_id=empresa.id, company_name=data.empresa_nombre)
    await db.configuration.insert_one(config.dict())

    # Create trial subscription
    trial_inicio = datetime.now(timezone.utc)
    trial_fin = trial_inicio + timedelta(days=await get_trial_dias())
    dia_facturacion = min(trial_fin.day, 28)
    suscripcion = Suscripcion(
        empresa_id=empresa.id,
        plan_nombre=await get_plan_nombre_suscripcion(),
        precio=await get_precio_suscripcion(),
        status=SuscripcionStatus.TRIAL,
        fecha_inicio=trial_inicio,
        fecha_vencimiento=trial_fin,
        dia_facturacion=dia_facturacion,
        plan_tipo="mensual",
        plan_tier="empresarial",
    )
    await db.suscripciones.insert_one(suscripcion.dict())

    # Create default branch
    default_branch = Branch(
        empresa_id=empresa.id,
        nombre="Sucursal Principal",
        direccion="",
    )
    await db.branches.insert_one(default_branch.dict())

    # Create default categories
    for cat_nombre in ["General", "Alimentos", "Bebidas", "Limpieza"]:
        category = Category(empresa_id=empresa.id, nombre=cat_nombre)
        await db.categories.insert_one(category.dict())

    # Return token (auto-login)
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.id, "empresa_id": empresa.id},
        expires_delta=access_token_expires
    )
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    empresa = await db.empresas.find_one({"id": current_user.empresa_id})
    plan = (empresa or {}).get("plan", "emprendedor")
    suscripcion_doc = await db.suscripciones.find_one({"empresa_id": current_user.empresa_id})
    usr_extra_packs = int((suscripcion_doc or {}).get("usuarios_extra_packs", 0))
    limite_base = {"empresarial": 15, "profesional": 5, "emprendedor": 2}.get(plan, 2)
    limite_usuarios = limite_base + (usr_extra_packs * 5 if plan == "empresarial" else 0)
    count = await db.users.count_documents({"empresa_id": current_user.empresa_id, "activo": True})
    if count >= limite_usuarios:
        raise HTTPException(status_code=400, detail=f"Tu plan permite hasta {limite_usuarios} usuario(s). Actualizá tu plan para agregar más.")
    # Check if email already exists within this empresa
    existing_user = await db.users.find_one({"email": user_data.email, "empresa_id": current_user.empresa_id})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Hash password
    hashed_password = bcrypt.hash(user_data.password)

    # Create user — empresa_id inherited from the admin creating it
    user_dict = user_data.dict()
    user_dict.pop('password')
    user = User(**user_dict, empresa_id=current_user.empresa_id)

    # Store in database (exclude computed/runtime fields)
    user_doc = user.dict(exclude={'active_branch_id', 'branch_id'})
    user_doc['password'] = hashed_password
    await db.users.insert_one(user_doc)

    return user

class PasswordResetEnviar(BaseModel):
    email: str

class PasswordResetVerificar(BaseModel):
    email: str
    codigo: str

class PasswordResetCambiar(BaseModel):
    reset_token: str
    nueva_password: str

@api_router.post("/auth/password-reset/enviar")
async def password_reset_enviar(data: PasswordResetEnviar):
    email_input = data.email.strip()
    # Búsqueda case-insensitive para no depender de cómo se guardó el email
    user_doc = await db.users.find_one({"email": {"$regex": f"^{email_input}$", "$options": "i"}})
    # Siempre responder OK para no revelar si el email existe
    if not user_doc:
        return {"ok": True, "mensaje": "Si el correo está registrado, recibirás un código"}

    # Usar el email exactamente como está guardado en la DB
    email = user_doc["email"]

    recientes = await db.password_resets.count_documents({
        "email": email,
        "expires_at": {"$gt": datetime.now(timezone.utc)},
        "usado": False,
    })
    if recientes >= 3:
        raise HTTPException(status_code=429, detail="Demasiados intentos. Esperá unos minutos.")

    codigo = str(random.randint(1000, 9999))
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=15)
    await db.password_resets.insert_one({
        "email": email,
        "codigo": codigo,
        "expires_at": expires_at,
        "usado": False,
    })

    try:
        await send_password_reset_email(email, codigo)
    except Exception as e:
        logging.exception("Error enviando email de recuperación")
        raise HTTPException(status_code=500, detail="No se pudo enviar el correo. Intentá de nuevo.")

    return {"ok": True, "mensaje": "Si el correo está registrado, recibirás un código"}


@api_router.post("/auth/password-reset/verificar")
async def password_reset_verificar(data: PasswordResetVerificar):
    email_input = data.email.strip()
    doc = await db.password_resets.find_one({
        "email": {"$regex": f"^{email_input}$", "$options": "i"},
        "codigo": data.codigo,
        "usado": False,
        "expires_at": {"$gt": datetime.now(timezone.utc)},
    })
    if not doc:
        raise HTTPException(status_code=400, detail="Código inválido o expirado")

    await db.password_resets.update_one({"_id": doc["_id"]}, {"$set": {"usado": True}})

    # Usar el email exactamente como está en el registro guardado
    email = doc["email"]
    token = jwt.encode(
        {"type": "password_reset", "email": email, "exp": datetime.now(timezone.utc) + timedelta(minutes=15)},
        SECRET_KEY,
        algorithm=ALGORITHM,
    )
    return {"ok": True, "reset_token": token}


@api_router.post("/auth/password-reset/cambiar")
async def password_reset_cambiar(data: PasswordResetCambiar):
    try:
        payload = jwt.decode(data.reset_token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("type") != "password_reset":
            raise ValueError()
        email = payload["email"]
    except Exception:
        raise HTTPException(status_code=400, detail="Token de recuperación inválido o expirado")

    if len(data.nueva_password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")

    hashed = bcrypt.hash(data.nueva_password)
    result = await db.users.update_one({"email": email}, {"$set": {"password": hashed}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    return {"ok": True, "mensaje": "Contraseña actualizada correctamente"}


@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    # Find user
    user_doc = await db.users.find_one({"email": user_data.email})
    if not user_doc or not bcrypt.verify(user_data.password, user_doc['password']):
        raise HTTPException(status_code=400, detail="Incorrect email or password")

    if not user_doc.get('activo', True):
        raise HTTPException(status_code=400, detail="User account is disabled")

    # Verificar que el correo de la empresa esté verificado
    empresa_doc = await db.empresas.find_one({"id": user_doc['empresa_id']})
    if empresa_doc and not empresa_doc.get('email_verificado', True):
        raise HTTPException(status_code=403, detail="Correo no verificado. Completá el registro de tu empresa.")

    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    user = User(**user_doc)

    token_data: dict = {"sub": user_doc['id'], "empresa_id": user_doc['empresa_id']}
    # Solo auto-asignar sucursal para usuarios no-admin con exactamente una sucursal asignada
    if user.rol != UserRole.ADMIN and len(user.branch_ids) == 1:
        token_data["active_branch_id"] = user.branch_ids[0]
        user = user.copy(update={"active_branch_id": user.branch_ids[0]})

    access_token = create_access_token(data=token_data, expires_delta=access_token_expires)
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(user: User = Depends(get_current_user)):
    return user

class SelectBranchRequest(BaseModel):
    branch_id: str

@api_router.post("/auth/select-branch", response_model=Token)
async def select_branch(data: SelectBranchRequest, current_user: User = Depends(get_current_user)):
    if data.branch_id not in current_user.branch_ids:
        if current_user.rol == UserRole.ADMIN:
            branch = await db.branches.find_one({"id": data.branch_id, "empresa_id": current_user.empresa_id})
            if not branch:
                raise HTTPException(status_code=403, detail="Sucursal no encontrada en esta empresa")
        else:
            raise HTTPException(status_code=403, detail="No tiene acceso a esa sucursal")
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": current_user.id, "empresa_id": current_user.empresa_id, "active_branch_id": data.branch_id},
        expires_delta=access_token_expires
    )
    updated_user = current_user.copy(update={"active_branch_id": data.branch_id})
    return Token(access_token=access_token, token_type="bearer", user=updated_user)

# Category routes
@api_router.post("/categories", response_model=Category)
async def create_category(category_data: CategoryCreate, user: User = Depends(require_role([UserRole.ADMIN]))):
    category = Category(**category_data.dict(), empresa_id=user.empresa_id)
    await db.categories.insert_one(category.dict())
    return category

@api_router.get("/categories", response_model=List[Category])
async def get_categories(user: User = Depends(get_current_user)):
    categories = await db.categories.find({"empresa_id": user.empresa_id}).to_list(1000)
    return [Category(**cat) for cat in categories]

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, category_data: CategoryCreate, user: User = Depends(require_role([UserRole.ADMIN]))):
    category = await db.categories.find_one({"id": category_id, "empresa_id": user.empresa_id})
    if not category:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    await db.categories.update_one(
        {"id": category_id},
        {"$set": {"nombre": category_data.nombre, "descripcion": category_data.descripcion, "icono": category_data.icono}}
    )
    updated = await db.categories.find_one({"id": category_id})
    return Category(**updated)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, user: User = Depends(require_role([UserRole.ADMIN]))):
    category = await db.categories.find_one({"id": category_id, "empresa_id": user.empresa_id})
    if not category:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    products_count = await db.products.count_documents({"categoria_id": category_id, "empresa_id": user.empresa_id})
    if products_count > 0:
        raise HTTPException(status_code=400, detail=f"No se puede eliminar: {products_count} producto(s) usan esta categoría")
    await db.categories.delete_one({"id": category_id})
    return {"message": "Categoría eliminada"}

# Branch Product routes
@api_router.post("/branch-products", response_model=BranchProduct)
async def create_branch_product(product_data: BranchProductCreate, user: User = Depends(require_role([UserRole.ADMIN]))):
    target_branch_id = product_data.branch_id or user.branch_id
    if not target_branch_id:
        raise HTTPException(status_code=400, detail="branch_id requerido")
    # Verify product exists for this empresa
    product = await db.products.find_one({"id": product_data.product_id, "empresa_id": user.empresa_id})
    if not product:
        raise HTTPException(status_code=400, detail="Product not found")
    # Check if product already exists in branch
    existing = await db.branch_products.find_one({
        "product_id": product_data.product_id,
        "branch_id": target_branch_id,
        "empresa_id": user.empresa_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Product already exists in this branch")
    data = {k: v for k, v in product_data.dict().items() if k != "branch_id"}
    branch_product = BranchProduct(**data, branch_id=target_branch_id, empresa_id=user.empresa_id)
    await db.branch_products.insert_one(branch_product.dict())
    return branch_product

@api_router.put("/branch-products/{branch_product_id}", response_model=BranchProduct)
async def update_branch_product(branch_product_id: str, product_data: BranchProductUpdate, user: User = Depends(require_role([UserRole.ADMIN]))):
    bp = await db.branch_products.find_one({"id": branch_product_id, "empresa_id": user.empresa_id})
    if not bp:
        raise HTTPException(status_code=404, detail="Branch product not found")
    update_data = {k: v for k, v in product_data.dict().items() if v is not None}
    if update_data:
        await db.branch_products.update_one({"id": branch_product_id, "empresa_id": user.empresa_id}, {"$set": update_data})
    updated = await db.branch_products.find_one({"id": branch_product_id, "empresa_id": user.empresa_id})
    return BranchProduct(**updated)

@api_router.post("/branch-products/bulk-deactivate")
async def bulk_deactivate_branch_products(data: BranchProductBulkDeleteRequest, user: User = Depends(require_role([UserRole.ADMIN]))):
    branch_id = data.branch_id
    if not branch_id:
        return {"updated": 0}

    if data.delete_all:
        query = {"empresa_id": user.empresa_id}
        if data.search:
            regex = {"$regex": re.escape(data.search), "$options": "i"}
            matching = await db.products.find(
                {"empresa_id": user.empresa_id, "$or": [{"nombre": regex}, {"codigo_barras": regex}]},
                {"id": 1}
            ).to_list(None)
            product_ids = [p["id"] for p in matching]
        else:
            all_products = await db.products.find({"empresa_id": user.empresa_id}, {"id": 1}).to_list(None)
            product_ids = [p["id"] for p in all_products]
    else:
        product_ids = data.ids or []

    if not product_ids:
        return {"updated": 0}

    existing_bps = await db.branch_products.find(
        {"product_id": {"$in": product_ids}, "branch_id": branch_id, "empresa_id": user.empresa_id}
    ).to_list(None)
    existing_map = {bp["product_id"]: bp for bp in existing_bps}

    updated = 0
    for pid in product_ids:
        if pid in existing_map:
            await db.branch_products.update_one(
                {"id": existing_map[pid]["id"]},
                {"$set": {"activo": False}}
            )
            updated += 1
        else:
            global_product = await db.products.find_one({"id": pid, "empresa_id": user.empresa_id})
            if global_product:
                bp = BranchProduct(
                    product_id=pid,
                    branch_id=branch_id,
                    empresa_id=user.empresa_id,
                    precio=global_product.get("precio", 0),
                    stock=global_product.get("stock", 0),
                    stock_minimo=global_product.get("stock_minimo", 10),
                    activo=False
                )
                await db.branch_products.insert_one(bp.dict())
                updated += 1

    return {"updated": updated}

@api_router.delete("/branch-products/bulk")
async def bulk_delete_branch_products(data: BranchProductBulkDeleteRequest, user: User = Depends(require_role([UserRole.ADMIN]))):
    if data.delete_all and data.branch_id:
        query = {"branch_id": data.branch_id, "empresa_id": user.empresa_id}
        if data.search:
            regex = {"$regex": re.escape(data.search), "$options": "i"}
            matching_products = await db.products.find(
                {"empresa_id": user.empresa_id, "$or": [{"nombre": regex}, {"codigo_barras": regex}]},
                {"id": 1}
            ).to_list(None)
            matching_product_ids = [p["id"] for p in matching_products]
            query["product_id"] = {"$in": matching_product_ids}
        result = await db.branch_products.delete_many(query)
        return {"deleted": result.deleted_count}
    ids = data.ids or []
    if not ids:
        return {"deleted": 0}
    result = await db.branch_products.delete_many({"id": {"$in": ids}, "empresa_id": user.empresa_id})
    return {"deleted": result.deleted_count}

@api_router.delete("/branch-products/{branch_product_id}")
async def delete_branch_product(branch_product_id: str, user: User = Depends(require_role([UserRole.ADMIN]))):
    bp = await db.branch_products.find_one({"id": branch_product_id, "empresa_id": user.empresa_id})
    if not bp:
        raise HTTPException(status_code=404, detail="Branch product not found")
    await db.branch_products.delete_one({"id": branch_product_id, "empresa_id": user.empresa_id})
    return {"message": "Producto eliminado de la sucursal"}

@api_router.get("/branch-products")
async def get_branch_products(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=10000),
    search: Optional[str] = Query(None),
    user: User = Depends(get_current_user)
):
    if not user.branch_id:
        raise HTTPException(status_code=400, detail="El usuario debe estar asignado a una sucursal")

    base_pipeline = [
        {"$match": {"branch_id": user.branch_id, "empresa_id": user.empresa_id, "activo": True}},
        {"$lookup": {"from": "products", "localField": "product_id", "foreignField": "id", "as": "product"}},
        {"$unwind": "$product"},
        {"$project": {
            "_id": 0, "id": 1, "product_id": 1, "branch_id": 1, "empresa_id": 1,
            "precio": 1, "precio_por_peso": 1, "stock": 1, "stock_minimo": 1, "activo": 1,
            "nombre": "$product.nombre", "codigo_barras": "$product.codigo_barras",
            "tipo": "$product.tipo", "categoria_id": "$product.categoria_id",
            "control_stock": "$product.control_stock"
        }}
    ]

    if search:
        regex = {"$regex": re.escape(search), "$options": "i"}
        base_pipeline.append({"$match": {"$or": [{"nombre": regex}, {"codigo_barras": regex}]}})

    count_result = await db.branch_products.aggregate(base_pipeline + [{"$count": "total"}]).to_list(1)
    total = count_result[0]["total"] if count_result else 0

    paginated_pipeline = base_pipeline + [{"$skip": (page - 1) * per_page}, {"$limit": per_page}]
    items = await db.branch_products.aggregate(paginated_pipeline).to_list(per_page)

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, -(-total // per_page))
    }

# Product routes
@api_router.post("/products", response_model=Product)
async def create_product(product_data: ProductCreate, user: User = Depends(require_role([UserRole.ADMIN]))):
    # Verify category exists for this empresa
    category = await db.categories.find_one({"id": product_data.categoria_id, "empresa_id": user.empresa_id})
    if not category:
        raise HTTPException(status_code=400, detail="Category not found")

    # Check if barcode already exists for this empresa
    if product_data.codigo_barras:
        existing_product = await db.products.find_one({"codigo_barras": product_data.codigo_barras, "empresa_id": user.empresa_id})
        if existing_product:
            raise HTTPException(status_code=400, detail="Barcode already exists")

    product_dict = product_data.dict()
    precio_costo = product_dict.pop("precio_costo", None)
    margen = round((product_dict["precio"] - precio_costo) / precio_costo * 100, 2) if precio_costo else None
    # Combos never control their own stock
    if product_dict.get("kind") == ProductKind.COMBO:
        product_dict["control_stock"] = False
    product = Product(**product_dict, empresa_id=user.empresa_id)
    _cb = product_dict.get('codigo_barras')
    _pdoc = {k: v for k, v in product.model_dump().items() if k != 'codigo_barras' and v is not None}
    if _cb:
        _pdoc['codigo_barras'] = _cb
    await db.products.insert_one(_pdoc)
    # Auto-sync: create branch_products for all active branches of this empresa
    branches = await db.branches.find({"activo": True, "empresa_id": user.empresa_id}).to_list(1000)
    for branch in branches:
        existing = await db.branch_products.find_one({
            "product_id": product.id,
            "branch_id": branch["id"],
            "empresa_id": user.empresa_id
        })
        if not existing:
            bp_kwargs = dict(
                empresa_id=user.empresa_id,
                product_id=product.id,
                branch_id=branch["id"],
                precio=product.precio,
                precio_por_peso=product.precio_por_peso,
                stock=product.stock,
                stock_minimo=product.stock_minimo,
            )
            if precio_costo is not None:
                bp_kwargs["costo"] = precio_costo
            if margen is not None:
                bp_kwargs["margen"] = margen
            bp = BranchProduct(**bp_kwargs)
            await db.branch_products.insert_one(bp.dict())
    return product

@api_router.get("/products")
async def get_products(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=10000),
    search: Optional[str] = Query(None),
    category_id: Optional[str] = Query(None),
    kind: Optional[str] = Query(None),
    activo: Optional[bool] = Query(None),
    user: User = Depends(get_current_user)
):
    query = {"empresa_id": user.empresa_id}
    query["activo"] = True if activo is None else activo
    if search:
        regex = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [{"nombre": regex}, {"codigo_barras": regex}]
    if category_id:
        query["categoria_id"] = category_id
    if kind:
        query["kind"] = kind
    total = await db.products.count_documents(query)
    skip = (page - 1) * per_page
    raw = await db.products.find(query).skip(skip).limit(per_page).to_list(per_page)
    result = []
    for p in raw:
        if p.get('kind') != 'combo' and p.get('control_stock') is None:
            p['control_stock'] = True
        if 'codigo_barras' in p and p['codigo_barras'] is None:
            del p['codigo_barras']
        try:
            result.append(Product(**p))
        except Exception as e:
            logger.warning(f"Skipping invalid product {p.get('id')}: {e}")
    return {
        "items": [p.dict() for p in result],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, -(-total // per_page))
    }

@api_router.get("/products/export")
async def export_products(
    format: str = Query("csv"),
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    if format not in ("csv", "xlsx"):
        format = "csv"
    products = await db.products.find({"empresa_id": user.empresa_id, "activo": True}).to_list(10000)
    categories = await db.categories.find({"empresa_id": user.empresa_id}).to_list(1000)
    cat_map = {c["id"]: c["nombre"] for c in categories}

    rows = []
    for p in products:
        rows.append({
            "nombre": p.get("nombre"),
            "codigo_barras": p.get("codigo_barras", ""),
            "tipo": p.get("tipo"),
            "clase": "Combo" if p.get("kind") == "combo" else "Normal",
            "precio": p.get("precio"),
            "precio_por_peso": p.get("precio_por_peso", ""),
            "categoria": cat_map.get(p.get("categoria_id"), ""),
            "stock": p.get("stock"),
            "stock_minimo": p.get("stock_minimo"),
            "activo": p.get("activo"),
        })

    df = pd.DataFrame(rows)

    if format == "xlsx":
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Productos")
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=productos.xlsx"},
        )
    else:
        csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=productos.csv"},
        )

@api_router.get("/products/import-template")
async def get_import_template(user: User = Depends(require_role([UserRole.ADMIN]))):
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # Obtener categorías existentes de la empresa
    db_cats = await db.categories.find({"empresa_id": user.empresa_id}).to_list(1000)
    cat_names = [c["nombre"] for c in db_cats] if db_cats else ["General", "Alimentos", "Bebidas", "Limpieza"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Hoja oculta con las categorías para el dropdown
    ws_cats = wb.create_sheet("Categorias")
    for i, nombre in enumerate(cat_names, start=1):
        ws_cats.cell(row=i, column=1, value=nombre)
    ws_cats.sheet_state = "hidden"

    HDR_FILL  = PatternFill("solid", fgColor="1a7a4a")
    REQ_FILL  = PatternFill("solid", fgColor="e8f5e9")
    OPT_FILL  = PatternFill("solid", fgColor="f3f4f6")
    THIN      = Border(
        left=Side(style="thin", color="cccccc"), right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),  bottom=Side(style="thin", color="cccccc"),
    )
    WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    NORM       = Font(name="Calibri", size=10)

    cols = [
        ("nombre",          28, "OBLIGATORIO. Nombre del producto."),
        ("tipo",            18, "OBLIGATORIO. Valores: codigo_barras  o  por_peso"),
        ("precio",          14, "OBLIGATORIO. Precio de venta (solo numeros). Se redondea al multiplo de 100 mas cercano."),
        ("categoria",       18, "OBLIGATORIO. Seleccionar de la lista desplegable."),
        ("precio_costo",    16, "Opcional. Precio de compra. Calcula el margen automaticamente con el precio de venta."),
        ("codigo_barras",   20, "Opcional. Si ya existe ese codigo el producto se ACTUALIZA."),
        ("stock",           12, "Opcional. Stock inicial. Por defecto: 0"),
        ("stock_minimo",    14, "Opcional. Alerta de stock bajo. Por defecto: 10"),
        ("clase",           14, "Opcional. Clase del producto: Normal o Combo. Por defecto: Normal"),
    ]

    for c, (header, width, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font      = WHITE_BOLD
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 22

    dv = DataValidation(type="list", formula1='"codigo_barras,por_peso"', allow_blank=False)
    dv.sqref = "B2:B1000"
    ws.add_data_validation(dv)

    # Dropdown de categorías referenciando la hoja oculta
    n_cats = len(cat_names)
    dv_cat = DataValidation(
        type="list",
        formula1=f"=Categorias!$A$1:$A${n_cats}",
        allow_blank=True,
    )
    dv_cat.sqref = "D2:D1000"
    ws.add_data_validation(dv_cat)

    dv_clase = DataValidation(type="list", formula1='"Normal,Combo"', allow_blank=True)
    dv_clase.sqref = "I2:I1000"
    ws.add_data_validation(dv_clase)

    for r in range(2, 202):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill   = REQ_FILL if c <= 4 else OPT_FILL
            cell.border = THIN
            cell.font   = NORM
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A2"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"},
    )

@api_router.post("/products/import")
async def import_products(
    file: UploadFile = File(...),
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    content = await file.read()
    filename = file.filename or ""

    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(content))
        elif filename.endswith(".csv"):
            df = pd.read_csv(io.StringIO(content.decode("utf-8-sig")))
        else:
            raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o XLSX.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {str(e)}")

    required_cols = {"nombre", "tipo", "precio", "categoria"}
    missing = required_cols - set(df.columns.str.strip().str.lower())
    if missing:
        raise HTTPException(status_code=400, detail=f"Columnas faltantes: {', '.join(missing)}")

    df.columns = df.columns.str.strip().str.lower()

    await db.products.update_many({"codigo_barras": None}, {"$unset": {"codigo_barras": ""}})
    try:
        await db.command("dropIndexes", "products", index="empresa_id_1_codigo_barras_1")
    except Exception:
        pass
    try:
        await db.command({
            "createIndexes": "products",
            "indexes": [{"key": {"empresa_id": 1, "codigo_barras": 1}, "name": "empresa_id_1_codigo_barras_1", "unique": True, "sparse": True}]
        })
    except Exception as _e:
        logger.error(f"FAILED to create sparse index: {_e}")

    categories = await db.categories.find({"empresa_id": user.empresa_id}).to_list(1000)
    cat_map = {c["nombre"].strip().lower(): c["id"] for c in categories}
    cat_name_map = {c["nombre"].strip().lower(): c["nombre"].strip() for c in categories}
    branches = await db.branches.find({"activo": True, "empresa_id": user.empresa_id}).to_list(1000)
    empresa_id = user.empresa_id
    total_rows = len(df)

    cfg = await db.configuration.find_one({"empresa_id": empresa_id}) or {}
    redondeo = cfg.get("redondeo_precio", 100)

    def _redondear(valor: float) -> float:
        import math
        if not redondeo:
            return round(valor, 2)
        return math.ceil(valor / redondeo) * redondeo

    async def generate():
        created = 0
        updated = 0
        errors = []
        new_categories = []

        for idx, row in df.iterrows():
            try:
                cat_nombre_raw = str(row.get("categoria", "")).strip()
                cat_nombre = cat_nombre_raw.lower()
                categoria_id = cat_map.get(cat_nombre)
                if not categoria_id:
                    new_cat = Category(empresa_id=empresa_id, nombre=cat_nombre_raw)
                    await db.categories.insert_one(new_cat.dict())
                    cat_map[cat_nombre] = new_cat.id
                    cat_name_map[cat_nombre] = cat_nombre_raw
                    new_categories.append(cat_nombre_raw)
                    categoria_id = new_cat.id

                raw_barcode = row.get("codigo_barras")
                if pd.notna(raw_barcode) and str(raw_barcode).strip() not in ("", "nan"):
                    try:
                        codigo_barras = str(int(float(str(raw_barcode).strip())))
                    except (ValueError, OverflowError):
                        codigo_barras = str(raw_barcode).strip()
                else:
                    codigo_barras = f"INT-{uuid.uuid4().hex[:10].upper()}"

                tipo = str(row.get("tipo", "codigo_barras")).strip()
                if tipo not in ("codigo_barras", "por_peso"):
                    tipo = "codigo_barras"

                raw_stock = row.get("stock")
                stock = int(float(raw_stock)) if pd.notna(raw_stock) and str(raw_stock).strip() not in ("", "nan") else 0

                raw_stock_min = row.get("stock_minimo")
                stock_minimo = int(float(raw_stock_min)) if pd.notna(raw_stock_min) and str(raw_stock_min).strip() not in ("", "nan") else 10

                raw_clase = row.get("clase", "")
                kind = "combo" if str(raw_clase).strip().lower() == "combo" else "normal"

                precio_raw = _redondear(float(row["precio"]))

                raw_costo = row.get("precio_costo")
                precio_costo = float(raw_costo) if pd.notna(raw_costo) and str(raw_costo).strip() not in ("", "nan") else None
                margen = round((precio_raw - precio_costo) / precio_costo * 100, 2) if precio_costo else None

                product_data = {
                    "nombre": str(row["nombre"]).strip(),
                    "codigo_barras": codigo_barras,
                    "tipo": tipo,
                    "kind": kind,
                    "precio": precio_raw,
                    "categoria_id": categoria_id,
                    "stock": stock,
                    "stock_minimo": stock_minimo,
                }
                existing = None
                if codigo_barras and not codigo_barras.startswith("INT-"):
                    existing = await db.products.find_one({"codigo_barras": codigo_barras, "empresa_id": empresa_id})

                if existing:
                    await db.products.update_one(
                        {"id": existing["id"], "empresa_id": empresa_id},
                        {"$set": product_data}
                    )
                    bp_update = {"precio": precio_raw}
                    if precio_costo is not None:
                        bp_update["costo"] = precio_costo
                    if margen is not None:
                        bp_update["margen"] = margen
                    await db.branch_products.update_many(
                        {"product_id": existing["id"], "empresa_id": empresa_id},
                        {"$set": bp_update}
                    )
                    updated += 1
                else:
                    product_data['codigo_barras'] = codigo_barras
                    product = Product(**product_data, empresa_id=empresa_id)
                    _pdoc = {k: v for k, v in product.model_dump().items() if v is not None}
                    await db.products.insert_one(_pdoc)
                    for branch in branches:
                        bp_exists = await db.branch_products.find_one({
                            "product_id": product.id,
                            "branch_id": branch["id"],
                            "empresa_id": empresa_id
                        })
                        if not bp_exists:
                            bp_data = dict(
                                empresa_id=empresa_id,
                                product_id=product.id,
                                branch_id=branch["id"],
                                precio=product.precio,
                                precio_por_peso=product.precio_por_peso,
                                stock=product.stock,
                                stock_minimo=product.stock_minimo,
                            )
                            if precio_costo is not None:
                                bp_data["costo"] = precio_costo
                            if margen is not None:
                                bp_data["margen"] = margen
                            bp = BranchProduct(**bp_data)
                            await db.branch_products.insert_one(bp.dict())
                    created += 1

            except Exception as e:
                errors.append(f"Fila {idx + 2}: {str(e)}")

            progress = int(((idx + 1) / total_rows) * 100)
            yield f"data: {json.dumps({'progress': progress, 'processed': idx + 1, 'total': total_rows})}\n\n"
            await asyncio.sleep(0)

        yield f"data: {json.dumps({'done': True, 'created': created, 'updated': updated, 'errors': errors, 'total_procesado': created + updated + len(errors), 'new_categories': new_categories})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


@api_router.get("/branch-products/import-template")
async def get_branch_import_template(
    branch_id: str,
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    branch = await db.branches.find_one({"id": branch_id, "empresa_id": user.empresa_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")

    # Obtener productos actuales de la sucursal (join products + branch_products)
    products = await db.products.find({"empresa_id": user.empresa_id, "activo": True}).to_list(5000)
    product_map = {p["id"]: p for p in products}
    branch_prods = await db.branch_products.find({"branch_id": branch_id, "empresa_id": user.empresa_id}).to_list(5000)
    bp_map = {bp["product_id"]: bp for bp in branch_prods}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Precios y Stock"

    HDR_FILL  = PatternFill("solid", fgColor="1a7a4a")
    REF_FILL  = PatternFill("solid", fgColor="e8e8e8")
    REQ_FILL  = PatternFill("solid", fgColor="fff3cd")
    OPT_FILL  = PatternFill("solid", fgColor="e8f5e9")
    THIN = Border(
        left=Side(style="thin", color="cccccc"), right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),  bottom=Side(style="thin", color="cccccc"),
    )
    WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    GRAY_BOLD  = Font(bold=True, color="555555", name="Calibri", size=10)
    NORM       = Font(name="Calibri", size=10)

    cols = [
        ("codigo_barras", 22, "REQUERIDO. Código de barras para identificar el producto."),
        ("nombre",        30, "Referencia (no se modifica)."),
        ("precio_venta",  16, "Opcional. Nuevo precio de venta en esta sucursal."),
        ("precio_costo",  16, "Opcional. Costo de compra. Recalcula el margen."),
        ("stock",         12, "Opcional. Stock en esta sucursal."),
    ]

    for c, (header, width, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    row = 2
    for prod in sorted(products, key=lambda p: p.get("nombre", "")):
        if not prod.get("codigo_barras") or prod["codigo_barras"].startswith("INT-"):
            continue
        bp = bp_map.get(prod["id"], {})
        cells_data = [
            (prod.get("codigo_barras", ""), REQ_FILL),
            (prod.get("nombre", ""),        REF_FILL),
            (bp.get("precio", prod.get("precio", "")), OPT_FILL),
            (bp.get("costo", ""),           OPT_FILL),
            (bp.get("stock", ""),           OPT_FILL),
        ]
        for c, (value, fill) in enumerate(cells_data, start=1):
            cell = ws.cell(row=row, column=c, value=value)
            cell.fill = fill
            cell.border = THIN
            cell.font = GRAY_BOLD if c <= 2 else NORM
            if c == 2:
                cell.protection = openpyxl.styles.Protection(locked=True)
        ws.row_dimensions[row].height = 18
        row += 1

    # Agregar filas vacías para productos nuevos a ingresar manualmente
    for _ in range(20):
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = REQ_FILL if c == 1 else OPT_FILL
            cell.border = THIN
            cell.font = NORM
        ws.row_dimensions[row].height = 18
        row += 1

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = branch.get("nombre", "sucursal").replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=plantilla_precios_{safe_name}.xlsx"},
    )


@api_router.post("/branch-products/import")
async def import_branch_products(
    branch_id: str,
    file: UploadFile = File(...),
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    branch = await db.branches.find_one({"id": branch_id, "empresa_id": user.empresa_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")

    content = await file.read()
    filename = file.filename or ""

    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(content))
        elif filename.endswith(".csv"):
            df = pd.read_csv(io.StringIO(content.decode("utf-8-sig")))
        else:
            raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o XLSX.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {str(e)}")

    df.columns = df.columns.str.strip().str.lower()

    if "codigo_barras" not in df.columns:
        raise HTTPException(status_code=400, detail="Columna 'codigo_barras' es requerida.")

    empresa_id = user.empresa_id
    total_rows = len(df)

    cfg = await db.configuration.find_one({"empresa_id": empresa_id}) or {}
    redondeo = cfg.get("redondeo_precio", 100)

    def _redondear(valor: float) -> float:
        import math
        if not redondeo:
            return round(valor, 2)
        return math.ceil(valor / redondeo) * redondeo

    async def generate():
        updated = 0
        skipped = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                raw_barcode = row.get("codigo_barras")
                if pd.isna(raw_barcode) or str(raw_barcode).strip() in ("", "nan"):
                    skipped += 1
                    yield f"data: {json.dumps({'progress': int(((idx+1)/total_rows)*100), 'processed': idx+1, 'total': total_rows})}\n\n"
                    await asyncio.sleep(0)
                    continue

                try:
                    codigo_barras = str(int(float(str(raw_barcode).strip())))
                except (ValueError, OverflowError):
                    codigo_barras = str(raw_barcode).strip()

                product = await db.products.find_one({"codigo_barras": codigo_barras, "empresa_id": empresa_id})
                if not product:
                    errors.append(f"Fila {idx+2}: código '{codigo_barras}' no encontrado.")
                    yield f"data: {json.dumps({'progress': int(((idx+1)/total_rows)*100), 'processed': idx+1, 'total': total_rows})}\n\n"
                    await asyncio.sleep(0)
                    continue

                bp_update = {}

                raw_precio = row.get("precio_venta")
                if pd.notna(raw_precio) and str(raw_precio).strip() not in ("", "nan"):
                    bp_update["precio"] = _redondear(float(raw_precio))

                raw_costo = row.get("precio_costo")
                precio_costo = None
                if pd.notna(raw_costo) and str(raw_costo).strip() not in ("", "nan"):
                    precio_costo = float(raw_costo)
                    bp_update["costo"] = precio_costo

                precio_venta_final = bp_update.get("precio")
                if precio_costo is not None and precio_venta_final is not None and precio_costo > 0:
                    bp_update["margen"] = round((precio_venta_final - precio_costo) / precio_costo * 100, 2)
                elif precio_costo is not None:
                    # Calcular margen con el precio actual del branch_product
                    bp_existing = await db.branch_products.find_one({"product_id": product["id"], "branch_id": branch_id, "empresa_id": empresa_id})
                    precio_actual = (bp_existing or {}).get("precio", product.get("precio", 0))
                    if precio_actual and precio_costo > 0:
                        bp_update["margen"] = round((precio_actual - precio_costo) / precio_costo * 100, 2)

                raw_stock = row.get("stock")
                if pd.notna(raw_stock) and str(raw_stock).strip() not in ("", "nan"):
                    bp_update["stock"] = int(float(raw_stock))

                if not bp_update:
                    skipped += 1
                    yield f"data: {json.dumps({'progress': int(((idx+1)/total_rows)*100), 'processed': idx+1, 'total': total_rows})}\n\n"
                    await asyncio.sleep(0)
                    continue

                existing_bp = await db.branch_products.find_one({"product_id": product["id"], "branch_id": branch_id, "empresa_id": empresa_id})
                if existing_bp:
                    await db.branch_products.update_one(
                        {"id": existing_bp["id"], "empresa_id": empresa_id},
                        {"$set": bp_update}
                    )
                else:
                    new_bp = BranchProduct(
                        empresa_id=empresa_id,
                        product_id=product["id"],
                        branch_id=branch_id,
                        precio=bp_update.get("precio", product.get("precio", 0)),
                        stock=bp_update.get("stock", 0),
                        costo=bp_update.get("costo"),
                        margen=bp_update.get("margen"),
                    )
                    await db.branch_products.insert_one(new_bp.dict())

                updated += 1

            except Exception as e:
                errors.append(f"Fila {idx+2}: {str(e)}")

            yield f"data: {json.dumps({'progress': int(((idx+1)/total_rows)*100), 'processed': idx+1, 'total': total_rows})}\n\n"
            await asyncio.sleep(0)

        yield f"data: {json.dumps({'done': True, 'updated': updated, 'skipped': skipped, 'errors': errors, 'total_procesado': updated + skipped + len(errors)})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, user: User = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id, "empresa_id": user.empresa_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return Product(**product)

@api_router.get("/products/barcode/{barcode}", response_model=Product)
async def get_product_by_barcode(barcode: str, user: User = Depends(get_current_user)):
    product = await db.products.find_one({"codigo_barras": barcode, "empresa_id": user.empresa_id, "activo": True})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return Product(**product)

@api_router.put("/products/bulk-control-stock")
async def bulk_update_control_stock(data: BulkUpdateControlStockRequest, user: User = Depends(require_role([UserRole.ADMIN]))):
    if data.delete_all:
        query = {"empresa_id": user.empresa_id}
        if data.search:
            regex = {"$regex": re.escape(data.search), "$options": "i"}
            query["$or"] = [{"nombre": regex}, {"codigo_barras": regex}]
        products = await db.products.find(query, {"id": 1}).to_list(None)
        ids = [p["id"] for p in products]
    else:
        ids = data.ids or []
    if not ids:
        return {"updated": 0}
    result = await db.products.update_many(
        {"id": {"$in": ids}, "empresa_id": user.empresa_id, "kind": {"$ne": "combo"}},
        {"$set": {"control_stock": data.control_stock}}
    )
    return {"updated": result.modified_count}

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product_data: ProductUpdate, user: User = Depends(require_role([UserRole.ADMIN]))):
    # Check if product exists for this empresa
    existing_product = await db.products.find_one({"id": product_id, "empresa_id": user.empresa_id})
    if not existing_product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Update fields (handle None-exclusive fields carefully)
    update_data = {k: v for k, v in product_data.dict().items() if v is not None}
    precio_costo = update_data.pop("precio_costo", None)
    # Explicitly include boolean and list fields even when falsy
    if product_data.control_stock is not None:
        update_data['control_stock'] = product_data.control_stock
    if product_data.activo is not None:
        update_data['activo'] = product_data.activo
    if product_data.combo_items is not None:
        update_data['combo_items'] = [ci.dict() for ci in product_data.combo_items]

    # Determine final kind (could be updated or existing)
    final_kind = update_data.get("kind") or existing_product.get("kind", "normal")
    if final_kind == ProductKind.COMBO:
        update_data["control_stock"] = False

    if update_data:
        await db.products.update_one({"id": product_id, "empresa_id": user.empresa_id}, {"$set": update_data})

    # Update costo/margen in all branch_products if precio_costo provided
    if precio_costo is not None:
        final_precio = update_data.get("precio") or existing_product.get("precio", 0)
        margen = round((final_precio - precio_costo) / precio_costo * 100, 2) if precio_costo else None
        bp_set = {"costo": precio_costo}
        if margen is not None:
            bp_set["margen"] = margen
        await db.branch_products.update_many(
            {"product_id": product_id, "empresa_id": user.empresa_id},
            {"$set": bp_set}
        )

    updated_product = await db.products.find_one({"id": product_id, "empresa_id": user.empresa_id})
    return Product(**updated_product)

@api_router.delete("/products/bulk")
async def bulk_delete_products(data: BulkDeleteRequest, user: User = Depends(require_role([UserRole.ADMIN]))):
    if data.delete_all:
        query = {"empresa_id": user.empresa_id}
        if data.search:
            regex = {"$regex": re.escape(data.search), "$options": "i"}
            query["$or"] = [{"nombre": regex}, {"codigo_barras": regex}]
        products_to_delete = await db.products.find(query, {"id": 1}).to_list(None)
        ids = [p["id"] for p in products_to_delete]
    else:
        ids = data.ids or []
    if not ids:
        return {"deleted": 0}
    result = await db.products.delete_many({"id": {"$in": ids}, "empresa_id": user.empresa_id})
    await db.branch_products.delete_many({"product_id": {"$in": ids}, "empresa_id": user.empresa_id})
    return {"deleted": result.deleted_count}

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, user: User = Depends(require_role([UserRole.ADMIN]))):
    product = await db.products.find_one({"id": product_id, "empresa_id": user.empresa_id})
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    await db.products.delete_one({"id": product_id, "empresa_id": user.empresa_id})
    await db.branch_products.delete_many({"product_id": product_id, "empresa_id": user.empresa_id})
    return {"ok": True}

# ===== CUSTOMERS =====

class Customer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nombre: str
    tipo_documento: Optional[str] = "dni"
    documento: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    direccion: Optional[str] = None
    fecha_nacimiento: Optional[str] = None
    observaciones: Optional[str] = None
    activo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    nombre: str
    tipo_documento: Optional[str] = "dni"
    documento: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    direccion: Optional[str] = None
    fecha_nacimiento: Optional[str] = None
    observaciones: Optional[str] = None
    activo: bool = True

class CustomerUpdate(BaseModel):
    nombre: Optional[str] = None
    tipo_documento: Optional[str] = None
    documento: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    direccion: Optional[str] = None
    fecha_nacimiento: Optional[str] = None
    observaciones: Optional[str] = None
    activo: Optional[bool] = None

class CustomerBulkDeleteRequest(BaseModel):
    ids: Optional[List[str]] = None
    delete_all: bool = False
    search: Optional[str] = None

class CustomerBulkUpdateRequest(BaseModel):
    ids: Optional[List[str]] = None
    delete_all: bool = False
    search: Optional[str] = None
    activo: bool

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR, UserRole.CAJERO]))):
    customer = Customer(**customer_data.dict(), empresa_id=user.empresa_id)
    await db.customers.insert_one(customer.model_dump())
    return customer

@api_router.get("/customers")
async def get_customers(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=10000),
    search: Optional[str] = Query(None),
    activo: Optional[bool] = Query(None),
    user: User = Depends(get_current_user)
):
    query = {"empresa_id": user.empresa_id}
    if activo is not None:
        query["activo"] = activo
    if search:
        regex = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [
            {"nombre": regex},
            {"documento": regex},
            {"email": regex},
            {"telefono": regex},
        ]
    total = await db.customers.count_documents(query)
    skip = (page - 1) * per_page
    raw = await db.customers.find(query).sort("nombre", 1).skip(skip).limit(per_page).to_list(per_page)
    result = []
    for c in raw:
        try:
            result.append(Customer(**c))
        except Exception as e:
            logger.warning(f"Skipping invalid customer {c.get('id')}: {e}")
    return {
        "items": [c.dict() for c in result],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, -(-total // per_page))
    }

@api_router.get("/customers/check-documento")
async def check_documento_existente(
    tipo_documento: str,
    documento: str,
    exclude_id: Optional[str] = None,
    user: User = Depends(get_current_user),
):
    tipo = tipo_documento.lower()
    clean = re.sub(r'\D', '', documento)
    base_q = {"empresa_id": user.empresa_id}
    if exclude_id:
        base_q["id"] = {"$ne": exclude_id}

    CUSTOMER_PROJ = {"_id": 0, "id": 1, "nombre": 1, "tipo_documento": 1, "documento": 1,
                     "telefono": 1, "email": 1, "direccion": 1, "activo": 1,
                     "fecha_nacimiento": 1, "observaciones": 1}

    def _hit(doc): return {"existe": True, **{k: doc.get(k) for k in CUSTOMER_PROJ if k != "_id"}}

    # 1. Coincidencia exacta (mismo tipo + mismo valor)
    exact = await db.customers.find_one(
        {**base_q, "tipo_documento": tipo, "documento": documento}, CUSTOMER_PROJ,
    )
    if exact:
        return _hit(exact)

    # Helper: strip guiones/espacios del documento almacenado
    strip_expr = {
        "$replaceAll": {
            "input": {"$replaceAll": {"input": "$documento", "find": "-", "replacement": ""}},
            "find": " ", "replacement": "",
        }
    }

    # 2. DNI → busca CUIT/CUIL que contengan ese DNI en las posiciones 2-9
    if tipo == "dni" and len(clean) >= 7:
        dni_padded = clean.zfill(8)
        pipeline = [
            {"$match": {**base_q, "tipo_documento": {"$in": ["cuit", "cuil"]}}},
            {"$addFields": {"doc_digits": strip_expr}},
            {"$match": {"$expr": {"$eq": [{"$substr": ["$doc_digits", 2, 8]}, dni_padded]}}},
            {"$limit": 1},
            {"$project": CUSTOMER_PROJ},
        ]
        rows = await db.customers.aggregate(pipeline).to_list(1)
        if rows:
            return _hit(rows[0])

    # 3. CUIT/CUIL → extrae el DNI de las posiciones 2-9 y busca clientes con ese DNI
    elif tipo in ("cuit", "cuil") and len(clean) == 11:
        dni_from_cuit = clean[2:10]
        dni_no_pad    = str(int(dni_from_cuit))
        pipeline = [
            {"$match": {**base_q, "tipo_documento": "dni"}},
            {"$addFields": {"doc_digits": strip_expr}},
            {"$match": {"$or": [{"doc_digits": dni_from_cuit}, {"doc_digits": dni_no_pad}]}},
            {"$limit": 1},
            {"$project": CUSTOMER_PROJ},
        ]
        rows = await db.customers.aggregate(pipeline).to_list(1)
        if rows:
            return _hit(rows[0])

    return {"existe": False}

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, user: User = Depends(get_current_user)):
    c = await db.customers.find_one({"id": customer_id, "empresa_id": user.empresa_id})
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return Customer(**c)

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer_data: CustomerUpdate, user: User = Depends(require_role([UserRole.ADMIN]))):
    c = await db.customers.find_one({"id": customer_id, "empresa_id": user.empresa_id})
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    update_data = {k: v for k, v in customer_data.dict().items() if v is not None}
    if update_data:
        await db.customers.update_one({"id": customer_id, "empresa_id": user.empresa_id}, {"$set": update_data})
    updated = await db.customers.find_one({"id": customer_id, "empresa_id": user.empresa_id})
    return Customer(**updated)

@api_router.put("/customers/bulk-status")
async def bulk_update_customer_status(data: CustomerBulkUpdateRequest, user: User = Depends(require_role([UserRole.ADMIN]))):
    query = {"empresa_id": user.empresa_id}
    if not data.delete_all:
        if not data.ids:
            raise HTTPException(status_code=400, detail="No se especificaron IDs")
        query["id"] = {"$in": data.ids}
    elif data.search:
        regex = {"$regex": re.escape(data.search), "$options": "i"}
        query["$or"] = [{"nombre": regex}, {"documento": regex}, {"email": regex}, {"telefono": regex}]
    result = await db.customers.update_many(query, {"$set": {"activo": data.activo}})
    return {"updated": result.modified_count}

@api_router.delete("/customers/bulk")
async def bulk_delete_customers(data: CustomerBulkDeleteRequest, user: User = Depends(require_role([UserRole.ADMIN]))):
    query = {"empresa_id": user.empresa_id}
    if not data.delete_all:
        if not data.ids:
            raise HTTPException(status_code=400, detail="No se especificaron IDs")
        query["id"] = {"$in": data.ids}
    elif data.search:
        regex = {"$regex": re.escape(data.search), "$options": "i"}
        query["$or"] = [{"nombre": regex}, {"documento": regex}, {"email": regex}, {"telefono": regex}]
    result = await db.customers.delete_many(query)
    return {"deleted": result.deleted_count}

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, user: User = Depends(require_role([UserRole.ADMIN]))):
    c = await db.customers.find_one({"id": customer_id, "empresa_id": user.empresa_id})
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    await db.customers.delete_one({"id": customer_id, "empresa_id": user.empresa_id})
    return {"ok": True}

async def _consultar_cuit_cuitonline(cuit_clean: str) -> dict:
    """Scraper de cuitonline.com como fallback cuando ARCA SOAP no está disponible."""
    async with httpx.AsyncClient(timeout=10, follow_redirects=True) as client:
        resp = await client.get(
            f"https://www.cuitonline.com/search.php?q={cuit_clean}",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                     "Accept": "text/html,application/xhtml+xml"},
        )
        if resp.status_code != 200:
            raise RuntimeError(f"cuitonline.com devolvió {resp.status_code}")
        html = resp.text

    # Extraer nombre del h2.denominacion
    m = re.search(r'<h2[^>]+class="denominacion"[^>]*>(.*?)</h2>', html, re.IGNORECASE | re.DOTALL)
    if not m:
        raise LookupError("CUIT no encontrado")
    nombre = re.sub(r'<[^>]+>', '', m.group(1)).strip()
    if not nombre:
        raise LookupError("CUIT no encontrado")

    # Tipo de persona
    tipo = "JURIDICA" if "persona jur" in html.lower() else "FISICA"

    return {"nombre": nombre, "tipo_persona": tipo, "estado_clave": "", "direccion": None}


@api_router.get("/afip/consultar-cuit/{cuit}")
async def consultar_cuit_afip(cuit: str, user: User = Depends(get_current_user)):
    cuit_clean = re.sub(r'\D', '', cuit)
    if len(cuit_clean) != 11:
        raise HTTPException(status_code=400, detail="CUIT inválido, debe tener 11 dígitos")
    try:
        return await _consultar_cuit_cuitonline(cuit_clean)
    except LookupError:
        raise HTTPException(status_code=404, detail="CUIT no encontrado en el padrón de AFIP")
    except Exception as e:
        logger.warning(f"cuitonline.com falló para CUIT {cuit_clean}: {e}")
        raise HTTPException(status_code=503, detail="Servicio de consulta AFIP no disponible en este momento")

# User management routes
@api_router.get("/users", response_model=List[User])
async def get_users(user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))):
    users = await db.users.find({"empresa_id": user.empresa_id}).to_list(1000)
    return [User(**u) for u in users]

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_data: UserUpdate, current_user: User = Depends(require_role([UserRole.ADMIN]))):
    user = await db.users.find_one({"id": user_id, "empresa_id": current_user.empresa_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    update_data = {k: v for k, v in user_data.dict().items() if v is not None}
    # Allow explicitly setting branch_ids to empty list (removing all branch assignments)
    if user_data.branch_ids is not None:
        update_data["branch_ids"] = user_data.branch_ids
    if update_data:
        await db.users.update_one({"id": user_id, "empresa_id": current_user.empresa_id}, {"$set": update_data})
    updated = await db.users.find_one({"id": user_id, "empresa_id": current_user.empresa_id})
    return User(**updated)

# Configuration routes
@api_router.get("/config", response_model=Configuration)
async def get_configuration(user: User = Depends(get_current_user)):
    config = await db.configuration.find_one({"empresa_id": user.empresa_id})
    if not config:
        # Create default configuration if none exists for this empresa
        default_config = Configuration(empresa_id=user.empresa_id)
        await db.configuration.insert_one(default_config.dict())
        return default_config
    return Configuration(**config)

@api_router.put("/config", response_model=Configuration)
async def update_configuration(config_data: ConfigurationUpdate, user: User = Depends(require_role([UserRole.ADMIN]))):
    # Get current configuration for this empresa
    current_config = await db.configuration.find_one({"empresa_id": user.empresa_id})
    if not current_config:
        current_config = Configuration(empresa_id=user.empresa_id).dict()
        await db.configuration.insert_one(current_config)

    # Update fields (allow company_logo to be explicitly set to null)
    raw = config_data.dict(exclude_unset=True)
    update_data = {k: v for k, v in raw.items() if v is not None or k == 'company_logo'}
    update_data['updated_at'] = datetime.now(timezone.utc)

    if update_data:
        await db.configuration.update_one({"empresa_id": user.empresa_id}, {"$set": update_data})

    updated_config = await db.configuration.find_one({"empresa_id": user.empresa_id})
    return Configuration(**updated_config)

@api_router.post("/config/upload-logo")
async def upload_logo(file: UploadFile = File(...), user: User = Depends(require_role([UserRole.ADMIN]))):
    # Validate file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")

    # Validate file size (max 2MB)
    content = await file.read()
    file_size = len(content)

    if file_size > 2 * 1024 * 1024:  # 2MB
        raise HTTPException(status_code=400, detail="File size must be less than 2MB")

    # Convert to base64
    base64_image = base64.b64encode(content).decode('utf-8')
    logo_data_url = f"data:{file.content_type};base64,{base64_image}"

    # Update configuration for this empresa
    await db.configuration.update_one({"empresa_id": user.empresa_id}, {"$set": {
        "company_logo": logo_data_url,
        "updated_at": datetime.now(timezone.utc)
    }})

    return {"message": "Logo uploaded successfully", "logo_url": logo_data_url}

# ─── Helpers de facturación ───────────────────────────────────────────────────

# Tipos de comprobante soportados (sin notas de crédito)
TIPOS_COMPROBANTE_FACTURA = {1, 6, 11, 2, 7}   # A, B, C, NDebA, NDebB

# Alícuotas IVA AFIP: tax_rate (decimal) → id interno WSFE
_IVA_ALICUOTA = [(0.27, 6), (0.21, 5), (0.105, 4), (0.05, 8), (0.025, 9), (0.0, 3)]

def prepare_invoice_payload(
    sale_dict: dict,
    afip_cfg: dict,
    tipo_cbte: int,
    tax_rate: float = 0.0,
) -> dict:
    """
    Arma la estructura mínima de datos que necesita cualquier
    librería AFIP/ARCA (ej. pyafipws, zeep, ARCA REST) para
    autorizar un comprobante electrónico (FECAESolicitar / fe_comp_consultar).

    No llama al WS — sólo prepara y devuelve el payload estructurado
    para su uso posterior o para almacenamiento en contingencia.
    """
    from datetime import datetime as _dt

    # Alícuota IVA más cercana al tax_rate configurado
    alicuota_id = min(_IVA_ALICUOTA, key=lambda x: abs(x[0] - tax_rate))[1]

    # Condición IVA receptor → DocTipo AFIP
    condicion_raw = (sale_dict.get("condicion_iva_receptor") or "consumidor_final").lower()
    if any(k in condicion_raw for k in ("responsable_inscripto", " ri", "monotributista", " mo", "exento", " ex")):
        doc_tipo = 80   # CUIT
        cuit_raw = str(sale_dict.get("cuit_receptor") or "").replace("-", "").replace(" ", "")
        try:
            doc_nro = int(cuit_raw) if cuit_raw.isdigit() else 0
        except (ValueError, TypeError):
            doc_nro = 0
    else:
        doc_tipo = 99   # Consumidor Final
        doc_nro = 0

    subtotal       = float(sale_dict.get("subtotal", 0))
    descuento      = float(sale_dict.get("descuento", 0))
    imp_trib       = float(sale_dict.get("impuestos_extra_total", 0))
    imp_total      = float(sale_dict.get("total", 0))

    # Neto gravado (base IVA) = subtotal menos descuento proporcional
    imp_neto = max(0.0, round(subtotal - descuento, 2))
    imp_iva  = round(imp_neto * tax_rate, 2)

    iva_detalle = []
    if alicuota_id != 3 and imp_iva > 0:
        iva_detalle = [{"id": alicuota_id, "base_imp": imp_neto, "importe": imp_iva}]

    items_payload = []
    for it in sale_dict.get("items", []):
        qty   = float(it.get("cantidad", 1))
        price = float(it.get("precio_unitario", 0))
        sub   = float(it.get("subtotal") or price * qty)
        items_payload.append({
            "descripcion":      it.get("nombre") or it.get("producto_id", ""),
            "cantidad":         qty,
            "precio_unitario":  round(price, 2),
            "subtotal_neto":    round(sub, 2),
            "imp_iva":          round(sub * tax_rate, 2),
            "alicuota_iva_id":  alicuota_id,
        })

    return {
        # ── Emisor ──────────────────────────────────────────────
        "cuit_emisor":            afip_cfg.get("cuit", ""),
        "razon_social":           afip_cfg.get("razon_social", ""),
        "domicilio_fiscal":       afip_cfg.get("domicilio_fiscal", ""),
        "condicion_iva_emisor":   afip_cfg.get("condicion_iva_emisor", "RI"),
        "iibb":                   afip_cfg.get("iibb", ""),
        "inicio_actividades":     afip_cfg.get("inicio_actividades", ""),
        "punto_venta":            afip_cfg.get("punto_venta", 1),
        # ── Comprobante ─────────────────────────────────────────
        "tipo_cbte":  tipo_cbte,         # 1=FactA 6=FactB 11=FactC 2=NDebA 7=NDebB
        "fecha_cbte": _dt.now().strftime("%Y%m%d"),
        "concepto":   afip_cfg.get("concepto_default", 1),  # 1=Productos 2=Servicios 3=Ambos
        # ── Receptor ────────────────────────────────────────────
        "tipo_doc":              doc_tipo,     # 80=CUIT 99=ConsumidorFinal
        "nro_doc":               doc_nro,
        "condicion_iva_receptor": condicion_raw,
        "cuit_receptor":          sale_dict.get("cuit_receptor"),
        # ── Moneda ──────────────────────────────────────────────
        "mon_id":    "PES",
        "mon_cotiz": 1.0,
        # ── Importes ────────────────────────────────────────────
        "imp_neto":      imp_neto,      # neto gravado (base IVA)
        "imp_iva":       imp_iva,       # total IVA calculado sobre neto
        "imp_trib":      round(imp_trib, 2),   # otros tributos / percepciones
        "imp_tot_conc":  0.0,           # no gravado
        "imp_op_ex":     0.0,           # operaciones exentas
        "imp_total":     round(imp_total, 2),
        "descuento":     round(descuento, 2),
        # ── IVA detalle (requerido por WSFE) ────────────────────
        "iva": iva_detalle,
        # ── Items (para PDF / display) ──────────────────────────
        "items": items_payload,
    }


# Sales routes
@api_router.post("/sales", response_model=Sale)
async def create_sale(sale_data: SaleCreate, user: User = Depends(get_current_user)):
    # Check if user has an open cash session for the active branch
    if not user.branch_id:
        raise HTTPException(status_code=400, detail="Debe seleccionar una sucursal antes de realizar ventas")
    current_session = await db.cash_sessions.find_one({
        "user_id": user.id,
        "empresa_id": user.empresa_id,
        "branch_id": user.branch_id,
        "status": CashSessionStatus.ABIERTA
    })
    if not current_session:
        raise HTTPException(status_code=400, detail="Debe abrir una caja en esta sucursal antes de realizar ventas")

    # Get current configuration for tax rate
    config = await db.configuration.find_one({"empresa_id": user.empresa_id})
    tax_rate = config['tax_rate'] if config else 0.12  # Default 12%
    auto_update_inventory = config.get('auto_update_inventory', True) if config else True

    # Verify products and calculate totals
    total_amount = 0
    validated_items = []

    for item in sale_data.items:
        precio_unitario = None
        product_nombre = None

        # Always fetch global product to know kind/control_stock
        global_product = await db.products.find_one({"id": item.producto_id, "empresa_id": user.empresa_id, "activo": True})
        if not global_product:
            raise HTTPException(status_code=400, detail="Producto no encontrado")

        product_kind = global_product.get('kind', 'normal')
        control_stock = global_product.get('control_stock', True)

        if product_kind == 'combo':
            # Resolve price from branch or global
            if user.branch_id:
                branch_product = await db.branch_products.find_one({
                    "product_id": item.producto_id,
                    "branch_id": user.branch_id,
                    "empresa_id": user.empresa_id,
                    "activo": True
                })
                if branch_product:
                    precio_unitario = branch_product.get('precio_por_peso') or branch_product['precio']
            if precio_unitario is None:
                precio_unitario = global_product.get('precio_por_peso') or global_product['precio']
            product_nombre = global_product['nombre']

            # Deduct stock from each combo component
            combo_items = global_product.get('combo_items', [])
            for ci in combo_items:
                comp = await db.products.find_one({"id": ci['product_id'], "empresa_id": user.empresa_id})
                comp_control = comp.get('control_stock', True) if comp else False
                if not comp or not comp_control or not auto_update_inventory:
                    continue
                comp_cantidad = ci['cantidad'] * item.cantidad
                deducted = False
                if user.branch_id:
                    comp_bp = await db.branch_products.find_one({
                        "product_id": ci['product_id'],
                        "branch_id": user.branch_id,
                        "empresa_id": user.empresa_id,
                        "activo": True
                    })
                    if comp_bp:
                        if comp_bp.get('stock', 0) < comp_cantidad:
                            raise HTTPException(status_code=400, detail=f"Stock insuficiente para {comp['nombre']} (componente del combo)")
                        await db.branch_products.update_one(
                            {"product_id": ci['product_id'], "branch_id": user.branch_id, "empresa_id": user.empresa_id},
                            {"$inc": {"stock": -int(comp_cantidad)}}
                        )
                        deducted = True
                if not deducted:
                    if comp.get('stock', 0) < comp_cantidad:
                        raise HTTPException(status_code=400, detail=f"Stock insuficiente para {comp['nombre']} (componente del combo)")
                    await db.products.update_one(
                        {"id": ci['product_id'], "empresa_id": user.empresa_id},
                        {"$inc": {"stock": -int(comp_cantidad)}}
                    )

        else:
            # Normal product: try branch first
            manage_stock = control_stock and auto_update_inventory
            if user.branch_id:
                branch_product = await db.branch_products.find_one({
                    "product_id": item.producto_id,
                    "branch_id": user.branch_id,
                    "empresa_id": user.empresa_id,
                    "activo": True
                })
                if branch_product:
                    if manage_stock and branch_product.get('stock', 0) < item.cantidad:
                        raise HTTPException(status_code=400, detail=f"Stock insuficiente para {global_product['nombre']}")
                    precio_unitario = branch_product.get('precio_por_peso') or branch_product.get('precio')
                    if manage_stock:
                        await db.branch_products.update_one(
                            {"product_id": item.producto_id, "branch_id": user.branch_id, "empresa_id": user.empresa_id},
                            {"$inc": {"stock": -int(item.cantidad)}}
                        )
                    product_nombre = global_product['nombre']

            # Fall back to global product
            if precio_unitario is None:
                if manage_stock and global_product.get('stock', 0) < item.cantidad:
                    raise HTTPException(status_code=400, detail=f"Stock insuficiente para {global_product['nombre']}")
                precio_unitario = global_product.get('precio_por_peso') or global_product.get('precio')
                product_nombre = global_product['nombre']
                if manage_stock:
                    await db.products.update_one(
                        {"id": item.producto_id, "empresa_id": user.empresa_id},
                        {"$inc": {"stock": -int(item.cantidad)}}
                    )

        # Calculate subtotal
        if precio_unitario is None:
            raise HTTPException(status_code=400, detail=f"No se pudo determinar el precio para {product_nombre or item.producto_id}")
        subtotal = item.cantidad * precio_unitario
        validated_items.append(SaleItem(
            producto_id=item.producto_id,
            nombre=product_nombre,
            cantidad=item.cantidad,
            precio_unitario=precio_unitario,
            subtotal=subtotal
        ))
        total_amount += subtotal

    # Generate invoice number
    branch_id_for_sale = current_session.get('branch_id') or user.branch_id
    last_sale = await db.sales.find_one(
        {"branch_id": branch_id_for_sale, "empresa_id": user.empresa_id},
        sort=[("fecha", -1)]
    )
    if last_sale and last_sale.get('numero_factura'):
        last_number = int(last_sale['numero_factura'].split('-')[-1])
        numero_factura = f"FAC-{last_number + 1:06d}"
    else:
        numero_factura = "FAC-000001"

    # Calculate totals
    subtotal = total_amount
    impuestos = total_amount * tax_rate
    base_total = total_amount * (1 + tax_rate)
    # Apply payment method adjustment
    adjustments = (config.get('payment_method_adjustments') or {}) if config else {}
    adjustment_pct = adjustments.get(sale_data.metodo_pago.value, 0.0)
    adjustment_amount = base_total * (adjustment_pct / 100.0)
    descuento = sale_data.descuento or 0.0
    impuestos_extra_total = sale_data.impuestos_extra_total or 0.0
    total = base_total + adjustment_amount - descuento + impuestos_extra_total

    # Create sale
    sale = Sale(
        empresa_id=user.empresa_id,
        cajero_id=user.id,
        branch_id=branch_id_for_sale,
        session_id=current_session['id'],
        items=validated_items,
        subtotal=subtotal,
        impuestos=impuestos,
        total=total,
        metodo_pago=sale_data.metodo_pago,
        numero_factura=numero_factura,
        cliente_id=sale_data.cliente_id,
        descuento=descuento,
        impuestos_extra_total=impuestos_extra_total,
        condicion_iva_receptor=sale_data.condicion_iva_receptor,
        observaciones_comprobante=sale_data.observaciones_comprobante,
        tipo_comprobante=sale_data.tipo_comprobante,
    )

    await db.sales.insert_one(sale.dict())

    # Update cash session
    await db.cash_sessions.update_one(
        {"id": current_session['id'], "empresa_id": user.empresa_id},
        {"$inc": {"monto_ventas": total}}
    )

    # Create cash movement
    movement = CashMovement(
        empresa_id=user.empresa_id,
        session_id=current_session['id'],
        tipo=MovementType.VENTA,
        monto=total,
        descripcion=f"Venta {numero_factura} - {sale_data.metodo_pago.value.capitalize()}",
        venta_id=sale.id
    )
    await db.cash_movements.insert_one(movement.dict())

    # ── Facturación electrónica ARCA/AFIP ─────────────────────────────────────
    # Solo se procesa cuando el panel de Factura seleccionó un tipo de comprobante
    # explícito (distinto de "ticket"). tipo_cbte=None significa venta sin comprobante.
    tipo_cbte = sale_data.tipo_comprobante  # None → ticket, no genera comprobante

    if tipo_cbte is not None and tipo_cbte in TIPOS_COMPROBANTE_FACTURA:
        afip_cfg = await db.afip_config.find_one({"empresa_id": user.empresa_id, "activo": True})
        if afip_cfg:
            # Armar payload mínimo (estructura para librería AFIP/ARCA)
            sale_dict_for_invoice = {
                **sale.dict(),
                "condicion_iva_receptor": sale_data.condicion_iva_receptor,
                "cuit_receptor": sale_data.cuit_receptor,
            }
            factura_payload = prepare_invoice_payload(
                sale_dict_for_invoice, afip_cfg, tipo_cbte, tax_rate
            )

            # Persistir payload y marcar tipo de comprobante
            await db.sales.update_one({"id": sale.id}, {"$set": {
                "factura_payload":   factura_payload,
                "tipo_comprobante":  tipo_cbte,
                "afip_estado":       "sin_cae",
            }})
            sale.factura_payload  = factura_payload
            sale.tipo_comprobante = tipo_cbte
            sale.afip_estado      = "sin_cae"

            # Si hay certificado disponible → solicitar CAE (contingencia si falla)
            if afip_cfg.get("cert_pem") and afip_cfg.get("key_pem_encrypted"):
                try:
                    token, sign = await _afip_service.get_token_sign(afip_cfg, db, SECRET_KEY)
                    cae_result = await _afip_service.solicitar_cae(
                        sale.dict(), afip_cfg, token, sign, tipo_cbte,
                        cuit_receptor=sale_data.cuit_receptor
                    )
                    afip_update = {
                        "cae":                 cae_result["cae"],
                        "cae_vencimiento":     cae_result["cae_vencimiento"],
                        "afip_estado":         "autorizado",
                        "nro_comprobante_afip": cae_result["nro_comprobante"],
                    }
                    await db.sales.update_one({"id": sale.id}, {"$set": afip_update})
                    sale.cae                 = cae_result["cae"]
                    sale.cae_vencimiento     = cae_result["cae_vencimiento"]
                    sale.afip_estado         = "autorizado"
                    sale.nro_comprobante_afip = cae_result["nro_comprobante"]
                except Exception as afip_err:
                    logger.warning(f"AFIP contingencia para venta {sale.id}: {afip_err}")
                    await db.sales.update_one(
                        {"id": sale.id},
                        {"$set": {"afip_estado": "contingencia", "afip_error": str(afip_err)}}
                    )
                    sale.afip_estado = "contingencia"
                    sale.afip_error  = str(afip_err)

    return sale

@api_router.put("/sales/{sale_id}", response_model=Sale)
async def update_sale(sale_id: str, sale_data: SaleCreate, user: User = Depends(get_current_user)):
    original_sale = await db.sales.find_one({"id": sale_id, "empresa_id": user.empresa_id})
    if not original_sale:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    if original_sale.get('estado') == 'cancelado':
        raise HTTPException(status_code=400, detail="No se puede modificar una venta cancelada")

    config = await db.configuration.find_one({"empresa_id": user.empresa_id})
    tax_rate = config['tax_rate'] if config else 0.12
    auto_update_inventory = config.get('auto_update_inventory', True) if config else True

    # Restore stock from original items
    if auto_update_inventory:
        for orig_item in original_sale.get('items', []):
            prod_id = orig_item['producto_id']
            cantidad = orig_item['cantidad']
            global_prod = await db.products.find_one({"id": prod_id, "empresa_id": user.empresa_id})
            if not global_prod:
                continue
            control_stock = global_prod.get('control_stock', True)
            if not control_stock:
                continue
            product_kind = global_prod.get('kind', 'normal')
            if product_kind == 'combo':
                for ci in global_prod.get('combo_items', []):
                    comp = await db.products.find_one({"id": ci['product_id'], "empresa_id": user.empresa_id})
                    if not comp or not comp.get('control_stock', True):
                        continue
                    comp_cantidad = ci['cantidad'] * cantidad
                    restored = False
                    if user.branch_id:
                        comp_bp = await db.branch_products.find_one({"product_id": ci['product_id'], "branch_id": user.branch_id, "empresa_id": user.empresa_id})
                        if comp_bp:
                            await db.branch_products.update_one({"product_id": ci['product_id'], "branch_id": user.branch_id, "empresa_id": user.empresa_id}, {"$inc": {"stock": int(comp_cantidad)}})
                            restored = True
                    if not restored:
                        await db.products.update_one({"id": ci['product_id'], "empresa_id": user.empresa_id}, {"$inc": {"stock": int(comp_cantidad)}})
            else:
                restored = False
                if user.branch_id:
                    bp = await db.branch_products.find_one({"product_id": prod_id, "branch_id": user.branch_id, "empresa_id": user.empresa_id, "activo": True})
                    if bp:
                        await db.branch_products.update_one({"product_id": prod_id, "branch_id": user.branch_id, "empresa_id": user.empresa_id}, {"$inc": {"stock": int(cantidad)}})
                        restored = True
                if not restored:
                    await db.products.update_one({"id": prod_id, "empresa_id": user.empresa_id}, {"$inc": {"stock": int(cantidad)}})

    # Process new items (same validation + stock deduction as create_sale)
    total_amount = 0
    validated_items = []
    for item in sale_data.items:
        precio_unitario = None
        product_nombre = None
        global_product = await db.products.find_one({"id": item.producto_id, "empresa_id": user.empresa_id, "activo": True})
        if not global_product:
            raise HTTPException(status_code=400, detail="Producto no encontrado")
        product_kind = global_product.get('kind', 'normal')
        control_stock = global_product.get('control_stock', True)
        manage_stock = control_stock and auto_update_inventory

        if product_kind == 'combo':
            if user.branch_id:
                branch_product = await db.branch_products.find_one({"product_id": item.producto_id, "branch_id": user.branch_id, "empresa_id": user.empresa_id, "activo": True})
                if branch_product:
                    precio_unitario = branch_product.get('precio_por_peso') or branch_product['precio']
            if precio_unitario is None:
                precio_unitario = global_product.get('precio_por_peso') or global_product['precio']
            product_nombre = global_product['nombre']
            for ci in global_product.get('combo_items', []):
                comp = await db.products.find_one({"id": ci['product_id'], "empresa_id": user.empresa_id})
                comp_control = comp.get('control_stock', True) if comp else False
                if not comp or not comp_control or not auto_update_inventory:
                    continue
                comp_cantidad = ci['cantidad'] * item.cantidad
                deducted = False
                if user.branch_id:
                    comp_bp = await db.branch_products.find_one({"product_id": ci['product_id'], "branch_id": user.branch_id, "empresa_id": user.empresa_id, "activo": True})
                    if comp_bp:
                        if comp_bp.get('stock', 0) < comp_cantidad:
                            raise HTTPException(status_code=400, detail=f"Stock insuficiente para {comp['nombre']}")
                        await db.branch_products.update_one({"product_id": ci['product_id'], "branch_id": user.branch_id, "empresa_id": user.empresa_id}, {"$inc": {"stock": -int(comp_cantidad)}})
                        deducted = True
                if not deducted:
                    if comp.get('stock', 0) < comp_cantidad:
                        raise HTTPException(status_code=400, detail=f"Stock insuficiente para {comp['nombre']}")
                    await db.products.update_one({"id": ci['product_id'], "empresa_id": user.empresa_id}, {"$inc": {"stock": -int(comp_cantidad)}})
        else:
            if user.branch_id:
                branch_product = await db.branch_products.find_one({"product_id": item.producto_id, "branch_id": user.branch_id, "empresa_id": user.empresa_id, "activo": True})
                if branch_product:
                    if manage_stock and branch_product.get('stock', 0) < item.cantidad:
                        raise HTTPException(status_code=400, detail=f"Stock insuficiente para {global_product['nombre']}")
                    precio_unitario = branch_product.get('precio_por_peso') or branch_product.get('precio')
                    if manage_stock:
                        await db.branch_products.update_one({"product_id": item.producto_id, "branch_id": user.branch_id, "empresa_id": user.empresa_id}, {"$inc": {"stock": -int(item.cantidad)}})
                    product_nombre = global_product['nombre']
            if precio_unitario is None:
                if manage_stock and global_product.get('stock', 0) < item.cantidad:
                    raise HTTPException(status_code=400, detail=f"Stock insuficiente para {global_product['nombre']}")
                precio_unitario = global_product.get('precio_por_peso') or global_product.get('precio')
                product_nombre = global_product['nombre']
                if manage_stock:
                    await db.products.update_one({"id": item.producto_id, "empresa_id": user.empresa_id}, {"$inc": {"stock": -int(item.cantidad)}})

        if precio_unitario is None:
            raise HTTPException(status_code=400, detail=f"No se pudo determinar el precio para {product_nombre or item.producto_id}")
        subtotal_item = item.cantidad * precio_unitario
        validated_items.append(SaleItem(producto_id=item.producto_id, nombre=product_nombre, cantidad=item.cantidad, precio_unitario=precio_unitario, subtotal=subtotal_item))
        total_amount += subtotal_item

    # Recalculate totals
    subtotal = total_amount
    impuestos = total_amount * tax_rate
    base_total = total_amount * (1 + tax_rate)
    adjustments = (config.get('payment_method_adjustments') or {}) if config else {}
    adjustment_pct = adjustments.get(sale_data.metodo_pago.value, 0.0)
    adjustment_amount = base_total * (adjustment_pct / 100.0)
    descuento = sale_data.descuento or 0.0
    impuestos_extra_total = sale_data.impuestos_extra_total or 0.0
    total = base_total + adjustment_amount - descuento + impuestos_extra_total

    # Update the sale
    old_total = original_sale.get('total', 0)
    update_fields = {
        "items": [i.dict() for i in validated_items],
        "subtotal": subtotal,
        "impuestos": impuestos,
        "total": total,
        "metodo_pago": sale_data.metodo_pago.value,
        "cliente_id": sale_data.cliente_id,
        "descuento": descuento,
        "impuestos_extra_total": impuestos_extra_total,
        "condicion_iva_receptor": sale_data.condicion_iva_receptor,
        "observaciones_comprobante": sale_data.observaciones_comprobante,
        "fecha_modificacion": datetime.now(timezone.utc).isoformat(),
    }
    await db.sales.update_one({"id": sale_id}, {"$set": update_fields})

    # Adjust cash session total
    diff = total - old_total
    if diff != 0:
        await db.cash_sessions.update_one({"id": original_sale['session_id']}, {"$inc": {"monto_ventas": diff}})

    updated = await db.sales.find_one({"id": sale_id})
    return Sale(**updated)


@api_router.get("/sales", response_model=List[Sale])
async def get_sales(customer_id: Optional[str] = None, user: User = Depends(get_current_user)):
    query: dict = {"empresa_id": user.empresa_id}
    if user.rol == UserRole.CAJERO:
        query["cajero_id"] = user.id
    if customer_id:
        query["cliente_id"] = customer_id
    limit = 500 if customer_id else (100 if user.rol == UserRole.CAJERO else 1000)
    sales = await db.sales.find(query).sort("fecha", -1).to_list(limit)

    result = []
    for sale in sales:
        try:
            result.append(Sale(**sale))
        except Exception as e:
            logger.warning(f"Skipping invalid sale {sale.get('id')}: {e}")
    return result

@api_router.get("/returns", response_model=List[SaleReturn])
async def get_all_returns(user: User = Depends(get_current_user)):
    returns = await db.sale_returns.find({"empresa_id": user.empresa_id}).sort("fecha", -1).to_list(10000)
    return [SaleReturn(**r) for r in returns]

@api_router.get("/sales/{sale_id}/returns", response_model=List[SaleReturn])
async def get_sale_returns(sale_id: str, user: User = Depends(get_current_user)):
    returns = await db.sale_returns.find({"sale_id": sale_id, "empresa_id": user.empresa_id}).sort("fecha", -1).to_list(100)
    return [SaleReturn(**r) for r in returns]

@api_router.get("/credit-notes", response_model=List[CreditNote])
async def get_all_credit_notes(user: User = Depends(get_current_user)):
    notes = await db.credit_notes.find({"empresa_id": user.empresa_id}).sort("fecha", -1).to_list(10000)
    return [CreditNote(**n) for n in notes]

@api_router.get("/sales/{sale_id}/credit-notes", response_model=List[CreditNote])
async def get_sale_credit_notes(sale_id: str, user: User = Depends(get_current_user)):
    notes = await db.credit_notes.find({"sale_id": sale_id, "empresa_id": user.empresa_id}).sort("fecha", -1).to_list(100)
    return [CreditNote(**n) for n in notes]

@api_router.get("/reportes/margenes")
async def get_reporte_margenes(
    fecha_desde: str = Query(...),
    fecha_hasta: str = Query(...),
    branch_id: Optional[str] = Query(None),
    user: User = Depends(get_current_user)
):
    AR_TZ = timezone(timedelta(hours=-3))

    # Convertir fechas a rango UTC
    desde_local = datetime.strptime(fecha_desde, "%Y-%m-%d").replace(
        hour=0, minute=0, second=0, tzinfo=AR_TZ
    )
    hasta_local = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(
        hour=23, minute=59, second=59, tzinfo=AR_TZ
    )
    desde_utc = desde_local.astimezone(timezone.utc)
    hasta_utc = hasta_local.astimezone(timezone.utc)

    filtro = {
        "empresa_id": user.empresa_id,
        "fecha": {"$gte": desde_utc, "$lte": hasta_utc},
        "estado": {"$ne": "cancelado"},
    }
    if branch_id and branch_id != "all":
        filtro["branch_id"] = branch_id

    sales = await db.sales.find(filtro).sort("fecha", 1).to_list(100000)

    # Cargar costos de branch_products en batch: {(product_id, branch_id): costo}
    product_ids = list({item["producto_id"] for s in sales for item in s.get("items", [])})
    branch_ids_needed = list({s.get("branch_id") for s in sales if s.get("branch_id")})
    bp_cursor = await db.branch_products.find(
        {"product_id": {"$in": product_ids}, "branch_id": {"$in": branch_ids_needed}, "empresa_id": user.empresa_id},
        {"product_id": 1, "branch_id": 1, "costo": 1}
    ).to_list(100000)
    costo_map = {(bp["product_id"], bp["branch_id"]): bp.get("costo") for bp in bp_cursor}

    # Cargar nombres de sucursales
    branches_list = await db.branches.find({"empresa_id": user.empresa_id}, {"id": 1, "nombre": 1}).to_list(1000)
    branch_name_map = {b["id"]: b["nombre"] for b in branches_list}

    DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    por_fecha: dict = {}
    por_dia_semana: dict = {}
    por_sucursal_agg: dict = {}
    por_venta_list = []
    total_ventas = 0.0
    total_costo = 0.0
    total_margen = 0.0
    num_ventas = 0

    for s in sales:
        fecha_dt = s.get("fecha")
        if isinstance(fecha_dt, str):
            fecha_dt = datetime.fromisoformat(fecha_dt.replace("Z", "+00:00"))
        if fecha_dt.tzinfo is None:
            fecha_dt = fecha_dt.replace(tzinfo=timezone.utc)
        fecha_local = fecha_dt.astimezone(AR_TZ)
        fecha_dia = fecha_local.strftime("%Y-%m-%d")
        dia_numero = fecha_local.weekday()  # 0=lunes

        sale_ventas = 0.0
        sale_costo = 0.0
        items_detalle = []

        for item in s.get("items", []):
            pid = item.get("producto_id")
            bid = s.get("branch_id")
            qty = float(item.get("cantidad", 0))
            precio = float(item.get("precio_unitario", 0))
            subtotal_venta = float(item.get("subtotal") or item.get("total") or (qty * precio))
            costo_unit = costo_map.get((pid, bid))
            subtotal_costo = qty * costo_unit if costo_unit is not None else None
            margen_item = (subtotal_venta - subtotal_costo) if subtotal_costo is not None else None
            margen_pct_item = (margen_item / subtotal_costo * 100) if (subtotal_costo and subtotal_costo > 0) else None

            sale_ventas += subtotal_venta
            if subtotal_costo is not None:
                sale_costo += subtotal_costo

            items_detalle.append({
                "nombre": item.get("nombre", ""),
                "cantidad": qty,
                "subtotal_venta": round(subtotal_venta, 2),
                "subtotal_costo": round(subtotal_costo, 2) if subtotal_costo is not None else None,
                "margen": round(margen_item, 2) if margen_item is not None else None,
                "margen_pct": round(margen_pct_item, 2) if margen_pct_item is not None else None,
            })

        sale_margen = sale_ventas - sale_costo
        sale_margen_pct = (sale_margen / sale_costo * 100) if sale_costo > 0 else None

        # Acumulado global
        total_ventas += sale_ventas
        total_costo += sale_costo
        total_margen += sale_margen
        num_ventas += 1

        # Por fecha
        if fecha_dia not in por_fecha:
            por_fecha[fecha_dia] = {"ventas_total": 0.0, "costo_total": 0.0, "margen_total": 0.0, "num_ventas": 0}
        por_fecha[fecha_dia]["ventas_total"] += sale_ventas
        por_fecha[fecha_dia]["costo_total"] += sale_costo
        por_fecha[fecha_dia]["margen_total"] += sale_margen
        por_fecha[fecha_dia]["num_ventas"] += 1

        # Por día de semana
        if dia_numero not in por_dia_semana:
            por_dia_semana[dia_numero] = {"dia_nombre": DIAS[dia_numero], "ventas_total": 0.0, "costo_total": 0.0, "margen_total": 0.0, "num_ventas": 0}
        por_dia_semana[dia_numero]["ventas_total"] += sale_ventas
        por_dia_semana[dia_numero]["costo_total"] += sale_costo
        por_dia_semana[dia_numero]["margen_total"] += sale_margen
        por_dia_semana[dia_numero]["num_ventas"] += 1

        # Por sucursal
        bid = s.get("branch_id") or "global"
        if bid not in por_sucursal_agg:
            por_sucursal_agg[bid] = {
                "branch_id": bid,
                "branch_nombre": branch_name_map.get(bid, "Sin sucursal"),
                "ventas_total": 0.0, "costo_total": 0.0, "margen_total": 0.0, "num_ventas": 0
            }
        por_sucursal_agg[bid]["ventas_total"] += sale_ventas
        por_sucursal_agg[bid]["costo_total"] += sale_costo
        por_sucursal_agg[bid]["margen_total"] += sale_margen
        por_sucursal_agg[bid]["num_ventas"] += 1

        # Por venta
        por_venta_list.append({
            "sale_id": s.get("id"),
            "fecha": s.get("fecha").isoformat() if hasattr(s.get("fecha"), "isoformat") else str(s.get("fecha")),
            "fecha_dia": fecha_dia,
            "dia_semana": dia_numero,
            "branch_id": s.get("branch_id"),
            "branch_nombre": branch_name_map.get(s.get("branch_id", ""), "Sin sucursal"),
            "numero_factura": s.get("numero_factura"),
            "ventas_total": round(sale_ventas, 2),
            "costo_total": round(sale_costo, 2),
            "margen_total": round(sale_margen, 2),
            "margen_pct": round(sale_margen_pct, 2) if sale_margen_pct is not None else None,
            "items": items_detalle,
        })

    def _pct(margen, costo):
        return round(margen / costo * 100, 2) if costo and costo > 0 else None

    margen_total_global = round(total_margen, 2)
    margen_pct_global = _pct(total_margen, total_costo)

    por_fecha_list = sorted([
        {
            "fecha": k,
            "ventas_total": round(v["ventas_total"], 2),
            "costo_total": round(v["costo_total"], 2),
            "margen_total": round(v["margen_total"], 2),
            "margen_pct": _pct(v["margen_total"], v["costo_total"]),
            "num_ventas": v["num_ventas"],
        }
        for k, v in por_fecha.items()
    ], key=lambda x: x["fecha"])

    por_dia_list = sorted([
        {
            "dia_numero": k,
            "dia_nombre": v["dia_nombre"],
            "ventas_total": round(v["ventas_total"], 2),
            "costo_total": round(v["costo_total"], 2),
            "margen_total": round(v["margen_total"], 2),
            "margen_pct": _pct(v["margen_total"], v["costo_total"]),
            "num_ventas": v["num_ventas"],
        }
        for k, v in por_dia_semana.items()
    ], key=lambda x: x["dia_numero"])

    por_sucursal_list = sorted([
        {
            **v,
            "ventas_total": round(v["ventas_total"], 2),
            "costo_total": round(v["costo_total"], 2),
            "margen_total": round(v["margen_total"], 2),
            "margen_pct": _pct(v["margen_total"], v["costo_total"]),
        }
        for v in por_sucursal_agg.values()
    ], key=lambda x: x["ventas_total"], reverse=True)

    return {
        "resumen": {
            "ventas_total": round(total_ventas, 2),
            "costo_total": round(total_costo, 2),
            "margen_total": margen_total_global,
            "margen_pct": margen_pct_global,
            "num_ventas": num_ventas,
        },
        "por_fecha": por_fecha_list,
        "por_dia_semana": por_dia_list,
        "por_sucursal": por_sucursal_list,
        "por_venta": por_venta_list,
    }

@api_router.get("/reportes/ingresos-egresos")
async def get_reporte_ingresos_egresos(
    fecha_desde: str = Query(...),
    fecha_hasta: str = Query(...),
    user: User = Depends(get_current_user)
):
    AR_TZ = timezone(timedelta(hours=-3))
    MESES_ES = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}

    desde_local = datetime.strptime(fecha_desde, "%Y-%m-%d").replace(hour=0, minute=0, second=0, tzinfo=AR_TZ)
    hasta_local = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=AR_TZ)
    desde_utc = desde_local.astimezone(timezone.utc)
    hasta_utc = hasta_local.astimezone(timezone.utc)

    sales = await db.sales.find(
        {"empresa_id": user.empresa_id, "fecha": {"$gte": desde_utc, "$lte": hasta_utc}, "estado": {"$ne": "cancelado"}},
        {"fecha": 1, "total": 1}
    ).to_list(100000)

    compras = await db.compras.find(
        {"empresa_id": user.empresa_id, "fecha": {"$gte": desde_utc, "$lte": hasta_utc}},
        {"fecha": 1, "total": 1}
    ).to_list(100000)

    monthly: dict = {}

    def _get_key(fecha_raw):
        if isinstance(fecha_raw, str):
            fecha_raw = datetime.fromisoformat(fecha_raw.replace("Z", "+00:00"))
        if fecha_raw and fecha_raw.tzinfo is None:
            fecha_raw = fecha_raw.replace(tzinfo=timezone.utc)
        return fecha_raw.astimezone(AR_TZ).strftime("%Y-%m") if fecha_raw else None

    for s in sales:
        key = _get_key(s.get("fecha"))
        if not key:
            continue
        if key not in monthly:
            monthly[key] = {"ingresos": 0.0, "egresos": 0.0}
        monthly[key]["ingresos"] += s.get("total", 0)

    for c in compras:
        key = _get_key(c.get("fecha"))
        if not key:
            continue
        if key not in monthly:
            monthly[key] = {"ingresos": 0.0, "egresos": 0.0}
        monthly[key]["egresos"] += c.get("total", 0)

    # Rellenar todos los meses del rango aunque no tengan datos
    meses = []
    cur = desde_local.replace(day=1)
    end = hasta_local.replace(day=1)
    while cur <= end:
        key = cur.strftime("%Y-%m")
        label = f"{MESES_ES[cur.month]} {cur.year}"
        d = monthly.get(key, {"ingresos": 0.0, "egresos": 0.0})
        ingresos = round(d["ingresos"], 2)
        egresos = round(d["egresos"], 2)
        meses.append({"mes": key, "mes_label": label, "ingresos": ingresos, "egresos": egresos, "balance": round(ingresos - egresos, 2)})
        cur = cur.replace(month=cur.month % 12 + 1, year=cur.year + (1 if cur.month == 12 else 0))

    total_ingresos = round(sum(m["ingresos"] for m in meses), 2)
    total_egresos = round(sum(m["egresos"] for m in meses), 2)

    return {
        "meses": meses,
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "balance_total": round(total_ingresos - total_egresos, 2),
    }

@api_router.get("/reportes/productos-vendidos")
async def get_reporte_productos_vendidos(
    fecha_desde: str = Query(...),
    fecha_hasta: str = Query(...),
    branch_id: Optional[str] = Query(None),
    user: User = Depends(get_current_user)
):
    AR_TZ = timezone(timedelta(hours=-3))

    desde_local = datetime.strptime(fecha_desde, "%Y-%m-%d").replace(
        hour=0, minute=0, second=0, tzinfo=AR_TZ
    )
    hasta_local = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(
        hour=23, minute=59, second=59, tzinfo=AR_TZ
    )
    desde_utc = desde_local.astimezone(timezone.utc)
    hasta_utc = hasta_local.astimezone(timezone.utc)

    filtro = {
        "empresa_id": user.empresa_id,
        "fecha": {"$gte": desde_utc, "$lte": hasta_utc},
        "estado": {"$ne": "cancelado"},
        "branch_id": {"$nin": ["global", "", None]},
    }
    if branch_id and branch_id != "all":
        filtro["branch_id"] = branch_id

    sales = await db.sales.find(
        filtro,
        {"items": 1, "branch_id": 1, "_id": 0}
    ).to_list(100000)

    products_cursor = await db.products.find(
        {"empresa_id": user.empresa_id},
        {"id": 1, "nombre": 1, "categoria_id": 1, "_id": 0}
    ).to_list(10000)
    product_map = {p["id"]: p for p in products_cursor}

    categories_cursor = await db.categories.find(
        {"empresa_id": user.empresa_id},
        {"id": 1, "nombre": 1, "_id": 0}
    ).to_list(1000)
    cat_map = {c["id"]: c["nombre"] for c in categories_cursor}

    branches_cursor = await db.branches.find(
        {"empresa_id": user.empresa_id},
        {"id": 1, "nombre": 1, "_id": 0}
    ).to_list(1000)
    branch_name_map = {b["id"]: b["nombre"] for b in branches_cursor}

    # Clave de agrupación: (nombre_normalizado, branch_id) para desagregar por sucursal
    # y fusionar productos duplicados con el mismo nombre dentro de la misma sucursal
    by_key: dict = {}
    for sale in sales:
        bid = sale.get("branch_id", "")
        branch_nombre = branch_name_map.get(bid, bid or "Sin sucursal")
        for item in sale.get("items", []):
            pid = item.get("producto_id", "")
            if not pid:
                continue
            prod_data = product_map.get(pid, {})
            nombre = (item.get("nombre") or prod_data.get("nombre", "Sin nombre")).strip()
            cantidad = float(item.get("cantidad", 0))
            subtotal = float(item.get("subtotal") or item.get("total") or 0)
            categoria_id = prod_data.get("categoria_id", "")
            categoria_nombre = cat_map.get(categoria_id, "Sin categoría") if categoria_id else "Sin categoría"

            key = (nombre.lower(), bid)
            if key not in by_key:
                by_key[key] = {
                    "nombre": nombre,
                    "branch_id": bid,
                    "branch_nombre": branch_nombre,
                    "categoria_id": categoria_id,
                    "categoria_nombre": categoria_nombre,
                    "cantidad_vendida": 0.0,
                    "total_recaudado": 0.0,
                    "num_ventas": 0,
                    "_product_ids": set(),
                }

            by_key[key]["cantidad_vendida"] += cantidad
            by_key[key]["total_recaudado"] += subtotal
            by_key[key]["num_ventas"] += 1
            by_key[key]["_product_ids"].add(pid)

    # Cargar stock actual desde branch_products
    all_pids = list({pid for entry in by_key.values() for pid in entry["_product_ids"]})
    all_bids = list({entry["branch_id"] for entry in by_key.values() if entry["branch_id"]})
    bp_cursor = await db.branch_products.find(
        {"product_id": {"$in": all_pids}, "branch_id": {"$in": all_bids}, "empresa_id": user.empresa_id},
        {"product_id": 1, "branch_id": 1, "stock": 1, "_id": 0}
    ).to_list(100000)
    # stock_map: (product_id, branch_id) -> stock
    stock_map = {(bp["product_id"], bp["branch_id"]): bp.get("stock", 0) for bp in bp_cursor}

    result = []
    for entry in by_key.values():
        qty = entry["cantidad_vendida"]
        total = entry["total_recaudado"]
        bid = entry["branch_id"]
        stock_actual = sum(stock_map.get((pid, bid), 0) for pid in entry["_product_ids"])
        entry["cantidad_vendida"] = round(qty, 2)
        entry["total_recaudado"] = round(total, 2)
        entry["precio_promedio"] = round(total / qty, 2) if qty > 0 else 0.0
        entry["stock_actual"] = stock_actual
        del entry["_product_ids"]
        result.append(entry)

    result.sort(key=lambda x: (x["nombre"].lower(), x["branch_nombre"]))
    return result


@api_router.post("/sales/{sale_id}/return")
async def create_sale_return(sale_id: str, return_data: ReturnCreate, user: User = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id, "empresa_id": user.empresa_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    sale_obj = Sale(**sale)

    if sale_obj.estado == "cancelado":
        raise HTTPException(status_code=400, detail="Esta venta ya fue cancelada completamente")

    # Quantities already returned for this sale
    existing_returns = await db.sale_returns.find({"sale_id": sale_id}).to_list(100)
    returned_qty = {}
    for ret in existing_returns:
        for item in ret["items"]:
            pid = item["producto_id"]
            returned_qty[pid] = returned_qty.get(pid, 0) + item["cantidad"]

    sale_items_map = {item.producto_id: item for item in sale_obj.items}
    return_items = []
    total_return = 0.0

    for ri in return_data.items:
        if ri.cantidad <= 0:
            continue
        original = sale_items_map.get(ri.producto_id)
        if not original:
            raise HTTPException(status_code=400, detail=f"Producto no encontrado en la venta")
        available = original.cantidad - returned_qty.get(ri.producto_id, 0)
        if ri.cantidad > available:
            prod = await db.products.find_one({"id": ri.producto_id})
            name = prod["nombre"] if prod else ri.producto_id
            raise HTTPException(status_code=400, detail=f"Solo quedan {available} unidades disponibles para devolver de '{name}'")
        subtotal = ri.cantidad * original.precio_unitario
        total_return += subtotal
        return_items.append(SaleItem(
            producto_id=ri.producto_id,
            nombre=original.nombre,
            cantidad=ri.cantidad,
            precio_unitario=original.precio_unitario,
            subtotal=subtotal
        ))

    if not return_items:
        raise HTTPException(status_code=400, detail="Debe seleccionar al menos un producto para devolver")

    # Restore stock (only for products with control_stock enabled)
    for item in return_items:
        gp = await db.products.find_one({"id": item.producto_id, "empresa_id": user.empresa_id})
        if not gp or not gp.get("control_stock", True):
            continue
        await db.products.update_one(
            {"id": item.producto_id, "empresa_id": user.empresa_id},
            {"$inc": {"stock": int(item.cantidad)}}
        )
        if sale_obj.branch_id and sale_obj.branch_id != "global":
            await db.branch_products.update_one(
                {"product_id": item.producto_id, "branch_id": sale_obj.branch_id, "empresa_id": user.empresa_id},
                {"$inc": {"stock": int(item.cantidad)}}
            )

    # Generate return number
    count = await db.sale_returns.count_documents({"empresa_id": user.empresa_id})
    numero_devolucion = f"DEV-{str(count + 1).zfill(6)}"

    sale_return = SaleReturn(
        empresa_id=user.empresa_id,
        sale_id=sale_id,
        cajero_id=user.id,
        items=return_items,
        total=total_return,
        motivo=return_data.motivo,
        numero_devolucion=numero_devolucion
    )
    await db.sale_returns.insert_one(sale_return.dict())

    # Update sale estado
    new_returned = dict(returned_qty)
    for item in return_items:
        new_returned[item.producto_id] = new_returned.get(item.producto_id, 0) + item.cantidad
    fully_returned = all(new_returned.get(si.producto_id, 0) >= si.cantidad for si in sale_obj.items)
    new_estado = "cancelado" if fully_returned else "devolucion_parcial"
    await db.sales.update_one({"id": sale_id}, {"$set": {"estado": new_estado}})

    # Create credit note
    cn_count = await db.credit_notes.count_documents({"empresa_id": user.empresa_id})
    numero_nota_credito = f"NC-{str(cn_count + 1).zfill(6)}"
    credit_note = CreditNote(
        empresa_id=user.empresa_id,
        sale_id=sale_id,
        return_id=sale_return.id,
        cajero_id=user.id,
        numero_nota_credito=numero_nota_credito,
        numero_factura_original=sale_obj.numero_factura,
        items=return_items,
        total=total_return,
        motivo=return_data.motivo,
        tipo="total" if fully_returned else "parcial"
    )
    await db.credit_notes.insert_one(credit_note.dict())

    # Request CAE for credit note if original sale was authorized by AFIP
    # Factura A(1)→NC-A(3)  Factura B(6)→NC-B(8)  Factura C(11)→NC-C(13)
    NC_TIPO_MAP = {1: 3, 6: 8, 11: 13}
    nc_tipo = NC_TIPO_MAP.get(sale_obj.tipo_comprobante) if sale_obj.tipo_comprobante else None
    if sale_obj.afip_estado == "autorizado" and nc_tipo and sale_obj.nro_comprobante_afip:
        afip_cfg = await db.afip_config.find_one({"empresa_id": user.empresa_id, "activo": True})
        if afip_cfg and afip_cfg.get("cert_pem") and afip_cfg.get("key_pem_encrypted"):
            try:
                token_a, sign_a = await _afip_service.get_token_sign(afip_cfg, db, SECRET_KEY)
                tax_rate = afip_cfg.get("tax_rate", 0.21) if nc_tipo != 13 else 0.0
                if tax_rate > 0:
                    nc_subtotal = round(total_return / (1 + tax_rate), 2)
                    nc_impuestos = round(total_return - nc_subtotal, 2)
                else:
                    nc_subtotal = round(total_return, 2)
                    nc_impuestos = 0.0
                nc_sale_dict = {
                    "total": round(total_return, 2),
                    "subtotal": nc_subtotal,
                    "impuestos": nc_impuestos,
                    "impuestos_extra_total": 0.0,
                }
                cbtes_asoc = [{
                    "Tipo": sale_obj.tipo_comprobante,
                    "PtoVta": afip_cfg["punto_venta"],
                    "Nro": sale_obj.nro_comprobante_afip,
                }]
                nc_result = await _afip_service.solicitar_cae(
                    nc_sale_dict, afip_cfg, token_a, sign_a, nc_tipo,
                    cuit_receptor=sale_obj.cuit_receptor,
                    cbtes_asoc=cbtes_asoc,
                )
                await db.credit_notes.update_one(
                    {"id": credit_note.id},
                    {"$set": {
                        "cae": nc_result["cae"],
                        "cae_vencimiento": nc_result["cae_vencimiento"],
                        "afip_estado": "autorizado",
                        "tipo_comprobante_nc": nc_tipo,
                        "nro_comprobante_afip": nc_result["nro_comprobante"],
                    }}
                )
                credit_note.cae = nc_result["cae"]
                credit_note.cae_vencimiento = nc_result["cae_vencimiento"]
                credit_note.afip_estado = "autorizado"
                credit_note.tipo_comprobante_nc = nc_tipo
                credit_note.nro_comprobante_afip = nc_result["nro_comprobante"]
            except Exception as e:
                logger.warning(f"Error al obtener CAE para nota de crédito {numero_nota_credito}: {e}")
                await db.credit_notes.update_one(
                    {"id": credit_note.id},
                    {"$set": {
                        "afip_estado": "contingencia",
                        "afip_error": str(e),
                        "tipo_comprobante_nc": nc_tipo,
                    }}
                )

    # Reduce cash session monto_ventas
    await db.cash_sessions.update_one(
        {"id": sale_obj.session_id},
        {"$inc": {"monto_ventas": -total_return}}
    )

    # Cash movement
    movement = CashMovement(
        empresa_id=user.empresa_id,
        session_id=sale_obj.session_id,
        tipo=MovementType.DEVOLUCION,
        monto=-total_return,
        descripcion=f"Nota de crédito {numero_nota_credito} - {numero_devolucion} - Factura {sale_obj.numero_factura}",
        venta_id=sale_id
    )
    await db.cash_movements.insert_one(movement.dict())

    return {
        "message": "Devolución procesada exitosamente",
        "numero_devolucion": numero_devolucion,
        "numero_nota_credito": numero_nota_credito,
        "cae_nc": credit_note.cae,
        "afip_estado_nc": credit_note.afip_estado,
        "total": total_return
    }

# Cash Reports routes
@api_router.get("/cash-sessions/{session_id}/movements", response_model=List[CashMovement])
async def get_session_movements(session_id: str, user: User = Depends(get_current_user)):
    movements = await db.cash_movements.find({"session_id": session_id, "empresa_id": user.empresa_id}).sort("fecha", 1).to_list(1000)
    return [CashMovement(**movement) for movement in movements]

@api_router.get("/cash-sessions/{session_id}/report")
async def get_cash_session_report(session_id: str, user: User = Depends(get_current_user)):
    session = await db.cash_sessions.find_one({"id": session_id, "empresa_id": user.empresa_id})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")

    # Get movements
    movements_docs = await db.cash_movements.find({"session_id": session_id, "empresa_id": user.empresa_id}).sort("fecha", 1).to_list(1000)
    movements = [CashMovement(**movement) for movement in movements_docs]

    # Get sales details
    sales_docs = await db.sales.find({"session_id": session_id, "empresa_id": user.empresa_id}).to_list(1000)
    sales = [Sale(**sale) for sale in sales_docs]

    # Get user info
    user_doc = await db.users.find_one({"id": session["user_id"], "empresa_id": user.empresa_id})
    user_info = User(**user_doc) if user_doc else None

    # Get branch info
    branch_doc = await db.branches.find_one({"id": session["branch_id"], "empresa_id": user.empresa_id})
    branch_info = Branch(**branch_doc) if branch_doc else None

    # Devoluciones de esta sesión para descontar del resumen
    sale_ids = [s.id for s in sales]
    returns_docs = await db.sale_returns.find({"sale_id": {"$in": sale_ids}, "empresa_id": user.empresa_id}).to_list(1000)
    returned_by_sale = {}
    for ret in returns_docs:
        returned_by_sale[ret["sale_id"]] = returned_by_sale.get(ret["sale_id"], 0) + ret["total"]

    def net_total(sale):
        return sale.total - returned_by_sale.get(sale.id, 0)

    return {
        "session": CashSession(**session),
        "movements": movements,
        "sales": sales,
        "user": user_info,
        "branch": branch_info,
        "resumen": {
            "total_ventas": len(sales),
            "ingresos_efectivo": sum(net_total(s) for s in sales if s.metodo_pago == 'efectivo'),
            "ingresos_tarjeta": sum(net_total(s) for s in sales if s.metodo_pago == 'tarjeta'),
            "ingresos_transferencia": sum(net_total(s) for s in sales if s.metodo_pago == 'transferencia'),
        }
    }

# Dashboard routes
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: User = Depends(get_current_user)):
    is_admin_or_supervisor = user.rol in [UserRole.ADMIN, UserRole.SUPERVISOR]

    # Today's sales
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)

    sales_filter = {"empresa_id": user.empresa_id, "fecha": {"$gte": today, "$lt": tomorrow}}
    if not is_admin_or_supervisor and user.branch_id:
        sales_filter["branch_id"] = user.branch_id

    today_sales = await db.sales.find(sales_filter).to_list(1000)
    today_sale_ids = [s['id'] for s in today_sales]
    today_returns = await db.sale_returns.find({"sale_id": {"$in": today_sale_ids}, "empresa_id": user.empresa_id}).to_list(1000)
    returned_by_sale_today = {}
    for ret in today_returns:
        returned_by_sale_today[ret["sale_id"]] = returned_by_sale_today.get(ret["sale_id"], 0) + ret["total"]
    total_ventas_hoy = sum(sale['total'] - returned_by_sale_today.get(sale['id'], 0) for sale in today_sales)
    numero_ventas_hoy = len(today_sales)

    # Per-branch breakdown for admin
    ventas_por_sucursal = []
    if user.rol == UserRole.ADMIN:
        branches_list = await db.branches.find({"empresa_id": user.empresa_id, "activo": True}).to_list(1000)
        branch_map = {b["id"]: b["nombre"] for b in branches_list}
        branch_totals = {}
        for sale in today_sales:
            bid = sale.get("branch_id") or "global"
            net = sale["total"] - returned_by_sale_today.get(sale["id"], 0)
            if bid not in branch_totals:
                branch_totals[bid] = {"total": 0, "cantidad": 0}
            branch_totals[bid]["total"] += net
            branch_totals[bid]["cantidad"] += 1
        for bid, data in branch_totals.items():
            nombre = branch_map.get(bid, "Sin sucursal") if bid != "global" else "Sin sucursal"
            ventas_por_sucursal.append({"branch_id": bid, "nombre": nombre, "total": data["total"], "cantidad": data["cantidad"]})
        ventas_por_sucursal.sort(key=lambda x: x["total"], reverse=True)

    # Total products
    total_productos = await db.products.count_documents({"empresa_id": user.empresa_id, "activo": True})

    # Low stock from branch_products (real per-branch stock)
    # Exclude combo products and products with control_stock=False
    excluded_dash = await db.products.find(
        {"empresa_id": user.empresa_id, "$or": [{"kind": "combo"}, {"control_stock": False}]},
        {"id": 1}
    ).to_list(100000)
    excluded_ids_dash = [p["id"] for p in excluded_dash]

    bp_filter = {
        "empresa_id": user.empresa_id,
        "activo": True,
        "$expr": {"$lte": ["$stock", "$stock_minimo"]}
    }
    if excluded_ids_dash:
        bp_filter["product_id"] = {"$nin": excluded_ids_dash}
    if not is_admin_or_supervisor and user.branch_id:
        bp_filter["branch_id"] = user.branch_id
    elif is_admin_or_supervisor and user.rol == UserRole.SUPERVISOR and user.branch_id:
        bp_filter["branch_id"] = user.branch_id

    # Check if low stock alerts are enabled in config
    empresa_config = await db.configuration.find_one({"empresa_id": user.empresa_id})
    low_stock_enabled = empresa_config.get("low_stock_alert_enabled", True) if empresa_config else True

    bajo_stock_count = 0
    preview = []
    if low_stock_enabled:
        bajo_stock_count = await db.branch_products.count_documents(bp_filter)
        bajo_stock_bps_preview = await db.branch_products.find(bp_filter).to_list(5)
    else:
        bajo_stock_bps_preview = []

    # Build preview list (up to 5) with product names
    for bp in bajo_stock_bps_preview:
        prod = await db.products.find_one({"id": bp["product_id"], "empresa_id": user.empresa_id})
        branch = await db.branches.find_one({"id": bp["branch_id"], "empresa_id": user.empresa_id})
        if prod:
            preview.append({
                "id": bp["id"],
                "nombre": prod.get("nombre", ""),
                "sucursal": branch.get("nombre", "") if branch else "",
                "stock": bp.get("stock", 0),
                "stock_minimo": bp.get("stock_minimo", 0)
            })

    sucursales_count = await db.branches.count_documents({"empresa_id": user.empresa_id, "activo": True})
    categorias_count = await db.categories.count_documents({"empresa_id": user.empresa_id})

    return {
        "ventas_hoy": {
            "total": total_ventas_hoy,
            "cantidad": numero_ventas_hoy,
            "por_sucursal": ventas_por_sucursal
        },
        "productos": {
            "total": total_productos,
            "bajo_stock": bajo_stock_count
        },
        "productos_bajo_stock": preview,
        "onboarding": {
            "sucursales": sucursales_count,
            "categorias": categorias_count,
            "productos": total_productos,
        }
    }

@api_router.get("/dashboard/ventas-diarias")
async def get_ventas_diarias(
    dias: int = Query(60, ge=7, le=90),
    branch_id: Optional[str] = Query(None),
    user: User = Depends(get_current_user)
):
    # Usar timezone local de Argentina (UTC-3, sin DST)
    AR_TZ = timezone(timedelta(hours=-3))
    hoy_local = datetime.now(AR_TZ).replace(hour=0, minute=0, second=0, microsecond=0)
    desde_local = hoy_local - timedelta(days=dias - 1)
    # Convertir a UTC para el filtro de MongoDB
    desde_utc = desde_local.astimezone(timezone.utc)

    filtro = {"empresa_id": user.empresa_id, "fecha": {"$gte": desde_utc}}
    if branch_id:
        filtro["branch_id"] = branch_id
    ventas = await db.sales.find(filtro, {"id": 1, "fecha": 1, "total": 1}).to_list(100000)

    # Traer devoluciones de esas ventas para descontar
    sale_ids = [v["id"] for v in ventas if v.get("id")]
    returns_docs = await db.sale_returns.find({"sale_id": {"$in": sale_ids}, "empresa_id": user.empresa_id}, {"sale_id": 1, "total": 1}).to_list(100000)
    returned_by_sale = {}
    for r in returns_docs:
        returned_by_sale[r["sale_id"]] = returned_by_sale.get(r["sale_id"], 0) + r.get("total", 0)

    # Agrupar por fecha local (no UTC)
    totals: dict = {}
    for v in ventas:
        fecha = v.get("fecha")
        if not fecha:
            continue
        if isinstance(fecha, str):
            day = fecha[:10]
        else:
            fecha_local = fecha.astimezone(AR_TZ) if fecha.tzinfo else fecha.replace(tzinfo=timezone.utc).astimezone(AR_TZ)
            day = fecha_local.strftime("%Y-%m-%d")
        net = v.get("total", 0) - returned_by_sale.get(v.get("id", ""), 0)
        totals[day] = totals.get(day, 0) + net

    # Rellenar todos los días del rango con 0 si no hay ventas
    result = []
    for i in range(dias):
        d = (desde_local + timedelta(days=i)).strftime("%Y-%m-%d")
        result.append({"fecha": d, "total": round(totals.get(d, 0), 2)})

    return result

@api_router.get("/dashboard/stock-alerts")
async def get_stock_alerts(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=10000),
    user: User = Depends(get_current_user)
):
    is_admin = user.rol == UserRole.ADMIN

    # Exclude combo products and products with control_stock=False
    excluded_prods = await db.products.find(
        {"empresa_id": user.empresa_id, "$or": [{"kind": "combo"}, {"control_stock": False}]},
        {"id": 1}
    ).to_list(100000)
    excluded_ids = [p["id"] for p in excluded_prods]

    bp_filter = {
        "empresa_id": user.empresa_id,
        "activo": True,
        "$expr": {"$lte": ["$stock", "$stock_minimo"]}
    }
    if excluded_ids:
        bp_filter["product_id"] = {"$nin": excluded_ids}
    # Supervisor and cajero: filter to their branch
    if user.rol != UserRole.ADMIN and user.branch_id:
        bp_filter["branch_id"] = user.branch_id

    empresa_config = await db.configuration.find_one({"empresa_id": user.empresa_id})
    if not (empresa_config.get("low_stock_alert_enabled", True) if empresa_config else True):
        return {"items": [], "total": 0, "page": page, "per_page": per_page}

    total = await db.branch_products.count_documents(bp_filter)
    skip = (page - 1) * per_page
    bajo_stock_bps = await db.branch_products.find(bp_filter).skip(skip).limit(per_page).to_list(per_page)

    # Fetch product and branch info
    items = []
    for bp in bajo_stock_bps:
        prod = await db.products.find_one({"id": bp["product_id"], "empresa_id": user.empresa_id})
        branch = await db.branches.find_one({"id": bp["branch_id"], "empresa_id": user.empresa_id})
        items.append({
            "branch_product_id": bp["id"],
            "product_id": bp["product_id"],
            "branch_id": bp["branch_id"],
            "nombre": prod.get("nombre", "") if prod else "",
            "codigo_barras": prod.get("codigo_barras", "") if prod else "",
            "sucursal": branch.get("nombre", "") if branch else "",
            "stock": bp.get("stock", 0),
            "stock_minimo": bp.get("stock_minimo", 0),
            "diferencia": bp.get("stock", 0) - bp.get("stock_minimo", 0)
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, -(-total // per_page))
    }

@api_router.get("/dashboard/stock-alerts/export")
async def export_stock_alerts(
    format: str = Query("xlsx"),
    user: User = Depends(get_current_user)
):
    excluded_prods_exp = await db.products.find(
        {"empresa_id": user.empresa_id, "$or": [{"kind": "combo"}, {"control_stock": False}]},
        {"id": 1}
    ).to_list(100000)
    excluded_ids_exp = [p["id"] for p in excluded_prods_exp]

    bp_filter = {
        "empresa_id": user.empresa_id,
        "activo": True,
        "$expr": {"$lte": ["$stock", "$stock_minimo"]}
    }
    if excluded_ids_exp:
        bp_filter["product_id"] = {"$nin": excluded_ids_exp}
    if user.rol != UserRole.ADMIN and user.branch_id:
        bp_filter["branch_id"] = user.branch_id

    bajo_stock_bps = await db.branch_products.find(bp_filter).to_list(10000)

    rows = []
    for bp in bajo_stock_bps:
        prod = await db.products.find_one({"id": bp["product_id"], "empresa_id": user.empresa_id})
        branch = await db.branches.find_one({"id": bp["branch_id"], "empresa_id": user.empresa_id})
        rows.append({
            "Producto": prod.get("nombre", "") if prod else "",
            "Código de Barras": prod.get("codigo_barras", "") if prod else "",
            "Sucursal": branch.get("nombre", "") if branch else "",
            "Stock Actual": bp.get("stock", 0),
            "Stock Mínimo": bp.get("stock_minimo", 0),
            "Diferencia": bp.get("stock", 0) - bp.get("stock_minimo", 0)
        })

    df = pd.DataFrame(rows)
    output = io.BytesIO()

    if format == "xlsx":
        df.to_excel(output, index=False)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=stock_bajo.xlsx"}
        )
    else:
        df.to_csv(output, index=False)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=stock_bajo.csv"}
        )

# Proveedores routes
@api_router.post("/proveedores", response_model=Proveedor)
async def create_proveedor(
    proveedor_data: ProveedorCreate,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))
):
    proveedor = Proveedor(**proveedor_data.dict(), empresa_id=user.empresa_id)
    await db.proveedores.insert_one(proveedor.dict())
    return proveedor

@api_router.get("/proveedores", response_model=List[Proveedor])
async def get_proveedores(
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))
):
    proveedores = await db.proveedores.find(
        {"empresa_id": user.empresa_id}
    ).sort("nombre", 1).to_list(1000)
    return [Proveedor(**p) for p in proveedores]

@api_router.put("/proveedores/{proveedor_id}", response_model=Proveedor)
async def update_proveedor(
    proveedor_id: str,
    proveedor_data: ProveedorUpdate,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))
):
    proveedor = await db.proveedores.find_one({"id": proveedor_id, "empresa_id": user.empresa_id})
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor not found")
    update_data = {k: v for k, v in proveedor_data.dict().items() if v is not None}
    if update_data:
        await db.proveedores.update_one({"id": proveedor_id}, {"$set": update_data})
    updated = await db.proveedores.find_one({"id": proveedor_id})
    return Proveedor(**updated)

@api_router.delete("/proveedores/{proveedor_id}")
async def delete_proveedor(
    proveedor_id: str,
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    proveedor = await db.proveedores.find_one({"id": proveedor_id, "empresa_id": user.empresa_id})
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor not found")
    await db.proveedores.update_one({"id": proveedor_id}, {"$set": {"activo": False}})
    return {"message": "Proveedor deactivated"}

# Compras routes
@api_router.post("/compras", response_model=Compra)
async def create_compra(
    compra_data: CompraCreate,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))
):
    cfg = await db.configuration.find_one({"empresa_id": user.empresa_id}) or {}
    redondeo = cfg.get("redondeo_precio", 100)

    def _redondear_precio(valor: float) -> float:
        import math
        if not redondeo:
            return round(valor, 2)
        return math.ceil(valor / redondeo) * redondeo

    proveedor_nombre = None
    if compra_data.proveedor_id:
        prov = await db.proveedores.find_one({"id": compra_data.proveedor_id, "empresa_id": user.empresa_id})
        if prov:
            proveedor_nombre = prov["nombre"]

    # Process items: capture costo_anterior and update branch products in one pass
    enriched_items = []
    for item in compra_data.items:
        item_dict = item.dict()
        if item.product_id and compra_data.sucursal_id:
            bp = await db.branch_products.find_one({
                "product_id": item.product_id,
                "branch_id": compra_data.sucursal_id,
                "empresa_id": user.empresa_id
            })
            if bp:
                item_dict["costo_anterior"] = bp.get("costo")
                bp_update: dict = {"costo": item.precio_unitario}
                if item.actualizar_precio:
                    if item.nuevo_precio is not None:
                        bp_update["precio"] = item.nuevo_precio
                    else:
                        margen = bp.get("margen") or 0
                        bp_update["precio"] = _redondear_precio(item.precio_unitario * (1 + margen / 100))
                    if item.nuevo_margen is not None:
                        bp_update["margen"] = item.nuevo_margen
                await db.branch_products.update_one(
                    {"id": bp["id"]},
                    {"$set": bp_update, "$inc": {"stock": int(item.cantidad)}}
                )
            else:
                global_product = await db.products.find_one({"id": item.product_id, "empresa_id": user.empresa_id})
                if global_product:
                    new_bp = BranchProduct(
                        empresa_id=user.empresa_id,
                        product_id=item.product_id,
                        branch_id=compra_data.sucursal_id,
                        precio=global_product.get("precio", 0),
                        stock=int(item.cantidad),
                        stock_minimo=global_product.get("stock_minimo", 10),
                        costo=item.precio_unitario,
                    )
                    await db.branch_products.insert_one(new_bp.dict())
        enriched_items.append(item_dict)

    compra_dict = compra_data.dict()
    compra_dict["items"] = enriched_items
    compra = Compra(
        **compra_dict,
        empresa_id=user.empresa_id,
        proveedor_nombre=proveedor_nombre,
        registrado_por=user.id
    )
    await db.compras.insert_one(compra.dict())
    return compra

@api_router.get("/compras", response_model=List[Compra])
async def get_compras(
    proveedor_id: Optional[str] = None,
    fecha_desde: Optional[datetime] = None,
    fecha_hasta: Optional[datetime] = None,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))
):
    query = {"empresa_id": user.empresa_id}
    if proveedor_id:
        query["proveedor_id"] = proveedor_id
    if fecha_desde or fecha_hasta:
        query["fecha"] = {}
        if fecha_desde:
            query["fecha"]["$gte"] = fecha_desde
        if fecha_hasta:
            query["fecha"]["$lte"] = fecha_hasta
    compras = await db.compras.find(query).sort("fecha", -1).to_list(1000)
    return [Compra(**c) for c in compras]

@api_router.get("/compras/{compra_id}", response_model=Compra)
async def get_compra(
    compra_id: str,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))
):
    compra = await db.compras.find_one({"id": compra_id, "empresa_id": user.empresa_id})
    if not compra:
        raise HTTPException(status_code=404, detail="Compra not found")
    return Compra(**compra)

@api_router.put("/compras/{compra_id}", response_model=Compra)
async def update_compra(
    compra_id: str,
    compra_data: CompraUpdate,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))
):
    compra = await db.compras.find_one({"id": compra_id, "empresa_id": user.empresa_id})
    if not compra:
        raise HTTPException(status_code=404, detail="Compra not found")
    update_data = {k: v for k, v in compra_data.dict().items() if v is not None}
    if "proveedor_id" in update_data:
        prov = await db.proveedores.find_one({"id": update_data["proveedor_id"], "empresa_id": user.empresa_id})
        update_data["proveedor_nombre"] = prov["nombre"] if prov else None
    if update_data:
        await db.compras.update_one({"id": compra_id}, {"$set": update_data})
    updated = await db.compras.find_one({"id": compra_id})
    return Compra(**updated)

@api_router.delete("/compras/{compra_id}")
async def delete_compra(
    compra_id: str,
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    compra = await db.compras.find_one({"id": compra_id, "empresa_id": user.empresa_id})
    if not compra:
        raise HTTPException(status_code=404, detail="Compra not found")
    await db.compras.delete_one({"id": compra_id})
    return {"message": "Compra deleted"}

@api_router.post("/compras/{compra_id}/distribuir", response_model=Compra)
async def distribuir_compra(
    compra_id: str,
    data: DistribucionCreate,
    user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR]))
):
    compra = await db.compras.find_one({"id": compra_id, "empresa_id": user.empresa_id})
    if not compra:
        raise HTTPException(status_code=404, detail="Compra not found")

    branch = await db.branches.find_one({"id": data.sucursal_id, "empresa_id": user.empresa_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Sucursal not found")

    cfg = await db.configuration.find_one({"empresa_id": user.empresa_id}) or {}
    redondeo = cfg.get("redondeo_precio", 100)
    def _redondear(v: float) -> float:
        import math
        return math.ceil(v / redondeo) * redondeo if redondeo else round(v, 2)

    # Calcular saldo disponible por product_id
    ya_distribuido: dict = {}
    for d in compra.get("distribuciones", []):
        for it in d.get("items", []):
            pid = it.get("product_id")
            if pid:
                ya_distribuido[pid] = ya_distribuido.get(pid, 0) + float(it.get("cantidad", 0))

    original_qty: dict = {}
    for it in compra.get("items", []):
        pid = it.get("product_id")
        if pid:
            original_qty[pid] = original_qty.get(pid, 0) + float(it.get("cantidad", 0))

    # Validar saldo
    for item in data.items:
        if not item.actualizar_stock or item.cantidad <= 0:
            continue
        saldo = original_qty.get(item.product_id, 0) - ya_distribuido.get(item.product_id, 0)
        if item.cantidad > saldo + 0.001:
            raise HTTPException(
                status_code=400,
                detail=f"Saldo insuficiente: disponible {saldo}, solicitado {item.cantidad}"
            )

    # Aplicar a sucursal
    distribucion_items = []
    for item in data.items:
        if not item.actualizar_stock and not item.actualizar_precio:
            continue

        bp = await db.branch_products.find_one({
            "product_id": item.product_id,
            "branch_id": data.sucursal_id,
            "empresa_id": user.empresa_id
        })

        update_op: dict = {}
        if item.actualizar_precio:
            set_fields: dict = {}
            if item.nuevo_precio is not None:
                set_fields["precio"] = item.nuevo_precio
            if item.nuevo_margen is not None:
                set_fields["margen"] = item.nuevo_margen
            if set_fields:
                update_op["$set"] = set_fields

        if item.actualizar_stock and item.cantidad > 0:
            update_op.setdefault("$inc", {})["stock"] = int(item.cantidad)
            await db.products.update_one(
                {"id": item.product_id, "empresa_id": user.empresa_id},
                {"$inc": {"stock": int(item.cantidad)}}
            )

        if update_op:
            if bp:
                await db.branch_products.update_one({"id": bp["id"]}, update_op)
            else:
                global_product = await db.products.find_one({"id": item.product_id, "empresa_id": user.empresa_id})
                if global_product:
                    bp_data = dict(
                        empresa_id=user.empresa_id,
                        product_id=item.product_id,
                        branch_id=data.sucursal_id,
                        precio=item.nuevo_precio or global_product.get("precio", 0),
                        stock=int(item.cantidad) if item.actualizar_stock else 0,
                        stock_minimo=global_product.get("stock_minimo", 10),
                    )
                    if item.nuevo_margen is not None:
                        bp_data["margen"] = item.nuevo_margen
                    new_bp = BranchProduct(**bp_data)
                    await db.branch_products.insert_one(new_bp.dict())

        distribucion_items.append({
            "product_id": item.product_id,
            "cantidad": item.cantidad,
            "nuevo_precio": item.nuevo_precio,
            "nuevo_margen": item.nuevo_margen,
            "actualizo_stock": item.actualizar_stock,
            "actualizo_precio": item.actualizar_precio,
        })

    if not distribucion_items:
        raise HTTPException(status_code=400, detail="No hay ítems para aplicar")

    distribucion = {
        "id": str(uuid.uuid4()),
        "sucursal_id": data.sucursal_id,
        "sucursal_nombre": branch.get("nombre", ""),
        "fecha": datetime.now(timezone.utc),
        "aplicado_por": user.id,
        "items": distribucion_items,
    }

    await db.compras.update_one({"id": compra_id}, {"$push": {"distribuciones": distribucion}})
    updated = await db.compras.find_one({"id": compra_id})
    return Compra(**updated)

# ─────────────────────────────────────────────
# CUENTA / SUSCRIPCIÓN routes
# ─────────────────────────────────────────────

async def _get_or_create_suscripcion(empresa_id: str) -> dict:
    """Devuelve la suscripción de la empresa; si no existe, crea una trial."""
    doc = await db.suscripciones.find_one({"empresa_id": empresa_id})
    if not doc:
        now = datetime.now(timezone.utc)
        dia_facturacion = min(now.day, 28)
        suscripcion = Suscripcion(
            empresa_id=empresa_id,
            plan_nombre=await get_plan_nombre_suscripcion(),
            precio=await get_precio_suscripcion(),
            status=SuscripcionStatus.TRIAL,
            fecha_inicio=now,
            fecha_vencimiento=now + timedelta(days=await get_trial_dias()),
            dia_facturacion=dia_facturacion,
            plan_tipo="mensual",
        )
        await db.suscripciones.insert_one(suscripcion.dict())
        doc = suscripcion.dict()
    # Normalizar vencimiento
    vencimiento = doc["fecha_vencimiento"]
    if isinstance(vencimiento, str):
        vencimiento = datetime.fromisoformat(vencimiento)
    if not vencimiento.tzinfo:
        vencimiento = vencimiento.replace(tzinfo=timezone.utc)
    now = datetime.now(timezone.utc)
    # Auto-actualizar status si corresponde
    if vencimiento < now:
        if doc["status"] == SuscripcionStatus.TRIAL:
            # Trial vencido: sin periodo de gracia, bloqueo inmediato
            await db.suscripciones.update_one(
                {"empresa_id": empresa_id},
                {"$set": {"status": SuscripcionStatus.VENCIDA}}
            )
            doc["status"] = SuscripcionStatus.VENCIDA
        elif doc["status"] == SuscripcionStatus.ACTIVA:
            # Suscripción paga: periodo de gracia configurable
            gracia_fin = vencimiento + timedelta(days=await get_grace_days())
            if now > gracia_fin:
                # Gracia expirada → bloquear
                await db.suscripciones.update_one(
                    {"empresa_id": empresa_id},
                    {"$set": {"status": SuscripcionStatus.VENCIDA}}
                )
                doc["status"] = SuscripcionStatus.VENCIDA
            # else: dentro de gracia, mantener ACTIVA en DB
    return doc


def _calcular_estado_suscripcion(doc: dict, grace_days: int = 15) -> dict:
    """Calcula campos derivados de una suscripción: en_gracia, gracia_vencimiento, bloqueado, dias_restantes."""
    now = datetime.now(timezone.utc)
    vencimiento = doc["fecha_vencimiento"]
    if isinstance(vencimiento, str):
        vencimiento = datetime.fromisoformat(vencimiento)
    if not vencimiento.tzinfo:
        vencimiento = vencimiento.replace(tzinfo=timezone.utc)

    en_gracia = False
    gracia_vencimiento = None
    if doc["status"] == SuscripcionStatus.ACTIVA and vencimiento < now:
        gracia_vencimiento = vencimiento + timedelta(days=grace_days)
        en_gracia = True

    if en_gracia:
        dias_restantes = max(0, (gracia_vencimiento - now).days)
    else:
        dias_restantes = max(0, (vencimiento - now).days)

    bloqueado = doc["status"] in (SuscripcionStatus.VENCIDA, SuscripcionStatus.SUSPENDIDA)
    return {
        "en_gracia": en_gracia,
        "gracia_vencimiento": gracia_vencimiento,
        "dias_restantes": dias_restantes,
        "bloqueado": bloqueado,
    }


@api_router.get("/cuenta/status")
async def get_cuenta_status(user: User = Depends(require_role([UserRole.ADMIN]))):
    doc = await _get_or_create_suscripcion(user.empresa_id)
    estado = _calcular_estado_suscripcion(doc, grace_days=await get_grace_days())
    plan_tier = doc.get("plan_tier", "profesional")
    modules_extra = doc.get("modules_extra", [])
    modules_removidos = doc.get("modules_removidos", [])
    empresa = await db.empresas.find_one({"id": user.empresa_id})
    empresa_nombre = empresa["nombre"] if empresa else ""
    return {
        "id": doc["id"],
        "empresa_nombre": empresa_nombre,
        "plan_nombre": doc["plan_nombre"],
        "precio": doc["precio"],
        "moneda": doc["moneda"],
        "status": doc["status"],
        "fecha_inicio": doc["fecha_inicio"],
        "fecha_vencimiento": doc["fecha_vencimiento"],
        "fue_pagada": doc.get("fue_pagada", False),
        "tipo_cobro": doc.get("tipo_cobro", "manual"),
        "mp_preapproval_id": doc.get("mp_preapproval_id"),
        "dia_facturacion": doc.get("dia_facturacion"),
        "plan_tier": plan_tier,
        "modules_activos": calcular_modules_activos(plan_tier, modules_extra, modules_removidos),
        "addon_tienda": doc.get("addon_tienda", False),
        "sucursales_extra": doc.get("sucursales_extra", 0),
        "usuarios_extra_packs": doc.get("usuarios_extra_packs", 0),
        "descuento_pct": doc.get("descuento_pct", 0),
        **estado,
    }


@api_router.get("/auth/suscripcion")
async def get_suscripcion_check(user: User = Depends(require_role([UserRole.ADMIN, UserRole.SUPERVISOR, UserRole.CAJERO]))):
    """Estado de suscripción para control de acceso. Accesible a todos los roles autenticados."""
    doc = await _get_or_create_suscripcion(user.empresa_id)
    estado = _calcular_estado_suscripcion(doc, grace_days=await get_grace_days())
    plan_tier = doc.get("plan_tier", "profesional")
    modules_extra = doc.get("modules_extra", [])
    modules_removidos = doc.get("modules_removidos", [])
    modules_activos = calcular_modules_activos(plan_tier, modules_extra, modules_removidos)
    return {
        "status": doc["status"],
        "fecha_vencimiento": doc["fecha_vencimiento"],
        "fue_pagada": doc.get("fue_pagada", False),
        "plan_tier": plan_tier,
        "modules_activos": modules_activos,
        **estado,
    }


@api_router.get("/cuenta/pagos")
async def get_cuenta_pagos(user: User = Depends(require_role([UserRole.ADMIN]))):
    pagos = await db.pagos_suscripcion.find(
        {"empresa_id": user.empresa_id}
    ).sort("fecha", -1).to_list(50)
    return [PagoSuscripcion(**p) for p in pagos]


@api_router.get("/cuenta/planes")
async def get_planes(user: User = Depends(require_role([UserRole.ADMIN]))):
    precio_emp = await get_precio_emprendedor()
    precio_pro = await get_precio_profesional()
    precio_ent = await get_precio_empresarial()
    whatsapp = await get_whatsapp_numero()
    suscripcion = await db.suscripciones.find_one({"empresa_id": user.empresa_id})
    descuento_pct = (suscripcion or {}).get("descuento_pct", 0)
    def aplicar_descuento(precio: float) -> float:
        if descuento_pct > 0:
            return round(precio * (1 - descuento_pct / 100))
        return precio
    return {
        "tiers": {
            "emprendedor": {"precio_mensual": aplicar_descuento(precio_emp), "precio_anual": aplicar_descuento(precio_emp) * 11, "precio_mensual_original": precio_emp},
            "profesional":  {"precio_mensual": aplicar_descuento(precio_pro), "precio_anual": aplicar_descuento(precio_pro) * 11, "precio_mensual_original": precio_pro},
            "empresarial":  {"precio_mensual": aplicar_descuento(precio_ent), "precio_anual": aplicar_descuento(precio_ent) * 11, "precio_mensual_original": precio_ent},
        },
        "addon_precios": {
            "tienda": await get_precio_tienda(),
            "sucursal_extra": await get_precio_sucursal_extra(),
            "pack_usuarios": await get_precio_pack_usuarios(),
        },
        "whatsapp_numero": whatsapp,
        "descuento_pct": descuento_pct,
    }

@api_router.get("/public/planes")
async def get_planes_public():
    precio_emp = await get_precio_emprendedor()
    precio_pro = await get_precio_profesional()
    precio_ent = await get_precio_empresarial()
    return {
        "tiers": {
            "emprendedor": {"precio_mensual": precio_emp, "precio_anual": precio_emp * 11},
            "profesional":  {"precio_mensual": precio_pro, "precio_anual": precio_pro * 11},
            "empresarial":  {"precio_mensual": precio_ent, "precio_anual": precio_ent * 11},
        },
        "trial_dias": await get_trial_dias(),
    }


@api_router.get("/notificaciones")
async def get_notificaciones(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    skip = (page - 1) * per_page
    total = await db.notificaciones.count_documents({"empresa_id": user.empresa_id})
    items = await db.notificaciones.find(
        {"empresa_id": user.empresa_id}
    ).sort("fecha", -1).skip(skip).limit(per_page).to_list(per_page)
    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "items": [Notificacion(**n).dict() for n in items],
    }


@api_router.get("/notificaciones/count")
async def get_notificaciones_count(user: User = Depends(require_role([UserRole.ADMIN]))):
    count = await db.notificaciones.count_documents(
        {"empresa_id": user.empresa_id, "leida": False}
    )
    return {"no_leidas": count}


@api_router.put("/notificaciones/leer-todas")
async def marcar_todas_leidas(user: User = Depends(require_role([UserRole.ADMIN]))):
    await db.notificaciones.update_many(
        {"empresa_id": user.empresa_id, "leida": False},
        {"$set": {"leida": True}},
    )
    return {"ok": True}


@api_router.put("/notificaciones/{notif_id}/leer")
async def marcar_notificacion_leida(notif_id: str, user: User = Depends(require_role([UserRole.ADMIN]))):
    result = await db.notificaciones.update_one(
        {"id": notif_id, "empresa_id": user.empresa_id},
        {"$set": {"leida": True}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notificación no encontrada")
    return {"ok": True}


class PagoCreate(BaseModel):
    plan_tipo: str = "mensual"   # "mensual" | "anual"
    plan_tier: str = "profesional"  # "emprendedor" | "profesional" | "empresarial"
    addon_tienda: bool = False
    sucursales_extra: int = 0
    usuarios_extra_packs: int = 0


@api_router.post("/cuenta/pago/crear")
async def crear_pago_suscripcion(data: PagoCreate = PagoCreate(), user: User = Depends(require_role([UserRole.ADMIN]))):
    if not MP_ACCESS_TOKEN or MP_ACCESS_TOKEN.startswith("APP_USR-your"):
        raise HTTPException(status_code=503, detail="MercadoPago no configurado. Configure MP_ACCESS_TOKEN en .env")

    # Cancelar preferencias abandonadas (pending sin mp_payment_id = nunca se pagaron)
    await db.pagos_suscripcion.update_many(
        {"empresa_id": user.empresa_id, "estado": "pending", "mp_payment_id": {"$in": [None, ""]}},
        {"$set": {"estado": "cancelled"}},
    )

    # Bloquear solo si hay un pago pendiente con mp_payment_id (pagó con método lento, esperando acreditación)
    pago_pendiente = await db.pagos_suscripcion.find_one({
        "empresa_id": user.empresa_id,
        "estado": "pending",
        "mp_payment_id": {"$nin": [None, ""]},
    })
    if pago_pendiente:
        raise HTTPException(status_code=409, detail="Ya tenés un pago en proceso de acreditación. Puede tardar unas horas. Esperá la confirmación antes de iniciar uno nuevo.")

    # Plan tipo, tier y precio
    plan_tipo = data.plan_tipo if data.plan_tipo in ("mensual", "anual") else "mensual"
    plan_tier = data.plan_tier if data.plan_tier in ("emprendedor", "profesional", "empresarial") else "profesional"
    precio_mensual_base = await get_precio_por_tier(plan_tier)

    # Obtener suscripción para descuento y ventana check
    suscripcion = await db.suscripciones.find_one({"empresa_id": user.empresa_id})
    descuento_pct = (suscripcion or {}).get("descuento_pct", 0)
    if descuento_pct > 0:
        precio_mensual_base = round(precio_mensual_base * (1 - descuento_pct / 100))

    # Addons
    addon_tienda = bool(data.addon_tienda)
    sucursales_extra = max(0, min(int(data.sucursales_extra), 9))
    usuarios_extra_packs = max(0, int(data.usuarios_extra_packs))
    precio_addon_mensual = (
        (await get_precio_tienda() if addon_tienda else 0) +
        sucursales_extra * await get_precio_sucursal_extra() +
        usuarios_extra_packs * await get_precio_pack_usuarios()
    )
    precio_mensual = precio_mensual_base + precio_addon_mensual

    tier_labels = {"emprendedor": "Emprendedor", "profesional": "Profesional", "empresarial": "Empresarial"}
    tier_label = tier_labels[plan_tier]
    if plan_tipo == "anual":
        precio = precio_mensual * 11
        plan_nombre = f"Plan {tier_label} Anual"
        ventana_dias = 60
        meses = 12
    else:
        precio = precio_mensual
        plan_nombre = f"Plan {tier_label}"
        ventana_dias = 30
        meses = 1

    # Verificar que la suscripción esté próxima a vencer o ya vencida
    if suscripcion:
        vencimiento = suscripcion.get("fecha_vencimiento")
        if isinstance(vencimiento, str):
            vencimiento = datetime.fromisoformat(vencimiento)
        if vencimiento and not vencimiento.tzinfo:
            vencimiento = vencimiento.replace(tzinfo=timezone.utc)
        if vencimiento:
            dias_restantes = (vencimiento - datetime.now(timezone.utc)).days
            if dias_restantes > ventana_dias:
                raise HTTPException(
                    status_code=400,
                    detail=f"Tu suscripción vence en {dias_restantes} días. Podrás renovar cuando queden {ventana_dias} días o menos."
                )

    # Obtener empresa
    empresa = await db.empresas.find_one({"id": user.empresa_id})
    empresa_nombre = empresa["nombre"] if empresa else user.empresa_id
    sdk = mercadopago.SDK(MP_ACCESS_TOKEN)
    preference_data = {
        "items": [
            {
                "title": f"{plan_nombre} - {empresa_nombre}",
                "quantity": 1,
                "unit_price": precio,
                "currency_id": "ARS",
            }
        ],
        "back_urls": {
            "success": f"{FRONTEND_URL}/cuenta?pago=success",
            "failure": f"{FRONTEND_URL}/cuenta?pago=failure",
            "pending": f"{FRONTEND_URL}/cuenta?pago=pending",
        },
        **({"auto_return": "approved"} if not FRONTEND_URL.startswith("http://localhost") else {}),
        "notification_url": f"{APP_URL}/api/mercadopago/webhook",
        "external_reference": user.empresa_id,
        "statement_descriptor": await get_statement_descriptor(),
    }
    response = sdk.preference().create(preference_data)
    if response["status"] not in (200, 201):
        import logging
        logging.error(f"MercadoPago error: status={response['status']}, response={response.get('response')}")
        detail = response.get("response", {}).get("message", "Error al crear preferencia en MercadoPago")
        raise HTTPException(status_code=502, detail=f"Error MP ({response['status']}): {detail}")

    preference = response["response"]

    # Registrar pago pendiente
    pago = PagoSuscripcion(
        empresa_id=user.empresa_id,
        monto=precio,
        moneda="ARS",
        estado="pending",
        concepto=f"{plan_nombre} - {empresa_nombre}",
        mp_preference_id=preference["id"],
        plan_tipo=plan_tipo,
        plan_tier=plan_tier,
        addon_tienda=addon_tienda,
        sucursales_extra=sucursales_extra,
        usuarios_extra_packs=usuarios_extra_packs,
    )
    await db.pagos_suscripcion.insert_one(pago.dict())

    return {
        "init_point": preference["init_point"],
        "sandbox_init_point": preference.get("sandbox_init_point"),
        "preference_id": preference["id"],
    }


@api_router.post("/cuenta/suscripcion/activar")
async def activar_suscripcion_automatica(user: User = Depends(require_role([UserRole.ADMIN]))):
    """Crea un preapproval en MercadoPago para habilitar el débito automático mensual."""
    if not MP_ACCESS_TOKEN or MP_ACCESS_TOKEN.startswith("APP_USR-your"):
        raise HTTPException(status_code=503, detail="MercadoPago no configurado.")

    # Obtener datos de empresa y admin
    empresa = await db.empresas.find_one({"id": user.empresa_id})
    empresa_nombre = empresa["nombre"] if empresa else user.empresa_id
    admin = await db.users.find_one({"empresa_id": user.empresa_id, "rol": "admin"})
    payer_email = MP_TEST_PAYER_EMAIL or (admin["email"] if admin else user.email)

    nombre_base = await get_plan_nombre_suscripcion()

    # Obtener suscripción para día de facturación, plan tier y descuento
    suscripcion = await db.suscripciones.find_one({"empresa_id": user.empresa_id})
    if suscripcion and suscripcion.get("tipo_cobro") == "automatico" and suscripcion.get("mp_preapproval_id"):
        raise HTTPException(status_code=409, detail="Ya tenés el débito automático activo.")
    dia_facturacion = (suscripcion or {}).get("dia_facturacion") or min(datetime.now(timezone.utc).day, 28)
    plan_tier = (suscripcion or {}).get("plan_tier", "profesional")
    precio_mensual = await get_precio_por_tier(plan_tier)
    descuento_pct = (suscripcion or {}).get("descuento_pct", 0)
    if descuento_pct > 0:
        precio_mensual = round(precio_mensual * (1 - descuento_pct / 100))

    back_url = f"{FRONTEND_URL}/cuenta?suscripcion=authorized"
    if back_url.startswith("http://localhost") or back_url.startswith("http://127.0.0.1"):
        back_url = "https://example.com"  # placeholder para dev: MP requiere URL pública
    preapproval_data = {
        "reason": f"{nombre_base} - {empresa_nombre}",
        "auto_recurring": {
            "frequency": 1,
            "frequency_type": "months",
            "transaction_amount": precio_mensual,
            "currency_id": "ARS",
            "billing_day": dia_facturacion,
        },
        "back_url": back_url,
        "payer_email": payer_email,
        "external_reference": user.empresa_id,
        "status": "pending",
    }

    try:
        async with httpx.AsyncClient() as client_http:
            resp = await client_http.post(
                "https://api.mercadopago.com/preapproval",
                json=preapproval_data,
                headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"},
                timeout=15,
            )
        if resp.status_code not in (200, 201):
            import logging
            logging.error(f"MP preapproval error: {resp.status_code} {resp.text}")
            raise HTTPException(status_code=502, detail=f"Error MP ({resp.status_code}): {resp.text[:500]}")
        data = resp.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Error al conectar con MercadoPago: {str(e)}")

    preapproval_id = data["id"]

    # Guardar preapproval_id en la suscripción (aún no es "automatico" hasta que el cliente autorice)
    if suscripcion:
        await db.suscripciones.update_one(
            {"empresa_id": user.empresa_id},
            {"$set": {"mp_preapproval_id": preapproval_id, "tipo_cobro": "pendiente_autorizacion"}},
        )
    else:
        # Crear suscripción si no existe (edge case)
        now = datetime.now(timezone.utc)
        nueva_fecha = calcular_siguiente_vencimiento(dia_facturacion, 1, now)
        await db.suscripciones.insert_one(Suscripcion(
            empresa_id=user.empresa_id,
            plan_nombre=nombre_base,
            precio=precio_mensual,
            status=SuscripcionStatus.TRIAL,
            fecha_inicio=now,
            fecha_vencimiento=nueva_fecha,
            dia_facturacion=dia_facturacion,
            mp_preapproval_id=preapproval_id,
            tipo_cobro="pendiente_autorizacion",
        ).dict())

    return {
        "init_point": data.get("init_point"),
        "sandbox_init_point": data.get("sandbox_init_point"),
        "preapproval_id": preapproval_id,
    }


@api_router.post("/cuenta/suscripcion/cancelar")
async def cancelar_suscripcion_automatica(user: User = Depends(require_role([UserRole.ADMIN]))):
    """Cancela el débito automático (preapproval) en MercadoPago."""
    suscripcion = await db.suscripciones.find_one({"empresa_id": user.empresa_id})
    if not suscripcion or not suscripcion.get("mp_preapproval_id"):
        raise HTTPException(status_code=404, detail="No tenés un débito automático activo.")

    preapproval_id = suscripcion["mp_preapproval_id"]

    try:
        async with httpx.AsyncClient() as client_http:
            resp = await client_http.put(
                f"https://api.mercadopago.com/preapproval/{preapproval_id}",
                json={"status": "cancelled"},
                headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"},
                timeout=15,
            )
        # 200/201 OK, también aceptamos 404 si ya fue cancelado en MP
        if resp.status_code not in (200, 201, 404):
            raise HTTPException(status_code=502, detail=f"Error MP ({resp.status_code}): {resp.json().get('message', 'Error al cancelar')}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Error al conectar con MercadoPago: {str(e)}")

    await db.suscripciones.update_one(
        {"empresa_id": user.empresa_id},
        {"$set": {"tipo_cobro": "manual", "mp_preapproval_id": None}},
    )
    return {"ok": True, "mensaje": "Débito automático cancelado. Tu suscripción actual sigue vigente hasta su vencimiento."}


@api_router.post("/cuenta/pago/simular-aprobado")
async def simular_pago_aprobado(
    plan_tipo: str = "mensual",
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Solo para desarrollo/testing: simula un pago aprobado y renueva la suscripción."""
    if not APP_URL.startswith("http://localhost"):
        raise HTTPException(status_code=403, detail="Solo disponible en entorno de desarrollo")

    plan_tipo = plan_tipo if plan_tipo in ("mensual", "anual") else "mensual"
    meses = 1 if plan_tipo == "mensual" else 12

    now = datetime.now(timezone.utc)
    fake_payment_id = f"TEST-{int(now.timestamp())}"
    precio_mensual = await get_precio_suscripcion()
    nombre_base = await get_plan_nombre_suscripcion()
    precio = precio_mensual if plan_tipo == "mensual" else precio_mensual * 11
    plan_nombre = nombre_base if plan_tipo == "mensual" else f"{nombre_base} Anual"

    empresa = await db.empresas.find_one({"id": user.empresa_id})
    empresa_nombre = empresa["nombre"] if empresa else user.empresa_id

    # Cancelar pendientes abandonados
    await db.pagos_suscripcion.update_many(
        {"empresa_id": user.empresa_id, "estado": "pending", "mp_payment_id": {"$in": [None, ""]}},
        {"$set": {"estado": "cancelled"}},
    )

    suscripcion = await db.suscripciones.find_one({"empresa_id": user.empresa_id})
    if suscripcion:
        base, nueva_fecha = _aplicar_renovacion(suscripcion, meses, now)
        await db.suscripciones.update_one(
            {"empresa_id": user.empresa_id},
            {"$set": {
                "status": SuscripcionStatus.ACTIVA,
                "fecha_vencimiento": nueva_fecha,
                "plan_tipo": plan_tipo,
                "fue_pagada": True,
                "plan_nombre": plan_nombre,
                "precio": precio,
            }},
        )
    else:
        dia_facturacion = min(now.day, 28)
        base = now
        nueva_fecha = calcular_siguiente_vencimiento(dia_facturacion, meses, now)
        await db.suscripciones.insert_one(Suscripcion(
            empresa_id=user.empresa_id,
            plan_nombre=plan_nombre,
            precio=precio,
            status=SuscripcionStatus.ACTIVA,
            fecha_inicio=now,
            fecha_vencimiento=nueva_fecha,
            dia_facturacion=dia_facturacion,
            plan_tipo=plan_tipo,
            fue_pagada=True,
        ).dict())

    # Registrar el pago simulado
    pago = PagoSuscripcion(
        empresa_id=user.empresa_id,
        monto=precio,
        moneda="ARS",
        estado="approved",
        concepto=f"[TEST] {plan_nombre} - {empresa_nombre}",
        mp_payment_id=fake_payment_id,
        periodo_inicio=base,
        periodo_fin=nueva_fecha,
        plan_tipo=plan_tipo,
    )
    await db.pagos_suscripcion.insert_one(pago.dict())

    return {"ok": True, "nueva_fecha_vencimiento": nueva_fecha.isoformat()}


# ─── Helpers internos del webhook ────────────────────────────────────────────

async def _procesar_pago_aprobado(empresa_id: str, payment_id: str, monto: float,
                                   plan_tipo_pago: str = "mensual", origen: str = "manual",
                                   preapproval_id: Optional[str] = None,
                                   plan_tier: Optional[str] = None,
                                   addon_tienda: bool = False,
                                   sucursales_extra: int = 0,
                                   usuarios_extra_packs: int = 0):
    """Aplica la renovación de suscripción cuando un pago es aprobado."""
    meses = 1 if plan_tipo_pago == "mensual" else 12
    suscripcion = await db.suscripciones.find_one({"empresa_id": empresa_id})
    now = datetime.now(timezone.utc)
    tier_efectivo = plan_tier or (suscripcion.get("plan_tier") if suscripcion else "profesional") or "profesional"
    tier_labels = {"emprendedor": "Emprendedor", "profesional": "Profesional", "empresarial": "Empresarial"}
    tier_label = tier_labels.get(tier_efectivo, "Profesional")
    nuevo_plan_nombre = f"Plan {tier_label}" if plan_tipo_pago == "mensual" else f"Plan {tier_label} Anual"
    if suscripcion:
        base, nueva_fecha = _aplicar_renovacion(suscripcion, meses, now)
        # Calcular modules_extra para incluir/excluir tienda
        existing_modules_extra = list(suscripcion.get("modules_extra", []))
        if addon_tienda and "tienda" not in existing_modules_extra:
            existing_modules_extra.append("tienda")
        elif not addon_tienda and "tienda" in existing_modules_extra:
            existing_modules_extra.remove("tienda")
        update_fields = {
            "status": SuscripcionStatus.ACTIVA,
            "fecha_vencimiento": nueva_fecha,
            "plan_tipo": plan_tipo_pago,
            "plan_tier": tier_efectivo,
            "fue_pagada": True,
            "plan_nombre": nuevo_plan_nombre,
            "precio": monto,
            "addon_tienda": addon_tienda,
            "sucursales_extra": sucursales_extra,
            "usuarios_extra_packs": usuarios_extra_packs,
            "modules_extra": existing_modules_extra,
        }
        # Si es preapproval, marcar como automático
        if preapproval_id:
            update_fields["tipo_cobro"] = "automatico"
            update_fields["mp_preapproval_id"] = preapproval_id
        await db.suscripciones.update_one({"empresa_id": empresa_id}, {"$set": update_fields})
        await db.pagos_suscripcion.update_one(
            {"empresa_id": empresa_id, "mp_payment_id": payment_id},
            {"$set": {"periodo_inicio": base, "periodo_fin": nueva_fecha}},
        )
    else:
        dia_facturacion = min(now.day, 28)
        nueva_fecha = calcular_siguiente_vencimiento(dia_facturacion, meses, now)
        suscripcion_nueva = Suscripcion(
            empresa_id=empresa_id,
            plan_nombre=nuevo_plan_nombre,
            precio=monto,
            status=SuscripcionStatus.ACTIVA,
            fecha_inicio=now,
            fecha_vencimiento=nueva_fecha,
            dia_facturacion=dia_facturacion,
            plan_tipo=plan_tipo_pago,
            plan_tier=tier_efectivo,
            fue_pagada=True,
            tipo_cobro="automatico" if preapproval_id else "manual",
            mp_preapproval_id=preapproval_id,
        )
        await db.suscripciones.insert_one(suscripcion_nueva.dict())


# Webhook de MercadoPago (sin autenticación JWT)
@app.post("/api/mercadopago/webhook")
async def mp_webhook(request: Request):
    try:
        body = await request.json()
    except Exception:
        return {"status": "ok"}

    if not MP_ACCESS_TOKEN or MP_ACCESS_TOKEN.startswith("APP_USR-your"):
        return {"status": "ok"}

    event_type = body.get("type", "")

    # ── Evento: cobro automático de suscripción ────────────────────────────────
    if event_type == "subscription_authorized_payment":
        authorized_payment_id = str(body.get("data", {}).get("id", ""))
        if not authorized_payment_id:
            return {"status": "ok"}
        try:
            async with httpx.AsyncClient() as client_http:
                resp = await client_http.get(
                    f"https://api.mercadopago.com/v1/subscription_authorized_payments/{authorized_payment_id}",
                    headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"},
                    timeout=10,
                )
            if resp.status_code != 200:
                return {"status": "ok"}
            auth_data = resp.json()
        except Exception:
            return {"status": "ok"}

        preapproval_id = auth_data.get("preapproval_id", "")
        payment_info = auth_data.get("payment", {})
        payment_id = str(payment_info.get("id", ""))
        estado_pago = payment_info.get("status", "")
        monto = float(auth_data.get("transaction_amount", payment_info.get("transaction_amount", 0)))

        # Buscar empresa por preapproval_id
        suscripcion = await db.suscripciones.find_one({"mp_preapproval_id": preapproval_id})
        if not suscripcion:
            return {"status": "ok"}
        empresa_id = suscripcion["empresa_id"]

        empresa = await db.empresas.find_one({"id": empresa_id})
        empresa_nombre = empresa["nombre"] if empresa else empresa_id

        # Registrar el pago si no existe
        existing = await db.pagos_suscripcion.find_one({"mp_payment_id": payment_id}) if payment_id else None
        if not existing:
            pago = PagoSuscripcion(
                empresa_id=empresa_id,
                monto=monto,
                moneda="ARS",
                estado=estado_pago if estado_pago else "approved",
                concepto=f"{SUSCRIPCION_PLAN_NOMBRE} Débito Automático - {empresa_nombre}",
                mp_payment_id=payment_id or authorized_payment_id,
                mp_preapproval_id=preapproval_id,
                origen="preapproval",
                plan_tipo="mensual",
            )
            await db.pagos_suscripcion.insert_one(pago.dict())

        if estado_pago in ("approved", "") or not estado_pago:
            await _procesar_pago_aprobado(
                empresa_id=empresa_id,
                payment_id=payment_id or authorized_payment_id,
                monto=monto,
                plan_tipo_pago="mensual",
                origen="preapproval",
                preapproval_id=preapproval_id,
            )
        return {"status": "ok"}

    # ── Evento: cambio de estado del preapproval ───────────────────────────────
    if event_type == "subscription_preapproval":
        preapproval_id = str(body.get("data", {}).get("id", ""))
        if not preapproval_id:
            return {"status": "ok"}
        try:
            async with httpx.AsyncClient() as client_http:
                resp = await client_http.get(
                    f"https://api.mercadopago.com/preapproval/{preapproval_id}",
                    headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"},
                    timeout=10,
                )
            if resp.status_code != 200:
                return {"status": "ok"}
            pa_data = resp.json()
        except Exception:
            return {"status": "ok"}

        pa_status = pa_data.get("status", "")
        # Si el cliente canceló desde MP, actualizar nuestra BD
        if pa_status in ("cancelled", "paused"):
            await db.suscripciones.update_many(
                {"mp_preapproval_id": preapproval_id},
                {"$set": {"tipo_cobro": "manual", "mp_preapproval_id": None}},
            )
        # Si fue autorizado (cliente aprobó el preapproval por primera vez)
        elif pa_status == "authorized":
            suscripcion = await db.suscripciones.find_one({"mp_preapproval_id": preapproval_id})
            if suscripcion:
                await db.suscripciones.update_one(
                    {"mp_preapproval_id": preapproval_id},
                    {"$set": {"tipo_cobro": "automatico"}},
                )
        return {"status": "ok"}

    # ── Evento: pago único (comportamiento previo) ─────────────────────────────
    if event_type != "payment":
        return {"status": "ok"}

    payment_id = str(body.get("data", {}).get("id", ""))
    if not payment_id:
        return {"status": "ok"}

    # Consultar detalle del pago a la API de MP
    try:
        async with httpx.AsyncClient() as client_http:
            resp = await client_http.get(
                f"https://api.mercadopago.com/v1/payments/{payment_id}",
                headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"},
                timeout=10,
            )
        if resp.status_code != 200:
            return {"status": "ok"}
        payment_data = resp.json()
    except Exception:
        return {"status": "ok"}

    empresa_id = payment_data.get("external_reference")
    estado = payment_data.get("status")  # approved, rejected, cancelled, pending
    monto = float(payment_data.get("transaction_amount", 0))

    if not empresa_id:
        return {"status": "ok"}

    # Actualizar o crear registro de pago
    existing_pago = await db.pagos_suscripcion.find_one({
        "empresa_id": empresa_id,
        "mp_payment_id": payment_id,
    })
    if not existing_pago:
        # Buscar pago pendiente por preference_id
        pref_id = str(payment_data.get("order", {}).get("id", "") or
                      payment_data.get("preference_id", ""))
        pago_doc = await db.pagos_suscripcion.find_one({
            "empresa_id": empresa_id,
            "mp_preference_id": pref_id,
            "estado": "pending",
        })
        if pago_doc:
            await db.pagos_suscripcion.update_one(
                {"id": pago_doc["id"]},
                {"$set": {"estado": estado, "mp_payment_id": payment_id}}
            )
        else:
            empresa = await db.empresas.find_one({"id": empresa_id})
            empresa_nombre = empresa["nombre"] if empresa else empresa_id
            pago = PagoSuscripcion(
                empresa_id=empresa_id,
                monto=monto,
                moneda="ARS",
                estado=estado,
                concepto=f"{SUSCRIPCION_PLAN_NOMBRE} - {empresa_nombre}",
                mp_payment_id=payment_id,
            )
            await db.pagos_suscripcion.insert_one(pago.dict())
    else:
        await db.pagos_suscripcion.update_one(
            {"id": existing_pago["id"]},
            {"$set": {"estado": estado}}
        )

    # Si el pago fue aprobado → extender suscripción
    if estado == "approved":
        pago_reg = await db.pagos_suscripcion.find_one({"empresa_id": empresa_id, "mp_payment_id": payment_id})
        if not pago_reg:
            pago_reg = await db.pagos_suscripcion.find_one({"empresa_id": empresa_id, "mp_preference_id": {"$exists": True}, "estado": "approved"})
        pr = pago_reg or {}
        await _procesar_pago_aprobado(
            empresa_id=empresa_id,
            payment_id=payment_id,
            monto=monto,
            plan_tipo_pago=pr.get("plan_tipo", "mensual"),
            plan_tier=pr.get("plan_tier"),
            addon_tienda=bool(pr.get("addon_tienda", False)),
            sucursales_extra=int(pr.get("sucursales_extra", 0)),
            usuarios_extra_packs=int(pr.get("usuarios_extra_packs", 0)),
        )

    return {"status": "ok"}

# ─────────────────────────────────────────────
# OWNER (System Admin) routes
# ─────────────────────────────────────────────

class OwnerLoginData(BaseModel):
    username: str
    password: str

class SuscripcionUpdate(BaseModel):
    status: Optional[SuscripcionStatus] = None
    fecha_vencimiento: Optional[datetime] = None
    plan_nombre: Optional[str] = None
    plan_tier: Optional[str] = None
    precio: Optional[float] = None
    dias_extra: Optional[int] = None
    descuento_pct: Optional[int] = None

class ModulosUpdate(BaseModel):
    modules_extra: Optional[List[str]] = None
    modules_removidos: Optional[List[str]] = None

class PagoManual(BaseModel):
    monto: float
    concepto: str
    plan_tipo: str = "mensual"  # "mensual" | "anual"

class ClienteDatosUpdate(BaseModel):
    empresa_nombre: Optional[str] = None
    admin_nombre: Optional[str] = None
    admin_email: Optional[str] = None

owner_router = APIRouter(prefix="/owner")

def create_owner_token():
    payload = {
        "sub": "owner",
        "role": "owner",
        "exp": datetime.now(timezone.utc) + timedelta(hours=24),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def verify_owner_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("role") != "owner":
            raise HTTPException(status_code=403, detail="Acceso denegado")
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except Exception:
        raise HTTPException(status_code=401, detail="Token inválido")

@owner_router.post("/login")
async def owner_login(data: OwnerLoginData):
    if data.username != OWNER_USERNAME or data.password != OWNER_PASSWORD:
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    token = create_owner_token()
    return {"access_token": token, "token_type": "bearer"}

@owner_router.post("/impersonate/{empresa_id}")
async def owner_impersonate(empresa_id: str, _=Depends(verify_owner_token)):
    empresa = await db.empresas.find_one({"id": empresa_id})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")

    admin = await db.users.find_one({"empresa_id": empresa_id, "rol": "admin", "activo": True})
    if not admin:
        raise HTTPException(status_code=404, detail="No se encontró usuario admin activo para esta empresa")

    access_token = create_access_token(
        data={"sub": admin["id"], "empresa_id": admin["empresa_id"]},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "empresa_nombre": empresa["nombre"],
    }

@owner_router.get("/stats")
async def owner_stats(_=Depends(verify_owner_token)):
    total = await db.empresas.count_documents({})
    activas = await db.suscripciones.count_documents({"status": SuscripcionStatus.ACTIVA})
    trial = await db.suscripciones.count_documents({"status": SuscripcionStatus.TRIAL})
    vencidas = await db.suscripciones.count_documents({"status": SuscripcionStatus.VENCIDA})
    suspendidas = await db.suscripciones.count_documents({"status": SuscripcionStatus.SUSPENDIDA})
    pagos = await db.pagos_suscripcion.find({"estado": "approved"}).to_list(10000)
    total_recaudado = sum(p.get("monto", 0) for p in pagos)
    alertas_sin_leer = await db.notificaciones.count_documents({"leida": False})
    return {
        "total_clientes": total,
        "activas": activas,
        "trial": trial,
        "vencidas": vencidas,
        "suspendidas": suspendidas,
        "total_recaudado": total_recaudado,
        "alertas_sin_leer": alertas_sin_leer,
    }

def _calc_dias_restantes(suscripcion: dict) -> int:
    if not suscripcion:
        return 0
    vencimiento = suscripcion.get("fecha_vencimiento")
    if isinstance(vencimiento, str):
        vencimiento = datetime.fromisoformat(vencimiento)
    if vencimiento and not vencimiento.tzinfo:
        vencimiento = vencimiento.replace(tzinfo=timezone.utc)
    if not vencimiento:
        return 0
    return max(0, (vencimiento - datetime.now(timezone.utc)).days)

@owner_router.post("/clientes")
async def owner_create_cliente(data: OwnerCreateCliente, _=Depends(verify_owner_token)):
    existing = await db.users.find_one({"email": data.admin_email.strip().lower()})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")

    empresa = Empresa(nombre=data.empresa_nombre, email_verificado=True)
    await db.empresas.insert_one(empresa.dict())

    hashed_password = bcrypt.hash(data.admin_password)
    user = User(empresa_id=empresa.id, nombre=data.admin_nombre, email=data.admin_email.strip().lower(), rol=UserRole.ADMIN)
    user_doc = user.dict()
    user_doc['password'] = hashed_password
    await db.users.insert_one(user_doc)

    config = Configuration(empresa_id=empresa.id, company_name=data.empresa_nombre)
    await db.configuration.insert_one(config.dict())

    trial_inicio = datetime.now(timezone.utc)
    trial_fin = trial_inicio + timedelta(days=await get_trial_dias())
    suscripcion = Suscripcion(
        empresa_id=empresa.id,
        plan_nombre=await get_plan_nombre_suscripcion(),
        precio=await get_precio_suscripcion(),
        status=SuscripcionStatus.TRIAL,
        fecha_inicio=trial_inicio,
        fecha_vencimiento=trial_fin,
        dia_facturacion=min(trial_fin.day, 28),
        plan_tipo="mensual",
        plan_tier="empresarial",
    )
    await db.suscripciones.insert_one(suscripcion.dict())

    default_branch = Branch(empresa_id=empresa.id, nombre="Sucursal Principal", direccion="")
    await db.branches.insert_one(default_branch.dict())

    for cat_nombre in ["General", "Alimentos", "Bebidas", "Limpieza"]:
        await db.categories.insert_one(Category(empresa_id=empresa.id, nombre=cat_nombre).dict())

    return {"id": empresa.id, "nombre": empresa.nombre, "email": data.admin_email}

@owner_router.get("/clientes")
async def owner_get_clientes(_=Depends(verify_owner_token)):
    empresas = await db.empresas.find({}).sort("created_at", -1).to_list(1000)
    result = []
    for emp in empresas:
        suscripcion = await db.suscripciones.find_one({"empresa_id": emp["id"]})
        admin = await db.users.find_one({"empresa_id": emp["id"], "rol": "admin"})
        pagos_count = await db.pagos_suscripcion.count_documents(
            {"empresa_id": emp["id"], "estado": "approved"}
        )
        dias_restantes = _calc_dias_restantes(suscripcion)
        # Convert ObjectId and datetime for JSON serialization
        sus_data = None
        if suscripcion:
            sus_data = {k: v for k, v in suscripcion.items() if k != "_id"}
        result.append({
            "id": emp["id"],
            "nombre": emp["nombre"],
            "activo": emp.get("activo", True),
            "created_at": emp.get("created_at"),
            "admin_email": admin["email"] if admin else None,
            "admin_nombre": admin["nombre"] if admin else None,
            "suscripcion": sus_data,
            "dias_restantes": dias_restantes,
            "pagos_aprobados": pagos_count,
        })
    return result

@owner_router.get("/clientes/{empresa_id}")
async def owner_get_cliente(empresa_id: str, _=Depends(verify_owner_token)):
    emp = await db.empresas.find_one({"id": empresa_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    suscripcion = await db.suscripciones.find_one({"empresa_id": empresa_id})
    admin = await db.users.find_one({"empresa_id": empresa_id, "rol": "admin"})
    pagos = await db.pagos_suscripcion.find(
        {"empresa_id": empresa_id}
    ).sort("fecha", -1).to_list(100)
    dias_restantes = _calc_dias_restantes(suscripcion)
    sus_data = {k: v for k, v in suscripcion.items() if k != "_id"} if suscripcion else None
    pagos_data = [{k: v for k, v in p.items() if k != "_id"} for p in pagos]
    plan_tier = suscripcion.get("plan_tier", "profesional") if suscripcion else "profesional"
    modules_extra = suscripcion.get("modules_extra", []) if suscripcion else []
    modules_removidos = suscripcion.get("modules_removidos", []) if suscripcion else []
    sucursales_docs = await db.branches.find(
        {"empresa_id": empresa_id, "activo": True}
    ).sort("created_at", 1).to_list(100)
    sucursales = [{"id": s["id"], "nombre": s["nombre"]} for s in sucursales_docs]
    return {
        "id": emp["id"],
        "nombre": emp["nombre"],
        "activo": emp.get("activo", True),
        "created_at": emp.get("created_at"),
        "admin_email": admin["email"] if admin else None,
        "admin_nombre": admin["nombre"] if admin else None,
        "suscripcion": sus_data,
        "dias_restantes": dias_restantes,
        "pagos": pagos_data,
        "plan_tier": plan_tier,
        "modules_extra": modules_extra,
        "modules_removidos": modules_removidos,
        "modules_activos": calcular_modules_activos(plan_tier, modules_extra, modules_removidos),
        "sucursales_activas": len(sucursales),
        "sucursales": sucursales,
    }

@owner_router.put("/clientes/{empresa_id}/suscripcion")
async def owner_update_suscripcion(
    empresa_id: str, data: SuscripcionUpdate, _=Depends(verify_owner_token)
):
    emp = await db.empresas.find_one({"id": empresa_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    suscripcion = await db.suscripciones.find_one({"empresa_id": empresa_id})
    update = {}

    if data.status is not None:
        update["status"] = data.status
    if data.plan_nombre is not None:
        update["plan_nombre"] = data.plan_nombre
    if data.plan_tier is not None and data.plan_tier in PLAN_MODULES:
        update["plan_tier"] = data.plan_tier
    if data.precio is not None:
        update["precio"] = data.precio
    if data.descuento_pct is not None:
        update["descuento_pct"] = max(0, min(100, data.descuento_pct))
    if data.dias_extra and data.dias_extra > 0:
        if suscripcion:
            vencimiento = suscripcion.get("fecha_vencimiento")
            if isinstance(vencimiento, str):
                vencimiento = datetime.fromisoformat(vencimiento)
            if vencimiento and not vencimiento.tzinfo:
                vencimiento = vencimiento.replace(tzinfo=timezone.utc)
        else:
            vencimiento = None
        now = datetime.now(timezone.utc)
        base = vencimiento if vencimiento and vencimiento > now else now
        update["fecha_vencimiento"] = base + timedelta(days=data.dias_extra)
        if not data.status:
            update["status"] = SuscripcionStatus.ACTIVA
    elif data.fecha_vencimiento is not None:
        update["fecha_vencimiento"] = data.fecha_vencimiento
    if update:
        if suscripcion:
            await db.suscripciones.update_one({"empresa_id": empresa_id}, {"$set": update})
        else:
            now = datetime.now(timezone.utc)
            nueva = Suscripcion(
                empresa_id=empresa_id,
                plan_nombre=update.get("plan_nombre", SUSCRIPCION_PLAN_NOMBRE),
                precio=update.get("precio", SUSCRIPCION_PRECIO),
                status=update.get("status", SuscripcionStatus.ACTIVA),
                fecha_inicio=now,
                fecha_vencimiento=update.get("fecha_vencimiento", now + timedelta(days=await get_trial_dias())),
            )
            await db.suscripciones.insert_one(nueva.dict())
    return {"message": "Suscripción actualizada"}

@owner_router.post("/clientes/{empresa_id}/pago")
async def owner_registrar_pago(
    empresa_id: str, data: PagoManual, _=Depends(verify_owner_token)
):
    emp = await db.empresas.find_one({"id": empresa_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    plan_tipo = data.plan_tipo if data.plan_tipo in ("mensual", "anual") else "mensual"
    meses = 1 if plan_tipo == "mensual" else 12
    pago = PagoSuscripcion(
        empresa_id=empresa_id,
        monto=data.monto,
        moneda="ARS",
        estado="approved",
        concepto=data.concepto,
        plan_tipo=plan_tipo,
    )
    await db.pagos_suscripcion.insert_one(pago.dict())
    suscripcion = await db.suscripciones.find_one({"empresa_id": empresa_id})
    now = datetime.now(timezone.utc)
    if suscripcion:
        base, nueva_fecha = _aplicar_renovacion(suscripcion, meses, now)
        await db.suscripciones.update_one(
            {"empresa_id": empresa_id},
            {"$set": {
                "status": SuscripcionStatus.ACTIVA,
                "fecha_vencimiento": nueva_fecha,
                "plan_tipo": plan_tipo,
            }}
        )
        await db.pagos_suscripcion.update_one(
            {"id": pago.id},
            {"$set": {"periodo_inicio": base, "periodo_fin": nueva_fecha}}
        )
    else:
        dia_facturacion = min(now.day, 28)
        nueva_fecha = calcular_siguiente_vencimiento(dia_facturacion, meses, now)
        nueva_sus = Suscripcion(
            empresa_id=empresa_id,
            plan_nombre=await get_plan_nombre_suscripcion(),
            precio=data.monto,
            status=SuscripcionStatus.ACTIVA,
            fecha_inicio=now,
            fecha_vencimiento=nueva_fecha,
            dia_facturacion=dia_facturacion,
            plan_tipo=plan_tipo,
        )
        await db.suscripciones.insert_one(nueva_sus.dict())
    return {"message": "Pago registrado y suscripción renovada"}

@owner_router.post("/clientes/{empresa_id}/suscripcion/cancelar-preapproval")
async def owner_cancelar_preapproval(empresa_id: str, _=Depends(verify_owner_token)):
    """Cancela el débito automático de un cliente desde el panel owner."""
    suscripcion = await db.suscripciones.find_one({"empresa_id": empresa_id})
    if not suscripcion or not suscripcion.get("mp_preapproval_id"):
        raise HTTPException(status_code=404, detail="Este cliente no tiene débito automático activo.")

    preapproval_id = suscripcion["mp_preapproval_id"]
    try:
        async with httpx.AsyncClient() as client_http:
            resp = await client_http.put(
                f"https://api.mercadopago.com/preapproval/{preapproval_id}",
                json={"status": "cancelled"},
                headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"},
                timeout=15,
            )
        if resp.status_code not in (200, 201, 404):
            raise HTTPException(status_code=502, detail=f"Error MP ({resp.status_code}): No se pudo cancelar el preapproval")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Error al conectar con MercadoPago: {str(e)}")

    await db.suscripciones.update_one(
        {"empresa_id": empresa_id},
        {"$set": {"tipo_cobro": "manual", "mp_preapproval_id": None}},
    )
    return {"ok": True, "mensaje": "Débito automático cancelado correctamente."}

@owner_router.put("/clientes/{empresa_id}/modulos")
async def owner_update_modulos(
    empresa_id: str, data: ModulosUpdate, _=Depends(verify_owner_token)
):
    """Reemplaza los overrides de módulos de un cliente (extra y removidos)."""
    emp = await db.empresas.find_one({"id": empresa_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    update = {}
    if data.modules_extra is not None:
        invalid = [m for m in data.modules_extra if m not in MODULES]
        if invalid:
            raise HTTPException(status_code=400, detail=f"Módulos inválidos: {invalid}")
        update["modules_extra"] = data.modules_extra
    if data.modules_removidos is not None:
        invalid = [m for m in data.modules_removidos if m not in MODULES]
        if invalid:
            raise HTTPException(status_code=400, detail=f"Módulos inválidos: {invalid}")
        update["modules_removidos"] = data.modules_removidos
    if update:
        suscripcion = await db.suscripciones.find_one({"empresa_id": empresa_id})
        if suscripcion:
            await db.suscripciones.update_one({"empresa_id": empresa_id}, {"$set": update})
        else:
            raise HTTPException(status_code=404, detail="La empresa no tiene suscripción aún")
    return {"ok": True, "mensaje": "Módulos actualizados"}


@owner_router.get("/modulos")
async def owner_get_modulos_catalogo(_=Depends(verify_owner_token)):
    """Devuelve el catálogo completo de módulos y los módulos por plan."""
    return {
        "modules": MODULES,
        "plan_modules": PLAN_MODULES,
    }


@owner_router.delete("/clientes/{empresa_id}")
async def owner_delete_cliente(empresa_id: str, _=Depends(verify_owner_token)):
    empresa = await db.empresas.find_one({"id": empresa_id})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    collections = [
        db.products, db.branch_products, db.categories, db.branches,
        db.users, db.configuration, db.suscripciones, db.pagos_suscripcion,
        db.notificaciones, db.sales, db.sale_returns, db.cash_sessions,
        db.cash_movements, db.compras, db.proveedores, db.afip_config,
    ]
    for col in collections:
        await col.delete_many({"empresa_id": empresa_id})
    await db.empresas.delete_one({"id": empresa_id})
    return {"ok": True}

@owner_router.put("/clientes/{empresa_id}/activo")
async def owner_toggle_empresa(empresa_id: str, _=Depends(verify_owner_token)):
    emp = await db.empresas.find_one({"id": empresa_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    new_status = not emp.get("activo", True)
    await db.empresas.update_one({"id": empresa_id}, {"$set": {"activo": new_status}})
    if not new_status:
        # Guardar el status actual antes de suspender para poder restaurarlo
        sus = await db.suscripciones.find_one({"empresa_id": empresa_id})
        prev_status = sus.get("status") if sus else None
        await db.suscripciones.update_one(
            {"empresa_id": empresa_id},
            {"$set": {"status": SuscripcionStatus.SUSPENDIDA, "status_antes_suspension": prev_status}}
        )
    else:
        # Al re-activar: restaurar el status anterior si no venció
        sus = await db.suscripciones.find_one({"empresa_id": empresa_id})
        if sus and sus.get("status") == SuscripcionStatus.SUSPENDIDA:
            vencimiento = sus.get("fecha_vencimiento")
            if isinstance(vencimiento, str):
                vencimiento = datetime.fromisoformat(vencimiento)
            if vencimiento and not vencimiento.tzinfo:
                vencimiento = vencimiento.replace(tzinfo=timezone.utc)
            now = datetime.now(timezone.utc)
            if vencimiento and vencimiento > now:
                status_previo = sus.get("status_antes_suspension")
                nuevo_status = status_previo if status_previo in (SuscripcionStatus.TRIAL, SuscripcionStatus.ACTIVA) else SuscripcionStatus.ACTIVA
            else:
                nuevo_status = SuscripcionStatus.VENCIDA
            await db.suscripciones.update_one(
                {"empresa_id": empresa_id},
                {"$set": {"status": nuevo_status}}
            )
    return {"activo": new_status, "message": f"Empresa {'activada' if new_status else 'suspendida'}"}

@owner_router.put("/clientes/{empresa_id}/datos")
async def owner_update_cliente_datos(empresa_id: str, data: ClienteDatosUpdate, _=Depends(verify_owner_token)):
    emp = await db.empresas.find_one({"id": empresa_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    if data.empresa_nombre is not None:
        nombre = data.empresa_nombre.strip()
        if not nombre:
            raise HTTPException(status_code=400, detail="El nombre de la empresa no puede estar vacío")
        await db.empresas.update_one({"id": empresa_id}, {"$set": {"nombre": nombre}})
    if data.admin_nombre is not None or data.admin_email is not None:
        admin = await db.users.find_one({"empresa_id": empresa_id, "rol": "admin"})
        if not admin:
            raise HTTPException(status_code=404, detail="Administrador no encontrado")
        admin_update = {}
        if data.admin_nombre is not None:
            nombre_admin = data.admin_nombre.strip()
            if not nombre_admin:
                raise HTTPException(status_code=400, detail="El nombre del administrador no puede estar vacío")
            admin_update["nombre"] = nombre_admin
        if data.admin_email is not None:
            email = data.admin_email.strip().lower()
            if not email:
                raise HTTPException(status_code=400, detail="El email no puede estar vacío")
            existing = await db.users.find_one({"email": email, "id": {"$ne": admin["id"]}})
            if existing:
                raise HTTPException(status_code=400, detail="El email ya está en uso por otro usuario")
            admin_update["email"] = email
        if admin_update:
            await db.users.update_one({"id": admin["id"]}, {"$set": admin_update})
    return {"message": "Datos actualizados correctamente"}

@owner_router.get("/clientes/{empresa_id}/config")
async def owner_get_empresa_config(empresa_id: str, _=Depends(verify_owner_token)):
    empresa = await db.empresas.find_one({"id": empresa_id})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    config = await db.configuration.find_one({"empresa_id": empresa_id})
    if not config:
        default_config = Configuration(empresa_id=empresa_id)
        await db.configuration.insert_one(default_config.dict())
        return default_config.dict()
    config.pop("_id", None)
    return config

@owner_router.put("/clientes/{empresa_id}/config")
async def owner_update_empresa_config(empresa_id: str, config_data: ConfigurationUpdate, _=Depends(verify_owner_token)):
    empresa = await db.empresas.find_one({"id": empresa_id})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    current_config = await db.configuration.find_one({"empresa_id": empresa_id})
    if not current_config:
        default_config = Configuration(empresa_id=empresa_id).dict()
        await db.configuration.insert_one(default_config)
    update_data = {k: v for k, v in config_data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    if update_data:
        await db.configuration.update_one({"empresa_id": empresa_id}, {"$set": update_data})
    updated = await db.configuration.find_one({"empresa_id": empresa_id})
    updated.pop("_id", None)
    return updated

@owner_router.get("/config")
async def owner_get_config(_=Depends(verify_owner_token)):
    smtp = await get_smtp_config()
    return {
        "suscripcion_precio": await get_precio_suscripcion(),
        "suscripcion_plan_nombre": await get_plan_nombre_suscripcion(),
        "precio_emprendedor": await get_precio_emprendedor(),
        "precio_profesional": await get_precio_profesional(),
        "precio_empresarial": await get_precio_empresarial(),
        "whatsapp_numero": await get_whatsapp_numero(),
        "trial_dias": await get_trial_dias(),
        "grace_days": await get_grace_days(),
        "dias_alerta": await get_dias_alerta(),
        "descuento_anual_pct": await get_descuento_anual_pct(),
        "saas_nombre": await get_saas_nombre(),
        "statement_descriptor": await get_statement_descriptor(),
        "smtp_host": smtp.get("host", SMTP_HOST),
        "smtp_port": smtp.get("port", SMTP_PORT),
        "smtp_user": smtp.get("user", SMTP_USER),
        "smtp_password": smtp.get("password", SMTP_PASSWORD),
        "smtp_from": smtp.get("from_address", EMAIL_FROM),
    }

class OwnerConfigUpdate(BaseModel):
    suscripcion_precio: Optional[float] = None
    suscripcion_plan_nombre: Optional[str] = None
    precio_emprendedor: Optional[float] = None
    precio_profesional: Optional[float] = None
    precio_empresarial: Optional[float] = None
    precio_tienda: Optional[float] = None
    precio_sucursal_extra: Optional[float] = None
    precio_pack_usuarios: Optional[float] = None
    whatsapp_numero: Optional[str] = None
    trial_dias: Optional[int] = None
    grace_days: Optional[int] = None
    dias_alerta: Optional[list] = None
    descuento_anual_pct: Optional[int] = None
    saas_nombre: Optional[str] = None
    statement_descriptor: Optional[str] = None
    smtp_host: Optional[str] = None
    smtp_port: Optional[int] = None
    smtp_user: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_from: Optional[str] = None

@owner_router.put("/config")
async def owner_update_config(data: OwnerConfigUpdate, _=Depends(verify_owner_token)):
    simple_keys = [
        ("suscripcion_precio", data.suscripcion_precio),
        ("suscripcion_plan_nombre", data.suscripcion_plan_nombre),
        ("precio_emprendedor", data.precio_emprendedor),
        ("precio_profesional", data.precio_profesional),
        ("precio_empresarial", data.precio_empresarial),
        ("precio_tienda", data.precio_tienda),
        ("precio_sucursal_extra", data.precio_sucursal_extra),
        ("precio_pack_usuarios", data.precio_pack_usuarios),
        ("whatsapp_numero", data.whatsapp_numero),
        ("trial_dias", data.trial_dias),
        ("grace_days", data.grace_days),
        ("dias_alerta", data.dias_alerta),
        ("descuento_anual_pct", data.descuento_anual_pct),
        ("saas_nombre", data.saas_nombre),
        ("statement_descriptor", data.statement_descriptor),
    ]
    for key, value in simple_keys:
        if value is not None:
            await db.system_config.update_one(
                {"key": key},
                {"$set": {"value": value}},
                upsert=True,
            )
    smtp_fields = {
        k: v for k, v in {
            "host": data.smtp_host,
            "port": data.smtp_port,
            "user": data.smtp_user,
            "password": data.smtp_password,
            "from_address": data.smtp_from,
        }.items() if v is not None
    }
    if smtp_fields:
        existing = await get_smtp_config()
        merged = {**existing, **smtp_fields}
        await db.system_config.update_one(
            {"key": "smtp_config"},
            {"$set": {"value": merged}},
            upsert=True,
        )
    return {"ok": True}

@owner_router.get("/alertas")
async def owner_get_alertas(_=Depends(verify_owner_token)):
    """Empresas con suscripciones próximas a vencer (≤ 10 días) y últimas alertas generadas."""
    now = datetime.now(timezone.utc)
    suscripciones = await db.suscripciones.find(
        {"status": {"$in": [SuscripcionStatus.TRIAL, SuscripcionStatus.ACTIVA]}}
    ).to_list(None)
    por_vencer = []
    for sus in suscripciones:
        vencimiento = sus.get("fecha_vencimiento")
        if isinstance(vencimiento, str):
            vencimiento = datetime.fromisoformat(vencimiento)
        if vencimiento and not vencimiento.tzinfo:
            vencimiento = vencimiento.replace(tzinfo=timezone.utc)
        if not vencimiento:
            continue
        dias_restantes = (vencimiento - now).days
        if dias_restantes <= 10:
            empresa = await db.empresas.find_one({"id": sus["empresa_id"]})
            por_vencer.append({
                "empresa_id": sus["empresa_id"],
                "empresa_nombre": empresa["nombre"] if empresa else sus["empresa_id"],
                "status": sus.get("status"),
                "plan_nombre": sus.get("plan_nombre"),
                "fecha_vencimiento": sus.get("fecha_vencimiento"),
                "dias_restantes": max(0, dias_restantes),
            })
    por_vencer.sort(key=lambda x: x["dias_restantes"])
    alertas_recientes = await db.notificaciones.find().sort("fecha", -1).limit(50).to_list(50)
    return {
        "por_vencer": por_vencer,
        "alertas_recientes": [Notificacion(**a).dict() for a in alertas_recientes],
    }

@owner_router.post("/alertas/generar")
async def owner_generar_alertas(_=Depends(verify_owner_token)):
    """Genera manualmente alertas para suscripciones próximas a vencer."""
    generadas = await generar_alertas_vencimiento()
    return {"ok": True, "generadas": generadas}

@owner_router.get("/pagos")
async def owner_get_pagos(_=Depends(verify_owner_token)):
    """Todos los pagos de suscripción ordenados por fecha descendente."""
    pagos = await db.pagos_suscripcion.find({}).sort("fecha", -1).to_list(1000)
    empresas_cache: dict = {}
    result = []
    for p in pagos:
        eid = p.get("empresa_id", "")
        if eid not in empresas_cache:
            emp = await db.empresas.find_one({"id": eid})
            empresas_cache[eid] = emp["nombre"] if emp else eid
        pago_dict = {k: v for k, v in p.items() if k != "_id"}
        pago_dict["empresa_nombre"] = empresas_cache[eid]
        result.append(pago_dict)
    return result

# ─── Router AFIP/ARCA ─────────────────────────────────────────────────────────

afip_router = APIRouter(prefix="/api/afip")
_afip_service = AfipService()

@afip_router.get("/config")
async def get_afip_config(user: User = Depends(require_role([UserRole.ADMIN]))):
    """Retorna la configuración AFIP de la empresa (sin exponer la clave privada)."""
    cfg = await db.afip_config.find_one({"empresa_id": user.empresa_id, "activo": True})
    if not cfg:
        return {"configurado": False}
    return {
        "configurado": True,
        "cuit": cfg.get("cuit"),
        "punto_venta": cfg.get("punto_venta"),
        "ambiente": cfg.get("ambiente"),
        "tipo_comprobante_default": cfg.get("tipo_comprobante_default"),
        "razon_social": cfg.get("razon_social"),
        "tiene_certificado": bool(cfg.get("cert_pem")),
        "tiene_clave": bool(cfg.get("key_pem_encrypted")),
    }

@afip_router.post("/config")
async def save_afip_config(data: AfipConfigCreate, user: User = Depends(require_role([UserRole.ADMIN]))):
    """Crea o actualiza la configuración AFIP (sin tocar cert/key)."""
    now = datetime.now(timezone.utc)
    existing = await db.afip_config.find_one({"empresa_id": user.empresa_id})
    if existing:
        await db.afip_config.update_one(
            {"empresa_id": user.empresa_id},
            {"$set": {
                "cuit": data.cuit,
                "punto_venta": data.punto_venta,
                "ambiente": data.ambiente,
                "tipo_comprobante_default": data.tipo_comprobante_default,
                "razon_social": data.razon_social or "",
                "domicilio_fiscal": data.domicilio_fiscal,
                "condicion_iva_emisor": data.condicion_iva_emisor or "RI",
                "inicio_actividades": data.inicio_actividades,
                "iibb": data.iibb,
                "concepto_default": data.concepto_default or 1,
                "activo": True,
                "updated_at": now,
            }}
        )
    else:
        cfg = AfipConfig(**data.dict(), empresa_id=user.empresa_id)
        await db.afip_config.insert_one(cfg.dict())
    return {"ok": True, "mensaje": "Configuración AFIP guardada"}

@afip_router.post("/config/upload-cert")
async def upload_afip_cert(
    cert_file: Optional[UploadFile] = File(None),
    key_file: Optional[UploadFile] = File(None),
    p12_file: Optional[UploadFile] = File(None),
    p12_password: Optional[str] = None,
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    """
    Sube el certificado y/o clave privada de AFIP.
    Acepta:
      - p12_file: archivo .p12 (extrae cert + key automáticamente)
      - cert_file + key_file: archivos .pem separados
    """
    cfg = await db.afip_config.find_one({"empresa_id": user.empresa_id})
    if not cfg:
        raise HTTPException(status_code=400, detail="Primero configure los datos básicos de AFIP (CUIT, punto de venta).")

    update = {"updated_at": datetime.now(timezone.utc)}

    if p12_file:
        p12_bytes = await p12_file.read()
        pwd = p12_password.encode() if p12_password else None
        try:
            cert_pem, key_pem = extract_p12(p12_bytes, pwd)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo .p12: {e}")
        update["cert_pem"] = base64.b64encode(cert_pem).decode()
        update["key_pem_encrypted"] = encrypt_private_key(key_pem, SECRET_KEY)

    if cert_file:
        cert_bytes = await cert_file.read()
        update["cert_pem"] = base64.b64encode(cert_bytes).decode()

    if key_file:
        key_bytes = await key_file.read()
        update["key_pem_encrypted"] = encrypt_private_key(key_bytes, SECRET_KEY)

    await db.afip_config.update_one({"empresa_id": user.empresa_id}, {"$set": update})
    return {"ok": True, "mensaje": "Certificado cargado correctamente"}

@afip_router.post("/test")
async def test_afip_conexion(user: User = Depends(require_role([UserRole.ADMIN]))):
    """Llama FEDummy para verificar conectividad con los servidores de ARCA."""
    cfg = await db.afip_config.find_one({"empresa_id": user.empresa_id, "activo": True})
    if not cfg:
        raise HTTPException(status_code=400, detail="AFIP no está configurado para esta empresa.")
    result = await _afip_service.test_conexion(cfg, SECRET_KEY, db)
    if not result["ok"]:
        raise HTTPException(status_code=503, detail=f"Error de conexión con ARCA: {result.get('error')}")
    return result

@afip_router.post("/reintentar/{sale_id}")
async def reintentar_cae(sale_id: str, user: User = Depends(require_role([UserRole.ADMIN]))):
    """Reintenta obtener el CAE para una venta en estado contingencia o error."""
    sale_doc = await db.sales.find_one({"id": sale_id, "empresa_id": user.empresa_id})
    if not sale_doc:
        raise HTTPException(status_code=404, detail="Venta no encontrada.")
    if sale_doc.get("afip_estado") == "autorizado":
        return {"ok": True, "mensaje": "La venta ya tiene CAE autorizado.", "cae": sale_doc.get("cae")}

    cfg = await db.afip_config.find_one({"empresa_id": user.empresa_id, "activo": True})
    if not cfg or not cfg.get("cert_pem") or not cfg.get("key_pem_encrypted"):
        raise HTTPException(status_code=400, detail="AFIP no está configurado o falta el certificado.")

    try:
        token, sign = await _afip_service.get_token_sign(cfg, db, SECRET_KEY)
        tipo_cbte = sale_doc.get("tipo_comprobante") or cfg.get("tipo_comprobante_default", 6)
        result = await _afip_service.solicitar_cae(
            sale_doc, cfg, token, sign, tipo_cbte,
            cuit_receptor=sale_doc.get("cuit_receptor")
        )
        await db.sales.update_one(
            {"id": sale_id},
            {"$set": {
                "cae": result["cae"],
                "cae_vencimiento": result["cae_vencimiento"],
                "afip_estado": "autorizado",
                "afip_error": None,
                "nro_comprobante_afip": result["nro_comprobante"],
            }}
        )
        return {"ok": True, "cae": result["cae"], "cae_vencimiento": result["cae_vencimiento"]}
    except Exception as e:
        await db.sales.update_one(
            {"id": sale_id},
            {"$set": {"afip_estado": "error", "afip_error": str(e)}}
        )
        raise HTTPException(status_code=503, detail=f"Error al obtener CAE: {e}")

@afip_router.get("/ventas-pendientes")
async def get_ventas_pendientes(user: User = Depends(require_role([UserRole.ADMIN]))):
    """Lista ventas con CAE pendiente (contingencia o error)."""
    ventas = await db.sales.find({
        "empresa_id": user.empresa_id,
        "afip_estado": {"$in": ["contingencia", "error"]}
    }).to_list(200)
    return [Sale(**v) for v in ventas]

@afip_router.post("/reintentar-nc/{credit_note_id}")
async def reintentar_cae_nc(
    credit_note_id: str,
    body: RetryNcRequest = Body(default=RetryNcRequest()),
    user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Reintenta obtener el CAE para una nota de crédito en contingencia."""
    nc_doc = await db.credit_notes.find_one({"id": credit_note_id, "empresa_id": user.empresa_id})
    if not nc_doc:
        raise HTTPException(status_code=404, detail="Nota de crédito no encontrada.")
    if nc_doc.get("afip_estado") == "autorizado":
        return {"ok": True, "mensaje": "La nota de crédito ya tiene CAE.", "cae": nc_doc.get("cae")}

    # Recuperar la venta original para obtener tipo_comprobante y nro_comprobante_afip
    sale_doc = await db.sales.find_one({"id": nc_doc["sale_id"], "empresa_id": user.empresa_id})
    if not sale_doc:
        raise HTTPException(status_code=404, detail="Venta original no encontrada.")
    if sale_doc.get("afip_estado") != "autorizado":
        raise HTTPException(status_code=400, detail="La venta original no tiene CAE autorizado.")

    NC_TIPO_MAP = {1: 3, 6: 8, 11: 13}
    nc_tipo = NC_TIPO_MAP.get(sale_doc.get("tipo_comprobante"))
    if not nc_tipo:
        raise HTTPException(status_code=400, detail="La venta original no es una factura electrónica.")

    cfg = await db.afip_config.find_one({"empresa_id": user.empresa_id, "activo": True})
    if not cfg or not cfg.get("cert_pem") or not cfg.get("key_pem_encrypted"):
        raise HTTPException(status_code=400, detail="AFIP no está configurado o falta el certificado.")

    try:
        token_a, sign_a = await _afip_service.get_token_sign(cfg, db, SECRET_KEY)
        total_return = nc_doc["total"]
        tax_rate = cfg.get("tax_rate", 0.21) if nc_tipo != 13 else 0.0
        if tax_rate > 0:
            nc_subtotal = round(total_return / (1 + tax_rate), 2)
            nc_impuestos = round(total_return - nc_subtotal, 2)
        else:
            nc_subtotal = round(total_return, 2)
            nc_impuestos = 0.0
        nc_sale_dict = {
            "total": round(total_return, 2),
            "subtotal": nc_subtotal,
            "impuestos": nc_impuestos,
            "impuestos_extra_total": 0.0,
        }
        cbtes_asoc = [{
            "Tipo": sale_doc["tipo_comprobante"],
            "PtoVta": cfg["punto_venta"],
            "Nro": sale_doc["nro_comprobante_afip"],
        }]
        effective_cuit = body.cuit_receptor or sale_doc.get("cuit_receptor")
        result = await _afip_service.solicitar_cae(
            nc_sale_dict, cfg, token_a, sign_a, nc_tipo,
            cuit_receptor=effective_cuit,
            cbtes_asoc=cbtes_asoc,
        )
        await db.credit_notes.update_one(
            {"id": credit_note_id},
            {"$set": {
                "cae": result["cae"],
                "cae_vencimiento": result["cae_vencimiento"],
                "afip_estado": "autorizado",
                "tipo_comprobante_nc": nc_tipo,
                "nro_comprobante_afip": result["nro_comprobante"],
                "afip_error": None,
            }}
        )
        return {"ok": True, "cae": result["cae"], "cae_vencimiento": result["cae_vencimiento"]}
    except Exception as e:
        await db.credit_notes.update_one(
            {"id": credit_note_id},
            {"$set": {
                "afip_estado": "contingencia",
                "afip_error": str(e),
                "tipo_comprobante_nc": nc_tipo,
            }}
        )
        raise HTTPException(status_code=503, detail=f"Error al obtener CAE para NC: {e}")

# Include the router in the main app
app.include_router(api_router)
app.include_router(owner_router)
app.include_router(afip_router)

origins = os.environ.get("CORS_ORIGINS", "")
allow_origins = [o.strip() for o in origins.split(",") if o.strip()]
is_wildcard = allow_origins == ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_origin_regex=r".*" if is_wildcard else None,
    allow_credentials=not is_wildcard,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_tasks():
    async def periodic_alerts():
        while True:
            try:
                await generar_alertas_vencimiento()
            except Exception as e:
                logger.error(f"Error generando alertas: {e}")
            await asyncio.sleep(24 * 60 * 60)  # cada 24 horas
    asyncio.create_task(periodic_alerts())

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
